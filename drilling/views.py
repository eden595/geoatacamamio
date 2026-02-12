from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from openpyxl import Workbook
import os
from django.conf import settings
from xhtml2pdf import pisa
from django.http import JsonResponse
from django.template.loader import get_template, render_to_string
from django.contrib import messages
from django.utils.datastructures import MultiValueDictKeyError
import datetime
from datetime import timedelta
from django.db.models import F, Q, Max, Value, CharField
from django.utils.timezone import localtime
from django.utils import timezone
import json
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ObjectDoesNotExist,MultipleObjectsReturned
from django.views.decorators.http import require_GET
from core.models import DetalleControlHorario, Diametros, TipoTerreno, Orientacion, Sondajes, Aditivos, Sondas
from .forms import FormReportesOperacionales, FormDetallesPerforaciones, FormControlesHorarios, FormInsumos, FormAditivos, FormLongitudPozos, FormObservacionesReportes, ExportDataForm
from .models import DetallesPerforaciones, ReportesOperacionales, LongitudPozos, DetalleAditivos, ControlesHorarios, Insumos, ObservacionesReportes
from core.decorators import sondaje_admin_or_controlador_or_base_datos_or_supervisor_required, sondaje_admin_or_base_datos_or_supervisor_required
from user.models import UsuarioProfile
from django.db import transaction
from checklist.forms import ChecklistMaterialesSondaEntradaForm, ChecklistMaterialesSondaSalidaForm, ChecklistMaterialesCasetaFormTop, ChecklistMaterialesCasetaFormBottom
from core.forms import FormMaterialesCaseta, FormMaterialesSonda
from core.models import MaterialesCaseta, MaterialesSonda
from checklist.models import ChecklistMaterialesSonda, EstadoEtapasReporteDigital
from core.utils import normalizar_decimal, normalizar_entero, sumar_totales_control_horario
from django.core.paginator import Paginator
from django.middleware.csrf import get_token
from django.db.models.functions import Concat, Cast
import threading
from core.dashboard.generate_csv_v2 import GenerateCsv
from core.choices import gemelo
@login_required
@sondaje_admin_or_controlador_or_base_datos_or_supervisor_required
def new_reporte_digital(request):
    usuario = UsuarioProfile.objects.get(user=request.user)
    if usuario.faena.faena == 'SIN ASIGNAR' and request.user.role not in ['ADMINISTRADOR', 'SUPERVISOR', 'BASE DATOS']:
        return redirect('edit_my_profile')
    else:
        control_horario = DetalleControlHorario.objects.filter(status=True).order_by('detalle')
        diametros = Diametros.objects.filter(status=True).order_by('diametro')
        tipo_terreno = TipoTerreno.objects.filter(status=True).order_by('tipoTerreno')
        orientacion = Orientacion.objects.filter(status=True).order_by('orientacion')
        aditivos = Aditivos.objects.filter(status=True).order_by('aditivo')
        usuario_profile = UsuarioProfile.objects.get(user=request.user)
        try:
            sondaje_codigo = Sondajes.objects.get(faena=usuario_profile.faena, status=True)
            disabled_sondaje_codigo=True
        except Sondajes.DoesNotExist:
            sondaje_codigo = Sondajes.objects.all()
            disabled_sondaje_codigo=False
            
        formEntrada = ChecklistMaterialesSondaEntradaForm()
        formSalida = ChecklistMaterialesSondaSalidaForm()
        materialesSonda = MaterialesSonda.objects.filter(status=True)
        
        context = {
            'formInicio': formEntrada,
            'formFinal': formSalida,
            'materiales': materialesSonda,
            'detalles_control_horario': control_horario,
            'diametros': diametros,
            'tipos_terreno': tipo_terreno,
            'orientaciones': orientacion,
            'aditivos': aditivos,
            'formreporteoperacional': FormReportesOperacionales(initial={
                'controlador': request.user.first_name+" "+request.user.last_name,
                'sondajeCodigo': sondaje_codigo,
            },
            disabled_sondaje_codigo=disabled_sondaje_codigo,
            ),  
            'formdetalleperforacion': FormDetallesPerforaciones,
            'formcontrolhorario': FormControlesHorarios,
            'forminsumos': FormInsumos,
            'formaditivos': FormAditivos,
            'formlongitudpozo': FormLongitudPozos,
            'formobservaciones': FormObservacionesReportes,
            'sidebar': 'manage_new_report',
            'sidebarmain': 'system_drilling', 
        }
        return render(request, 'pages/drilling/new_reporte_digital.html', context)

@login_required
@sondaje_admin_or_controlador_or_base_datos_or_supervisor_required
def save_reporte_digital(request):
    if request.method == 'POST':

        total_horas = sumar_totales_control_horario(request.POST)
        if total_horas != timedelta(hours=12):
            horas = int(total_horas.total_seconds() // 3600)
            minutos = int((total_horas.total_seconds() % 3600) // 60)
            mensaje = f'Total de Horas Control Horario: {horas:02}:{minutos:02}'
            return JsonResponse(
                {
                    'titleText': mensaje,
                    'text': 'El total de horas en Control Horario debe ser exactamente 12:00 hrs.'
                },
                status=400
                )
        reporte = ReportesOperacionales.objects.filter(sondajeCodigo=request.POST['sondajeCodigo'], sondajeSerie=request.POST['sondajeSerie'], sondajeEstado=request.POST['sondajeEstado'], status=True).exclude(progreso='Eliminado').order_by('-id').first()
        if reporte == None:
            correlativo = 1
        else:
            correlativo = reporte.correlativo + 1
        try:
            # Iniciar la transacción atómica
            with transaction.atomic():
                # Crear Reporte Operacional
                try:
                    reporte_operacional = ReportesOperacionales.objects.create(
                        turno=request.POST['turno'],
                        perforista_id=request.POST['perforista'],
                        sonda_id=request.POST['sonda'],
                        sondajeCodigo_id=request.POST['sondajeCodigo'],
                        sondajeSerie=request.POST['sondajeSerie'],
                        sondajeEstado=request.POST['sondajeEstado'],
                        metroInicial=normalizar_decimal(request.POST['metroInicial']),
                        metroFinal=normalizar_decimal(request.POST['metroFinal']),
                        totalPerforado=normalizar_decimal(request.POST['totalPerforado']),
                        controlador=request.user, 
                        creador=request.POST['controlador'],
                        progreso='Por Revisar',
                        correlativo=correlativo,
                        id_checklist=request.POST['id'],
                        status=True,
                    )
                    print(f"Reporte creado con id_checklist {reporte_operacional.id_checklist}")
                except Exception as e:
                    raise Exception("Error en Sección:\n Datos Principales")

                # Crear DetallesPerforaciones dinámicamente
                try:
                    perforacion_count = len([key for key in request.POST.keys() if key.startswith('diametros_')])
                    for i in range(1, perforacion_count + 1):
                        DetallesPerforaciones.objects.create(
                            reporte=reporte_operacional,
                            diametros_id=request.POST.get(f'diametros_{i}'),
                            perforado=normalizar_decimal(request.POST.get(f'perforado_{i}')),
                            desde=normalizar_decimal(request.POST.get(f'desde_{i}')),
                            hasta=normalizar_decimal(request.POST.get(f'hasta_{i}')),
                            recuperacion=normalizar_decimal(request.POST.get(f'recuperacion_{i}')),
                            porcentajeRecuperacion=normalizar_decimal(request.POST.get(f'porcentajeRecuperacion_{i}')),
                            barra=int(request.POST.get(f'barra_{i}')) if request.POST.get(f'barra_{i}') and request.POST.get(f'barra_{i}').isdigit() else 0,
                            largoBarra=normalizar_decimal(request.POST.get(f'largoBarra_{i}')),
                            totalHtas=normalizar_decimal(request.POST.get(f'totalHtas_{i}')),
                            contra=normalizar_decimal(request.POST.get(f'contra_{i}')),
                            tipoTerreno_id=request.POST.get(f'tipoTerreno_{i}'),
                            orientacion_id=request.POST.get(f'orientacion_{i}'),
                            status=True,
                        )
                except Exception as e:
                    raise Exception("Error en Sección:\n Detalle Perforación")

                # Crear ControlesHorarios dinámicamente
                try:
                    control_horario_count = len([key for key in request.POST.keys() if key.startswith('inicio_')])
                    for i in range(1, control_horario_count + 1):
                        inicio_time = datetime.datetime.strptime(request.POST.get(f'inicio_{i}'), '%H:%M').time()
                        final_time = datetime.datetime.strptime(request.POST.get(f'final_{i}'), '%H:%M').time()
                        total_time = datetime.datetime.strptime(request.POST.get(f'total_{i}'), '%H:%M').time()
                        detalle_id = request.POST.get(f'detalleControlHorario_{i}')
                        if not detalle_id:
                            raise Exception("Debe seleccionar un Detalle para todos los registros de Control Horario.")

                        ControlesHorarios.objects.create(
                            reporte=reporte_operacional,
                            posicion=i, 
                            inicio=inicio_time,
                            final=final_time,
                            total=total_time,
                            detalleControlHorario_id=detalle_id,
                            status=True,
                        )
                except Exception as e:
                    if str(e) == "Debe seleccionar un Detalle para todos los registros de Control Horario.":
                        raise e
                    raise Exception("Error en Sección:\n Control Horario")

                # Crear Insumos
                try:
                    Insumos.objects.create(
                        reporte=reporte_operacional,
                        corona=request.POST['corona'],
                        escareador=request.POST['escareador'],
                        cantidadAgua_id=request.POST['cantidadAgua'],
                        casing=normalizar_decimal(request.POST['casing']),
                        zapata=request.POST['zapata'],
                        status=True,
                    )
                except Exception as e:
                    raise Exception("Error en Sección:\n Control de Insumos")

                # Crear DetalleAditivos dinámicamente
                try:
                    aditivo_count = len([key for key in request.POST.keys() if key.startswith('aditivo_')])
                    for i in range(1, aditivo_count + 1):
                        DetalleAditivos.objects.create(
                            reporte=reporte_operacional,
                            aditivo_id=request.POST.get(f'aditivo_{i}'),
                            cantidad=normalizar_entero(request.POST.get(f'cantidad_{i}')),    
                            status=True,
                        )
                except Exception as e:
                    raise Exception("Error en Sección:\n Aditivos")

                # Crear LongitudPozos
                try:
                    LongitudPozos.objects.create(
                        reporte=reporte_operacional,
                        largoBarril=normalizar_decimal(request.POST['largoBarril']),
                        #largoBarra=request.POST['largoBarra'],
                        puntoMuerto=normalizar_decimal(request.POST['puntoMuerto']),
                        restoBarra=normalizar_decimal(request.POST['restoBarra']),
                        numeroBarras=request.POST['numeroBarras'],
                        longitudPozo=normalizar_decimal(request.POST['longitudPozo']),
                        htaEnPozo=request.POST['htaEnPozo'],
                        mtsDeHta=normalizar_decimal(request.POST['mtsDeHta']),
                        profundidadHta=normalizar_decimal(request.POST['profundidadHta']),
                        status=True,
                    )
                except Exception as e:
                    raise Exception("Error en Sección:\n Longitud de Pozo")

                # Crear ObservacionesReportes
                try:
                    ObservacionesReportes.objects.create(
                        reporte=reporte_operacional,
                        observaciones=request.POST['observaciones'],
                        status=True,
                    )
                except Exception as e:
                    raise Exception("Error en Sección:\n Observaciones")
                
                EstadoEtapasReporteDigital.objects.update_or_create(
                    id_checklist=request.POST['id'],
                    defaults={
                        'reporteDigital': True,
                    }
                )
                
                # Si todo es exitoso
                return redirect('manage_mis_reportes_digitales')

        except Exception as e:
            # Muestra un mensaje con el error específico
            return JsonResponse({'titleText': str(e), 'text': "Revise los datos e intente nuevamente"}, status=400)

    else:
        return redirect('new_reporte_digital')

@require_GET
@sondaje_admin_or_controlador_or_base_datos_or_supervisor_required
def obtener_metro_inicial(request):
    sondaje_codigo = request.GET.get('sondajeCodigo')
    sondaje_serie = request.GET.get('sondajeSerie')
    sondaje_estado = request.GET.get('sondajeEstado')
    id_checklist = request.GET.get('id_checklist')
    
    print(id_checklist)
    """ reporteSonda = ChecklistMaterialesSonda.objects.filter(id_checklist=id_checklist).order_by('id').first()
        if not reporteSonda:
        print("Checklist no encontrado para id_checklist:", id_checklist)
        return JsonResponse({'error': 'Checklist no encontrado'}, status=404)
    print("progreso:", reporteSonda.progreso)"""

    try:
        sondajeCodigoObj = Sondajes.objects.get(sondaje=sondaje_codigo)
    except Sondajes.DoesNotExist:
        sondajeCodigoObj = None

    # --- LÓGICA DE DOBLE ORIGEN ---
    reporte_principal = None   # Para Metro Inicial e IDs (Datos Actuales)
    reporte_referencia = None  # Para Barra, Largo y Contra (Datos Anteriores)

    # 1. Buscamos si estamos editando el reporte actual
    reporte_actual = ReportesOperacionales.objects.filter(id_checklist=id_checklist, status=True).first()

    if reporte_actual:
        # MODO EDICIÓN:
        reporte_principal = reporte_actual
        
        # La referencia es el reporte HISTÓRICO (anterior en el tiempo)
        reporte_referencia = ReportesOperacionales.objects.filter(
            sondajeCodigo=reporte_actual.sondajeCodigo,
            sondajeSerie=reporte_actual.sondajeSerie,
            sondajeEstado=reporte_actual.sondajeEstado,
            status=True,
            fechacreacion__lt=reporte_actual.fechacreacion 
        ).exclude(progreso='Eliminado').order_by('-fechacreacion', '-id').first()
    else:
        # MODO CREACIÓN:
        # Buscamos el último reporte válido del pozo.
        ultimo_reporte = ReportesOperacionales.objects.filter(
            sondajeCodigo=sondajeCodigoObj, 
            sondajeSerie=sondaje_serie, 
            sondajeEstado=sondaje_estado, 
            status=True
        ).exclude(progreso='Eliminado').order_by('-fechacreacion', '-id').first()
        
        reporte_principal = ultimo_reporte
        reporte_referencia = ultimo_reporte

    # --- CONSTRUCCIÓN DE RESPUESTA ---
    if reporte_principal:
        # Datos del reporte principal
        reporte_data = {
            'id': reporte_principal.id,
            'turno': reporte_principal.turno,
            'fechacreacion': reporte_principal.fechacreacion,
            'controlador': str(reporte_principal.controlador), 
            'perforista': str(reporte_principal.perforista),  
            'sonda': str(reporte_principal.sonda),              
            'sondajeCodigo': reporte_principal.sondajeCodigo.id if reporte_principal.sondajeCodigo else None,  
            'sondajeSerie': reporte_principal.sondajeSerie,
            'sondajeEstado': reporte_principal.sondajeEstado,
            'metroInicial': reporte_principal.metroInicial,
            'metroFinal': reporte_principal.metroFinal,
            'totalPerforado': reporte_principal.totalPerforado,
            'creador': str(reporte_principal.creador), 
            'status': reporte_principal.status,
            'id_checklist': reporte_principal.id_checklist,
        }
        
        # Datos de referencia (Solo si existe un reporte anterior)
        perforacion_data = None
        if reporte_referencia:
            perforacion_data = DetallesPerforaciones.objects.filter(reporte=reporte_referencia).order_by('-id').values(
                'reporte', 'diametros', 'perforado', 'desde', 'hasta', 'recuperacion',
                'porcentajeRecuperacion', 'barra', 'largoBarra', 'totalHtas',
                'contra', 'tipoTerreno', 'orientacion', 'fechacreacion',
            ).first()

        insumos_data = Insumos.objects.filter(reporte=reporte_principal).order_by('-id').values(
            'reporte', 'corona', 'escareador', 'cantidadAgua', 'casing', 'zapata',
        ).first()
        
        longitudpozo_data = LongitudPozos.objects.filter(reporte=reporte_principal).order_by('-id').values(
            'reporte', 'largoBarril', 'largoBarra', 'puntoMuerto', 'restoBarra',
            'numeroBarras', 'longitudPozo',
        ).first()

        # AQUÍ ESTÁ LA CORRECCIÓN CLAVE:
        # 'existe_reporte_anterior' ahora depende de si encontramos datos de perforación previos.
        # Si es el primer reporte (perforacion_data es None), esto será False y el frontend mostrará "N/A".
        tiene_referencia = True if perforacion_data else False

        return JsonResponse({
            'reporte': reporte_data,
            # Enviamos vacíos en el fallback para no romper cálculos si el JS falla, pero el flag controlará la visualización
            'perforacion': perforacion_data if perforacion_data else { 'barra': '', 'largoBarra': 0, 'contra': 0 }, 
            'insumos': insumos_data,
            'longitudPozo': longitudpozo_data,
            'existe_reporte_anterior': tiene_referencia, 
        }, status=200)
    else:
        # Caso: Pozo nuevo absoluto
        return JsonResponse({
            'reporte': {
                'id': 0, 'turno': 0, 'fechacreacion': '', 'controlador': '', 'perforista': '', 'sonda': '', 
                'sondajeCodigo': '', 'sondajeSerie': '', 'sondajeEstado': '', 'metroInicial': 0, 'metroFinal': 0, 
                'totalPerforado': 0, 'creador': '', 'status': 0, 'id_checklist': 0,
            },
            'perforacion': {
                'reporte': 0, 'diametros': 0, 'perforado': 0, 'desde': 0, 'hasta': 0, 'recuperacion': 0, 
                'porcentajeRecuperacion': 0, 'barra': "", 'largoBarra': 0, 'totalHtas': 0, 'contra': 0, 
                'tipoTerreno': '', 'orientacion': '', 'fechacreacion': '',
            },
            'insumos': {
                'reporte': '', 'corona': '', 'escareador': '', 'cantidadAgua': 0, 'casing': '', 'zapata': '',
            },
            'longitudPozo': {
                'reporte': 0, 'largoBarril': 0, 'largoBarra': 0, 'puntoMuerto': 0, 'restoBarra': 0, 
                'numeroBarras': 0, 'longitudPozo': 0,
            },
            'existe_reporte_anterior': False,
        }, status=200)

@login_required
@sondaje_admin_or_controlador_or_base_datos_or_supervisor_required
def manage_mis_reportes_digitales(request): 
    usuario = UsuarioProfile.objects.get(user=request.user)
    """
    if usuario.faena.faena == 'SIN ASIGNAR' and request.user.role == 'ADMINISTRADOR':
        #lista = (ChecklistMaterialesSonda.objects.filter(Q(progreso='Creado') | Q(progreso='Por Corregir'),etapa="Entrada",status=True).values('id_checklist').annotate(max_fechacreacion=Max('fechacreacion')))
        lista = (ChecklistMaterialesSonda.objects.filter(progreso='Por Corregir',etapa="Entrada",status=True).values('id_checklist').annotate(max_fechacreacion=Max('fechacreacion')))
    else:
        #lista = (ChecklistMaterialesSonda.objects.filter(Q(progreso='Creado') | Q(progreso='Por Corregir'),etapa="Entrada",status=True,sondajeCodigo__faena=usuario.faena,).values('id_checklist').annotate(max_fechacreacion=Max('fechacreacion')))
        lista = (ChecklistMaterialesSonda.objects.filter(progreso='Por Corregir',etapa="Entrada",status=True,sondajeCodigo__faena=usuario.faena,).values('id_checklist').annotate(max_fechacreacion=Max('fechacreacion')))
    resultados = (ChecklistMaterialesSonda.objects.filter(id_checklist__in=[item['id_checklist'] for item in lista],fechacreacion__in=[item['max_fechacreacion'] for item in lista]).order_by('-fecha_checklist'))
    """

    if usuario.faena.faena == 'SIN ASIGNAR' and request.user.role in ['ADMINISTRADOR', 'SUPERVISOR', 'BASE DATOS']:
        lista = ChecklistMaterialesSonda.objects.filter(
            Q(progreso='Por Corregir'),
            etapa="Entrada",
            status=True
        ).values('id_checklist').annotate(max_fechacreacion=Max('fechacreacion'))
    else:
        lista = ChecklistMaterialesSonda.objects.filter(
            Q(progreso='Por Corregir'),
            etapa="Entrada",
            status=True,
            sondajeCodigo__faena=usuario.faena,
        ).values('id_checklist').annotate(max_fechacreacion=Max('fechacreacion'))

    # Generar filtros precisos por cada combinación id_checklist + max_fechacreacion
    resultados = []
    for item in lista:
        r = (ChecklistMaterialesSonda.objects
            .filter(id_checklist=item['id_checklist'], fechacreacion=item['max_fechacreacion'])
            .order_by('-fechacreacion')  # o por fecha_checklist si quieres
            .first())
        if r:
            resultados.append(r)
    
    # Obtener todos los registros de EstadoEtapasReporteDigital relacionados
    estados = EstadoEtapasReporteDigital.objects.filter(
        id_checklist__in=[r.id_checklist for r in resultados]
    )

    # Agrupar por id_checklist
    from collections import defaultdict

    estado_por_checklist = defaultdict(list)
    for estado in estados:
        estado_por_checklist[estado.id_checklist].append(estado)

    # Evaluar completado por cada resultado
    for r in resultados:
        estados_asociados = estado_por_checklist.get(r.id_checklist)

        if estados_asociados:
            # Si hay uno o más registros para este id_checklist, revisamos si TODOS tienen los campos en True
            completado = all(
                e.reporteDigital is True and
                e.checklistEntrada is True and
                e.checklistSalida is True
                for e in estados_asociados
            )
            r.completado = completado
        else:
            # Si no hay ningún registro, según tu lógica, es completado = True
            r.completado = True 
    
    resultados.sort(key=lambda x: x.fechacreacion, reverse=True)
    
    context = {
        'reportes': resultados,
        'sidebar': 'manage_my_digital_reports',
        'sidebarmain': 'system_drilling', 
    }
    return render(request, 'pages/drilling/manage_mis_reportes_digitales.html', context)

@login_required
@sondaje_admin_or_base_datos_or_supervisor_required
def manage_revisar_reportes_digitales(request):
    usuario = UsuarioProfile.objects.get(user=request.user)
    """
    if usuario.faena.faena == 'SIN ASIGNAR' and request.user.role == 'ADMINISTRADOR':
        lista = (ChecklistMaterialesSonda.objects.filter(Q(progreso='Corregido') | Q(progreso='Por Revisar'),etapa="Entrada",status=True).values('id_checklist').annotate(max_fechacreacion=Max('fechacreacion')))
    else:
        lista = (ChecklistMaterialesSonda.objects.filter(Q(progreso='Corregido') | Q(progreso='Por Revisar'),etapa="Entrada",status=True,sondajeCodigo__faena=usuario.faena,).values('id_checklist').annotate(max_fechacreacion=Max('fechacreacion')))
    resultados = (ChecklistMaterialesSonda.objects.filter(id_checklist__in=[item['id_checklist'] for item in lista],fechacreacion__in=[item['max_fechacreacion'] for item in lista]).order_by('-fecha_checklist'))
    """
    
    if usuario.faena.faena == 'SIN ASIGNAR' and request.user.role in ['ADMINISTRADOR', 'SUPERVISOR', 'BASE DATOS']:
        lista = ChecklistMaterialesSonda.objects.filter(
            Q(progreso='Corregido') | Q(progreso='Por Revisar'),
            etapa="Entrada",
            status=True
        ).values('id_checklist').annotate(max_fechacreacion=Max('fechacreacion'))
    else:
        lista = ChecklistMaterialesSonda.objects.filter(
            Q(progreso='Corregido') | Q(progreso='Por Revisar'),
            etapa="Entrada",
            status=True,
            sondajeCodigo__faena=usuario.faena,
        ).values('id_checklist').annotate(max_fechacreacion=Max('fechacreacion'))

    # Generar filtros precisos por cada combinación id_checklist + max_fechacreacion
    resultados = []
    for item in lista:
        r = (ChecklistMaterialesSonda.objects
            .filter(id_checklist=item['id_checklist'], fechacreacion=item['max_fechacreacion'])
            .order_by('-fechacreacion')  # o por fecha_checklist si quieres
            .first())
        if r:
            resultados.append(r)
    
    # Obtener todos los registros de EstadoEtapasReporteDigital relacionados
    estados = EstadoEtapasReporteDigital.objects.filter(
        id_checklist__in=[r.id_checklist for r in resultados]
    )

    # Agrupar por id_checklist
    from collections import defaultdict

    estado_por_checklist = defaultdict(list)
    for estado in estados:
        estado_por_checklist[estado.id_checklist].append(estado)

    # Evaluar completado por cada resultado
    for r in resultados:
        estados_asociados = estado_por_checklist.get(r.id_checklist)

        if estados_asociados:
            # Si hay uno o más registros para este id_checklist, revisamos si TODOS tienen los campos en True
            completado = all(
                e.reporteDigital is True and
                e.checklistEntrada is True and
                e.checklistSalida is True
                for e in estados_asociados
            )
            r.completado = completado
        else:
            # Si no hay ningún registro, según tu lógica, es completado = True
            r.completado = True 
    
    resultados.sort(key=lambda x: x.fechacreacion, reverse=True)
    
    context = {
        'reportes': resultados,
        'sidebar': 'manage_check_digital_reports',
        'sidebarmain': 'system_drilling', 
    }
    return render(request,'pages/drilling/manage_revisar_reportes_digitales.html', context)

@login_required
@sondaje_admin_or_base_datos_or_supervisor_required
def manage_todos_reportes_digitales(request):
    usuario = UsuarioProfile.objects.get(user=request.user)

    # --- LÓGICA AJAX (DataTables) ---
    if request.headers.get('x-requested-with') == 'XMLHttpRequest' and request.GET.get('draw'):
        
        # 1. Recuperar parámetros
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search_value = request.GET.get('search[value]', '')
        
        # Parámetros de Ordenamiento (Sorting)
        order_column_index = int(request.GET.get('order[0][column]', 0))
        order_direction = request.GET.get('order[0][dir]', 'desc')

        # 2. Query Base
        if usuario.faena.faena == 'SIN ASIGNAR' and request.user.role in ['ADMINISTRADOR', 'SUPERVISOR', 'BASE DATOS']:
            queryset = ChecklistMaterialesSonda.objects.filter(etapa="Entrada", status=True).exclude(progreso__in=('Eliminado', 'Por Revisar', 'Por Corregir'))
        else:
            queryset = ChecklistMaterialesSonda.objects.filter(etapa="Entrada", status=True, sondajeCodigo__faena=usuario.faena).exclude(progreso__in=('Eliminado', 'Por Revisar', 'Por Corregir'))

        # 3. FILTRO DE BÚSQUEDA AVANZADO
        if search_value:
            queryset = queryset.annotate(
                sondaje_full=Concat(
                    'sondajeCodigo__sondaje', Value('-'), 
                    Cast('sondajeSerie', CharField()), 
                    output_field=CharField()
                )
            ).filter(
                Q(id_checklist__icontains=search_value) |
                Q(sondaje_full__icontains=search_value) |
                Q(sondajeCodigo__sondaje__icontains=search_value) |
                Q(sondajeSerie__icontains=search_value) |
                Q(sondajeEstado__icontains=search_value) |
                Q(sonda__sonda__icontains=search_value) |
                Q(creador__icontains=search_value) |
                Q(turno__icontains=search_value)
            )

        # 4. AGRUPACIÓN INTELIGENTE
        grouped_queryset = queryset.values(
            'id_checklist', 
            'sondajeCodigo__sondaje', 
            'sondajeSerie', 
            'sondajeEstado',
            'sonda__sonda',
            'creador',
            'turno'
        ).annotate(max_fechacreacion=Max('fechacreacion'))

        # 5. LÓGICA DE ORDENAMIENTO
        column_mapping = {
            0: None,
            1: 'id_checklist',
            2: 'max_fechacreacion',
            3: 'sondajeSerie',
            4: 'sonda__sonda',
            5: 'creador',
            6: 'turno',
            7: None
        }

        col_name = column_mapping.get(order_column_index)

        if col_name:
            if order_direction == 'desc':
                grouped_queryset = grouped_queryset.order_by(f'-{col_name}')
            else:
                grouped_queryset = grouped_queryset.order_by(col_name)
        else:
            grouped_queryset = grouped_queryset.order_by('-max_fechacreacion')

        # Conteo para la paginación
        total_records = queryset.values('id_checklist').distinct().count()
        records_filtered = grouped_queryset.count()

        # 6. PAGINACIÓN
        page_items = grouped_queryset[start:start+length]

        # 7. RECUPERAR DATOS COMPLETOS
        data = []
        csrf_token = get_token(request)

        for item in page_items:
            reporte_obj = (ChecklistMaterialesSonda.objects
                .filter(id_checklist=item['id_checklist'], fechacreacion=item['max_fechacreacion'])
                .select_related('sondajeCodigo', 'sonda')
                .first())

            if reporte_obj:
                # --- MODIFICACIÓN AQUÍ ---
                # Convertimos la fecha UTC a la hora local configurada en settings (America/Santiago)
                fecha_local = timezone.localtime(reporte_obj.fecha_checklist)
                fecha_str = f"{fecha_local.strftime('%Y-%m-%d')} - {fecha_local.strftime('%H:%M')}"
                # -------------------------

                sondaje_str = f"{item['sondajeCodigo__sondaje']}-{item['sondajeSerie']}"
                if item['sondajeEstado']:
                    sondaje_str += f" {reporte_obj.get_sondajeEstado_display()}"

                botones_html = f"""
                <div class="form-inline-btn">
                    <form action="/editar_reporte_digital" method="POST">
                        <input type="hidden" name="csrfmiddlewaretoken" value="{csrf_token}">
                        <input type="hidden" name="id" value="{item['id_checklist']}">
                        <input type="hidden" name="origin" value="all_report">
                        <button type="submit" class="btn btn-info btn-size-edit">Ver</button>
                    </form>
                    <form class="report-digital-form" id="formulario-{reporte_obj.id}" method="POST">
                        <input type="hidden" name="csrfmiddlewaretoken" value="{csrf_token}">
                        <input type="hidden" name="id" value="{reporte_obj.id}">
                        <button type="submit" class="btn btn-secondary btn-size-edit">Pdf</button>
                    </form>
                """
                if request.user.role == "ADMINISTRADOR":
                    botones_html += f"""
                    <form action="/eliminar_reporte_digital_action" method="POST" class="form-delete">
                        <input type="hidden" name="csrfmiddlewaretoken" value="{csrf_token}">
                        <input type="hidden" name="id" value="{reporte_obj.id}">
                        <button type="submit" class="btn btn-danger btn-size-edit">Eliminar</button>
                    </form>
                    """
                botones_html += "</div>"

                data.append({
                    "id_checklist": item['id_checklist'],
                    "fecha": fecha_str,  # Usamos la fecha convertida
                    "sondaje": sondaje_str,
                    "sonda": item['sonda__sonda'],
                    "creador": item['creador'],
                    "turno": reporte_obj.get_turno_display(),
                    "acciones": botones_html
                })

        return JsonResponse({
            "draw": draw,
            "recordsTotal": total_records,
            "recordsFiltered": records_filtered,
            "data": data
        })

    # Carga normal HTML
    context = {
        'sidebar': 'manage_all_digital_reports',
        'sidebarmain': 'system_drilling', 
    }
    return render(request,'pages/drilling/manage_todos_reportes_digitales.html', context)

@login_required
@sondaje_admin_or_controlador_or_base_datos_or_supervisor_required
def progreso_reporte_digital(request):
    if request.method == 'POST':
        try:
            checklist = ChecklistMaterialesSonda.objects.get(id=request.POST['id'])
            
            # --- VALIDACIÓN DE CADENA COMPLETA (ANTES DE ENVIAR) ---
            # 1. Filtro Flexible para Gemelo
            estado_gemelo = checklist.sondajeEstado
            if not estado_gemelo: 
                filtro_gemelo = Q(sondajeEstado__isnull=True) | Q(sondajeEstado__exact='')
            else:
                filtro_gemelo = Q(sondajeEstado=estado_gemelo)

            # 2. Buscar huecos anteriores
            checklists_anteriores = ChecklistMaterialesSonda.objects.filter(
                filtro_gemelo,
                sondajeCodigo=checklist.sondajeCodigo,
                sondajeSerie=checklist.sondajeSerie,
                fechacreacion__lt=checklist.fechacreacion,
                status=True,
                etapa="Entrada" 
            ).exclude(progreso='Eliminado')

            ids_con_reporte = ReportesOperacionales.objects.filter(status=True).values('id_checklist')
            
            # Reportes anteriores que NO tienen operacional
            reportes_incompletos = checklists_anteriores.exclude(id_checklist__in=ids_con_reporte).order_by('fechacreacion')
            blocker = reportes_incompletos.first()

            if blocker:
                fecha_local = localtime(blocker.fechacreacion)
                messages.error(request, f"Secuencia Incompleta: Existe un reporte pendiente anterior ({fecha_local.strftime('%Y-%m-%d %H:%M')}). Complétalo antes de enviar este.", extra_tags='Acción Denegada')
                return redirect('manage_mis_reportes_digitales')
            # --- FIN VALIDACIÓN ---

            reporte = ReportesOperacionales.objects.get(id_checklist=checklist.id_checklist, status=True)
            controles = ControlesHorarios.objects.filter(reporte=reporte.id)
            total_horas = timedelta()
            for control in controles:
                total_horas += timedelta(hours=control.total.hour, minutes=control.total.minute, seconds=control.total.second)

            # Validación de 12 horas
            if total_horas < timedelta(hours=12) or total_horas > timedelta(hours=12):
                horas = total_horas.total_seconds() // 3600
                minutos = (total_horas.total_seconds() % 3600) // 60
                mensaje = f'Total de Horas Control Horario: {int(horas):02}:{int(minutos):02}'
                messages.error(request, mensaje, extra_tags="El total debe ser igual a 12:00 hrs")
                return redirect('manage_mis_reportes_digitales') 
            else:
                with transaction.atomic():
                    if (reporte.progreso == 'Creado'): 
                        ReportesOperacionales.objects.filter(id_checklist=checklist.id_checklist, status=True).update(progreso='Por Revisar')
                        ChecklistMaterialesSonda.objects.filter(id_checklist=checklist.id_checklist).update(progreso='Por Revisar')
                        messages.success(request, 'Reporte Enviado Correctamente')
                    elif (reporte.progreso == 'Por Corregir'):
                        ReportesOperacionales.objects.filter(id_checklist=checklist.id_checklist, status=True).update(progreso='Corregido')
                        ChecklistMaterialesSonda.objects.filter(id_checklist=checklist.id_checklist).update(progreso='Corregido')
                        messages.success(request, 'Reporte Corregido', extra_tags= 'Enviado Correctamente')
                    else:
                        messages.error(request, 'Error al Enviar el Reporte', extra_tags= 'Intente Nuevamente')
                
            return redirect('manage_mis_reportes_digitales') 
        except Exception as e:
            messages.error(request, f'Error inesperado: {str(e)}')
            return redirect('manage_mis_reportes_digitales')
    else:
        return redirect('manage_mis_reportes_digitales')

@login_required
@sondaje_admin_or_base_datos_or_supervisor_required
def progreso_reporte_digital_corregir(request):
    if request.method == 'POST': 
        checklist = ChecklistMaterialesSonda.objects.get(id=request.POST['id'])
        try:
            reporte = ReportesOperacionales.objects.get(id_checklist=checklist.id_checklist, status=True)
            if (reporte.progreso == 'Por Revisar' or reporte.progreso == 'Corregido'):
                # --- INICIO ATOMIC ---
                with transaction.atomic():
                    ReportesOperacionales.objects.filter(id_checklist=checklist.id_checklist, status=True).update(progreso='Por Corregir')
                    ChecklistMaterialesSonda.objects.filter(id_checklist=checklist.id_checklist).update(progreso='Por Corregir')
                # --- FIN ATOMIC ---
                messages.success(request, 'Reporte Enviado Para su Correción')
            else:
                messages.success(request, 'Error al Enviar el Reporte')
        except:
            messages.error(request, 'Error, complete primero el reporte antes de enviarlo a corregir')
        return redirect('manage_revisar_reportes_digitales') 
    else:
        messages.error(request, 'Error al Enviar el Reporte') 
        return redirect('manage_revisar_reportes_digitales') 
    
@login_required
@sondaje_admin_or_base_datos_or_supervisor_required
def progreso_reporte_digital_aprobar(request):
    if request.method == 'POST': 
        try:
            checklist = ChecklistMaterialesSonda.objects.get(id=request.POST['id'])
            
            # --- VALIDACIÓN (Igual que antes) ---
            estado_gemelo = checklist.sondajeEstado
            if not estado_gemelo: 
                filtro_gemelo = Q(sondajeEstado__isnull=True) | Q(sondajeEstado__exact='')
            else:
                filtro_gemelo = Q(sondajeEstado=estado_gemelo)

            checklists_anteriores = ChecklistMaterialesSonda.objects.filter(
                filtro_gemelo,
                sondajeCodigo=checklist.sondajeCodigo,
                sondajeSerie=checklist.sondajeSerie,
                fechacreacion__lt=checklist.fechacreacion,
                status=True,
                etapa="Entrada" 
            ).exclude(progreso='Eliminado')

            ids_con_reporte = ReportesOperacionales.objects.filter(status=True).values('id_checklist')
            reportes_incompletos = checklists_anteriores.exclude(id_checklist__in=ids_con_reporte).order_by('fechacreacion')
            blocker = reportes_incompletos.first()

            if blocker:
                fecha_local = localtime(blocker.fechacreacion)
                messages.error(request, f"Secuencia Incompleta: El reporte anterior ({fecha_local.strftime('%Y-%m-%d %H:%M')}) no está completo.", extra_tags='Acción Denegada')
                return redirect('manage_revisar_reportes_digitales')
            # --- FIN VALIDACIÓN ---

            reporte = ReportesOperacionales.objects.get(id_checklist=checklist.id_checklist, status=True)
            
            # --- FUNCIÓN AUXILIAR PARA EL HILO ---
            def correr_dashboard():
                # Esta función solo se ejecutará cuando la DB esté libre
                try:
                    GenerateCsv().actualizar_solo_sondajes()
                    print(f"--> Dashboard actualizado exitosamente para checklist {checklist.id_checklist}")
                except Exception as e_thread:
                    print(f"--> Error en hilo Dashboard: {e_thread}")

            # --- EJECUCIÓN ATÓMICA ---
            with transaction.atomic():
                if (reporte.progreso == 'Por Revisar' or reporte.progreso == 'Corregido'):
                    # 1. Actualizamos DB (Escribimos)
                    ReportesOperacionales.objects.filter(id_checklist=checklist.id_checklist, status=True).update(progreso='Aprobado')
                    ChecklistMaterialesSonda.objects.filter(id_checklist=checklist.id_checklist).update(progreso='Aprobado')
                    
                    # 2. Programamos el hilo para DESPUÉS del commit (Evita el Deadlock)
                    transaction.on_commit(lambda: threading.Thread(target=correr_dashboard).start())

                    messages.success(request, 'Reporte Aprobado Correctamente')
                else:
                    messages.warning(request, 'El reporte no está en estado válido para aprobación.')

        except Exception as e:
            messages.error(request, 'Error al procesar la solicitud.')
        
        return redirect('manage_revisar_reportes_digitales') 
    else:
        return redirect('manage_revisar_reportes_digitales')
    
@login_required
@sondaje_admin_or_controlador_or_base_datos_or_supervisor_required
def editar_reporte_digital(request):
    try:
        # 1. Recuperar datos básicos
        if request.method == 'POST':
            id_checklist_request = request.POST.get('id')
            origin_request = request.POST.get('origin')
        else:
            id_checklist_request = request.session.get('edit_reporte_digital')
            origin_request = request.session.get('edit_reporte_digital_origin')

        if not id_checklist_request:
            return redirect('dashboardSondaje')

        # 2. Identificar el Checklist Actual
        checklist_actual = ChecklistMaterialesSonda.objects.filter(
            id_checklist=id_checklist_request, 
            etapa='Entrada'
        ).first()
        
        if not checklist_actual:
            checklist_actual = ChecklistMaterialesSonda.objects.filter(id_checklist=id_checklist_request).first()

        if not checklist_actual:
            messages.error(request, 'No se encontraron datos para el checklist solicitado.')
            return redirect('dashboardSondaje')

        # --- DETECCIÓN DE MODO LECTURA ---
        es_solo_lectura = checklist_actual.progreso in ['Aprobado', 'Eliminado']

        # --- VALIDACIÓN DE CADENA COMPLETA (Solo si NO es lectura) ---
        if not es_solo_lectura:
            
            # B) Filtro Flexible para Gemelo
            estado_gemelo = checklist_actual.sondajeEstado
            if not estado_gemelo: 
                filtro_gemelo = Q(sondajeEstado__isnull=True) | Q(sondajeEstado__exact='')
            else:
                filtro_gemelo = Q(sondajeEstado=estado_gemelo)

            # C) BUSCAR HUECOS EN LA HISTORIA
            checklists_anteriores = ChecklistMaterialesSonda.objects.filter(
                filtro_gemelo,
                sondajeCodigo=checklist_actual.sondajeCodigo,
                sondajeSerie=checklist_actual.sondajeSerie,
                fechacreacion__lt=checklist_actual.fechacreacion, # Todo lo anterior
                status=True,
                etapa="Entrada" 
            ).exclude(progreso='Eliminado')

            ids_con_reporte = ReportesOperacionales.objects.filter(status=True).values('id_checklist')

            reportes_incompletos = checklists_anteriores.exclude(
                id_checklist__in=ids_con_reporte
            ).order_by('fechacreacion')

            blocker = reportes_incompletos.first()

            if blocker:
                # Definir redirección
                redirect_target = 'dashboardSondaje'
                if origin_request == 'my_report':
                    redirect_target = 'manage_mis_reportes_digitales'
                elif origin_request == 'check_report':
                    redirect_target = 'manage_revisar_reportes_digitales'
                elif origin_request == 'all_report':
                    redirect_target = 'manage_todos_reportes_digitales'
                elif origin_request == 'deleted_report':
                    redirect_target = 'manage_mis_reportes_digitales_eliminados'

                fecha_local = localtime(blocker.fechacreacion)
                fecha_formateada = fecha_local.strftime('%Y-%m-%d %H:%M')

                messages.error(
                    request, 
                    f"Secuencia Incompleta: Existe un reporte pendiente anterior con fecha {fecha_formateada} para este pozo. Debes completar o eliminar los reportes previos antes de editar este.",
                    extra_tags='Acceso Denegado'
                )
                return redirect(redirect_target)

        # --- FIN VALIDACIÓN ---

        # 3. Lógica de carga de sesión y Datos
        request.session['edit_reporte_digital'] = id_checklist_request
        request.session['edit_reporte_digital_origin'] = origin_request
        origin = request.session['edit_reporte_digital_origin']
        disabled_sondaje_codigo = True
        
        # Carga Checklist Entrada
        checklistEntrada = ChecklistMaterialesSonda.objects.filter(id_checklist= request.session['edit_reporte_digital'], etapa='Entrada', status=True)
        first_checklist_entrada = checklistEntrada.first()
        
        formEntrada = ChecklistMaterialesSondaEntradaForm(initial={
            'fecha_checklist_entrada': first_checklist_entrada.fecha_checklist,
            'turno_entrada': first_checklist_entrada.turno,
            'jornada_entrada': first_checklist_entrada.jornada,
            'sonda_entrada': first_checklist_entrada.sonda,
            'sondaje_entrada': first_checklist_entrada.sondajeCodigo,
            'serie_entrada': first_checklist_entrada.sondajeSerie,
            'estado_entrada': first_checklist_entrada.sondajeEstado,
            'id_checklist': first_checklist_entrada.id_checklist,
            },
            modo_edicion=True, 
        )

        # Carga Checklist Salida
        checklistSalida = ChecklistMaterialesSonda.objects.filter(id_checklist=request.session['edit_reporte_digital'], etapa='Salida', status=True)
        first_checklist_salida = checklistSalida.first()

        if first_checklist_salida:
            formSalida = ChecklistMaterialesSondaSalidaForm(initial={'fecha_checklist_salida': first_checklist_salida.fecha_checklist, 'turno_salida': first_checklist_salida.turno, 'jornada_salida': first_checklist_salida.jornada, 'sonda_salida': first_checklist_salida.sonda.id, 'sondaje_salida': first_checklist_salida.sondajeCodigo.id, 'serie_salida': first_checklist_salida.sondajeSerie, 'estado_salida': first_checklist_salida.sondajeEstado}, modo_edicion=True)
            existe_checklist_salida = "Existe"
        else:
            formSalida = ChecklistMaterialesSondaSalidaForm(initial={'turno_salida': first_checklist_entrada.turno, 'jornada_salida': '2', 'sonda_salida': first_checklist_entrada.sonda.id, 'sondaje_salida': first_checklist_entrada.sondajeCodigo.id, 'serie_salida': first_checklist_entrada.sondajeSerie, 'estado_salida': first_checklist_entrada.sondajeEstado}, modo_edicion=True)
            existe_checklist_salida = "No Existe"

        materialesSonda = MaterialesSonda.objects.filter(status=True)
        
        # Sidebar logic
        if first_checklist_entrada.progreso == 'Creado' or first_checklist_entrada.progreso == 'Por Corregir':
                sidebar = 'manage_my_digital_reports'
        elif first_checklist_entrada.progreso == 'Por Revisar' or first_checklist_entrada.progreso == 'Corregido':
            sidebar = 'manage_check_digital_reports'
        elif first_checklist_entrada.progreso == 'Eliminado':
            sidebar = 'manage_my_digital_reports_deleted'
        else:
            sidebar = 'manage_all_digital_reports'

        try:
            reporte = ReportesOperacionales.objects.get(id_checklist=request.session['edit_reporte_digital'], status=True)
            existe_reporte = "Existe"
            # Carga de datos JSON (Serialización)
            reporte_detalle_perforacion = DetallesPerforaciones.objects.filter(reporte=reporte, status=True)
            reporte_detalle_perforacion_json = json.dumps([{ 'reporte': item.reporte.id, 'diametros': item.diametros.id if item.diametros else None, 'perforado': f"{(item.perforado or 0):.2f}", 'desde': f"{(item.desde or 0):.2f}", 'hasta': f"{(item.hasta or 0):.2f}", 'recuperacion': f"{(item.recuperacion or 0):.2f}", 'porcentajeRecuperacion': f"{(item.porcentajeRecuperacion or 0):.2f}", 'barra': item.barra, 'largoBarra': f"{(item.largoBarra or 0):.2f}", 'totalHtas': f"{(item.totalHtas or 0):.2f}", 'contra': f"{(item.contra or 0):.2f}", 'tipoTerreno': item.tipoTerreno.id if item.tipoTerreno else None, 'orientacion': item.orientacion.id if item.orientacion else None, 'fechacreacion': item.fechacreacion.isoformat(), } for item in reporte_detalle_perforacion])
            reporte_control_horario = ControlesHorarios.objects.filter(reporte=reporte, status=True)
            reporte_control_horario_json = json.dumps([{ 'reporte': item.reporte.id, 'posicion': item.posicion, 'inicio': item.inicio.strftime('%H:%M'), 'final': item.final.strftime('%H:%M'), 'total': item.total.strftime('%H:%M'), 'detalleControlHorario': item.detalleControlHorario.id if item.detalleControlHorario else None, 'fechacreacion': item.fechacreacion.isoformat(), } for item in reporte_control_horario])
            reporte_detalle_aditivos = DetalleAditivos.objects.filter(reporte=reporte, status=True)
            reporte_detalle_aditivos_json = json.dumps([{ 'reporte': item.reporte.id, 'aditivo': item.aditivo.id if item.aditivo else None, 'cantidad': item.cantidad, 'fechacreacion': item.fechacreacion.isoformat(), } for item in reporte_detalle_aditivos])
            
            reporte_insumos = Insumos.objects.get(reporte=reporte, status=True)
            reporte_longitud_pozos = LongitudPozos.objects.get(reporte=reporte, status=True)
            reporte_observaciones = ObservacionesReportes.objects.get(reporte=reporte, status=True)
            control_horario = DetalleControlHorario.objects.filter(status=True).order_by('detalle')
            diametros = Diametros.objects.filter(status=True).order_by('diametro')
            tipo_terreno = TipoTerreno.objects.filter(status=True).order_by('tipoTerreno')
            orientacion = Orientacion.objects.filter(status=True).order_by('orientacion')
            aditivos = Aditivos.objects.filter(status=True).order_by('aditivo')
            usuario_profile = UsuarioProfile.objects.get(user=request.user)
            try:
                Sondajes.objects.get(faena=usuario_profile.faena, status=True)
                disabled_sondaje_codigo=True
            except Sondajes.DoesNotExist:
                disabled_sondaje_codigo=True
            
            disabled_reporte_detalle_perforacion = True
            
            # Instanciar Formularios
            form_reporte_operacional = FormReportesOperacionales(initial={
                    'fechacreacion': reporte.fechacreacion,
                    'turno': reporte.turno,
                    'perforista': reporte.perforista,
                    'controlador': reporte.controlador.first_name+" "+reporte.controlador.last_name,
                    'sonda': reporte.sonda,
                    'sondajeCodigo': reporte.sondajeCodigo,
                    'sondajeSerie': reporte.sondajeSerie,
                    'sondajeEstado': reporte.sondajeEstado,
                    'metroInicial': f"{reporte.metroInicial:.2f}" if reporte.metroInicial is not None else "",
                    'metroFinal': f"{reporte.metroFinal:.2f}" if reporte.metroFinal is not None else "",
                    'totalPerforado': f"{reporte.totalPerforado:.2f}" if reporte.totalPerforado is not None else "",
                },
                disabled_sondaje_codigo=disabled_sondaje_codigo,
                disabled_reporte_detalle_perforacion=disabled_reporte_detalle_perforacion,
            )
            form_insumos = FormInsumos(initial={
                    'corona': reporte_insumos.corona,
                    'escareador': reporte_insumos.escareador,
                    'cantidadAgua': reporte_insumos.cantidadAgua,
                    'casing': reporte_insumos.casing,
                    'zapata': reporte_insumos.zapata,
                })
            form_longitud = FormLongitudPozos(initial={
                    'largoBarril': f"{reporte_longitud_pozos.largoBarril:.2f}" if reporte_longitud_pozos.largoBarril is not None else "",
                    'puntoMuerto': f"{reporte_longitud_pozos.puntoMuerto:.2f}" if reporte_longitud_pozos.puntoMuerto is not None else "",
                    'restoBarra': f"{reporte_longitud_pozos.restoBarra:.2f}" if reporte_longitud_pozos.restoBarra is not None else "",
                    'numeroBarras': reporte_longitud_pozos.numeroBarras,
                    'longitudPozo': f"{reporte_longitud_pozos.longitudPozo:.2f}" if reporte_longitud_pozos.longitudPozo is not None else "",
                    'htaEnPozo': reporte_longitud_pozos.htaEnPozo,
                    'mtsDeHta': f"{reporte_longitud_pozos.mtsDeHta:.2f}" if reporte_longitud_pozos.mtsDeHta is not None else "",
                    'profundidadHta': f"{reporte_longitud_pozos.profundidadHta:.2f}" if reporte_longitud_pozos.profundidadHta is not None else "",
                })
            form_observaciones = FormObservacionesReportes(initial={
                    'observaciones': reporte_observaciones.observaciones,
                })

            # --- APLICAR MODO SOLO LECTURA ---
            if es_solo_lectura:
                def disable_form(form):
                    for field in form.fields.values():
                        field.widget.attrs['disabled'] = True
                        field.required = False
                
                disable_form(formEntrada)
                disable_form(formSalida)
                disable_form(form_reporte_operacional)
                disable_form(form_insumos)
                disable_form(form_longitud)
                disable_form(form_observaciones)

            context = {
                'sondajeCodigo': first_checklist_entrada.sondajeCodigo,
                'sondajeSerie': first_checklist_entrada.sondajeSerie,
                'sondajeEstado': first_checklist_entrada.sondajeEstado,
                'formInicio': formEntrada,
                'formFinal': formSalida,
                'checklistEntrada': checklistEntrada,
                'checklistSalida': checklistSalida,
                'materialesSonda': materialesSonda,
                'existe_reporte': existe_reporte,
                'existe_checklist_salida': existe_checklist_salida,
                'id_checklist': request.session['edit_reporte_digital'],
                'origen': origin,
                'reporte': reporte,
                'reporte_detalle_perforacion': reporte_detalle_perforacion_json,
                'reporte_control_horario': reporte_control_horario_json,
                'reporte_detalle_aditivos': reporte_detalle_aditivos_json,
                'detalles_control_horario': control_horario,
                'diametros': diametros,
                'tipos_terreno': TipoTerreno.objects.filter(status=True).order_by('tipoTerreno'),
                'orientaciones': Orientacion.objects.filter(status=True).order_by('orientacion'),
                'aditivos': aditivos,
                'formreporteoperacional': form_reporte_operacional,
                'formdetalleperforacion': FormDetallesPerforaciones,
                'formcontrolhorario': FormControlesHorarios,
                'forminsumos': form_insumos,
                'formaditivos': FormAditivos,
                'formlongitudpozo': form_longitud,
                'formobservaciones': form_observaciones,
                'sidebar': sidebar,
                'sidebarmain': 'system_drilling', 
                'es_solo_lectura': es_solo_lectura,
            }
            return render(request, 'pages/drilling/edit_reporte_digital.html', context)
        
        except ReportesOperacionales.DoesNotExist:
            existe_reporte = "No Existe"
            control_horario = DetalleControlHorario.objects.filter(status=True).order_by('detalle')
            diametros = Diametros.objects.filter(status=True).order_by('diametro')
            tipo_terreno = TipoTerreno.objects.filter(status=True).order_by('tipoTerreno')
            orientacion = Orientacion.objects.filter(status=True).order_by('orientacion')
            aditivos = Aditivos.objects.filter(status=True).order_by('aditivo')
            usuario_profile = UsuarioProfile.objects.get(user=request.user)
            
            reporte_previo = ReportesOperacionales.objects.filter(
                sondajeCodigo=first_checklist_entrada.sondajeCodigo,
                sondajeSerie=first_checklist_entrada.sondajeSerie,
                sondajeEstado=first_checklist_entrada.sondajeEstado if first_checklist_entrada.sondajeEstado else None
            ).order_by('-fechacreacion').first()
            
            if not reporte_previo and not first_checklist_entrada.sondajeEstado:
                 reporte_previo = ReportesOperacionales.objects.filter(
                    Q(sondajeEstado__isnull=True) | Q(sondajeEstado__exact=''),
                    sondajeCodigo=first_checklist_entrada.sondajeCodigo,
                    sondajeSerie=first_checklist_entrada.sondajeSerie
                ).order_by('-fechacreacion').first()

            form_reporte_operacional = FormReportesOperacionales(initial={
                    'controlador': request.user.first_name+" "+request.user.last_name,
                    'sondajeCodigo': first_checklist_entrada.sondajeCodigo,
                    'sondajeSerie': first_checklist_entrada.sondajeSerie,
                    'sondajeEstado': first_checklist_entrada.sondajeEstado,
                    'turno': first_checklist_entrada.turno,
                    'sonda': first_checklist_entrada.sonda,
                    'metroInicial': (f"{reporte_previo.metroInicial:.2f}" if reporte_previo is not None and reporte_previo.metroInicial is not None else "0.00"),
                },
                disabled_sondaje_codigo=disabled_sondaje_codigo,
            )

            # Si el checklist principal está cerrado pero no hay reporte (caso raro), también bloqueamos
            if es_solo_lectura:
                def disable_form(form):
                    for field in form.fields.values():
                        field.widget.attrs['disabled'] = True
                        field.required = False
                disable_form(formEntrada)
                disable_form(formSalida)
                disable_form(form_reporte_operacional)
                # Los otros formularios están vacíos aquí, no es crítico desactivarlos, pero si quieres:
                # disable_form(FormInsumos()) etc.

            context = {
                'reporte': checklist_actual,
                'sondajeCodigo': first_checklist_entrada.sondajeCodigo,
                'sondajeSerie': first_checklist_entrada.sondajeSerie,
                'sondajeEstado': first_checklist_entrada.sondajeEstado,
                'origen': origin,
                'existe_reporte': existe_reporte,
                'existe_checklist_salida': existe_checklist_salida,
                'id_checklist': request.session['edit_reporte_digital'],
                'formInicio': formEntrada,
                'formFinal': formSalida,
                'checklistEntrada': checklistEntrada,
                'checklistSalida': checklistSalida,
                'materialesSonda': materialesSonda,
                'detalles_control_horario': control_horario,
                'diametros': diametros,
                'tipos_terreno': tipo_terreno,
                'orientaciones': orientacion,
                'aditivos': aditivos,
                'formreporteoperacional': form_reporte_operacional,
                'formdetalleperforacion': FormDetallesPerforaciones,
                'formcontrolhorario': FormControlesHorarios,
                'forminsumos': FormInsumos,
                'formaditivos': FormAditivos,
                'formlongitudpozo': FormLongitudPozos,
                'formobservaciones': FormObservacionesReportes,
                'sidebar': sidebar,
                'sidebarmain': 'system_drilling', 
                'es_solo_lectura': es_solo_lectura,
            }
            return render(request, 'pages/drilling/edit_reporte_digital.html', context)

    except Exception as e:
        messages.error(request, f'Ocurrió un error inesperado al intentar acceder a la edición: {str(e)}')
        return redirect('dashboardSondaje')

@login_required
@sondaje_admin_or_controlador_or_base_datos_or_supervisor_required
def save_editar_reporte_digital(request):
    if request.method == 'POST':
        # 1. Validación de Horas (Siempre primero)
        total_horas = sumar_totales_control_horario(request.POST)
        if total_horas != timedelta(hours=12):
            horas = int(total_horas.total_seconds() // 3600)
            minutos = int((total_horas.total_seconds() % 3600) // 60)
            return JsonResponse({
                'titleText': f'Total de Horas: {horas:02}:{minutos:02}',
                'text': 'El total de horas debe ser exactamente 12:00 hrs.'
            }, status=400)

        sondaje_estado = "" if request.POST['sondajeEstado'] == '---------' else request.POST['sondajeEstado']

        try:
            with transaction.atomic():
                # ------------------------------------------------------------------
                # 2. RECUPERAR Y BLOQUEAR (select_for_update)
                # ------------------------------------------------------------------
                try:
                    # select_for_update() "congela" el registro en la BD.
                    # Si la Ventana B intenta guardar, tendrá que esperar a que Ventana A termine.
                    reporte_original = ReportesOperacionales.objects.select_for_update().get(
                        id_checklist=request.POST['id'], 
                        status=True
                    )
                
                except ReportesOperacionales.DoesNotExist:
                    # Caso: Ventana A ya guardó y puso status=False. Ventana B llega tarde.
                    return JsonResponse({
                        'titleText': 'Reporte No Disponible', 
                        'text': 'El reporte ya fue modificado o eliminado en otra ventana. Refresque la página.'
                    }, status=404)
                
                except MultipleObjectsReturned:
                    # Caso: YA existen duplicados (Tu error "entrega 2 y pide 1").
                    # Aquí detenemos todo para no ensuciar más la base de datos.
                    return JsonResponse({
                        'titleText': 'Error Crítico de Consistencia', 
                        'text': 'Existen múltiples registros activos para este reporte. Contacte a soporte para limpiar la base de datos.'
                    }, status=500)

                # ------------------------------------------------------------------
                # 3. VALIDACIÓN DE ESTADO (Seguridad Anti-Aprobación)
                # ------------------------------------------------------------------
                if reporte_original.progreso in ['Aprobado', 'Eliminado']:
                    return JsonResponse({
                        'titleText': 'Acción Denegada', 
                        'text': f'El reporte cambió a estado "{reporte_original.progreso}" en otra ventana. No se pueden guardar cambios.'
                    }, status=403)

                # ------------------------------------------------------------------
                # 4. DESACTIVAR REPORTE ANTERIOR (Update seguro)
                # ------------------------------------------------------------------
                reporte_original.status = False
                reporte_original.save()
                
                # Desactivar hijos de forma masiva
                DetallesPerforaciones.objects.filter(reporte=reporte_original, status=True).update(status=False)
                ControlesHorarios.objects.filter(reporte=reporte_original, status=True).update(status=False)
                Insumos.objects.filter(reporte=reporte_original, status=True).update(status=False)
                DetalleAditivos.objects.filter(reporte=reporte_original, status=True).update(status=False)
                LongitudPozos.objects.filter(reporte=reporte_original, status=True).update(status=False)
                ObservacionesReportes.objects.filter(reporte=reporte_original, status=True).update(status=False)

                # ------------------------------------------------------------------
                # 5. CREAR NUEVO REPORTE
                # ------------------------------------------------------------------
                try:
                    reporte_operacional = ReportesOperacionales.objects.create(
                        fechacreacion=reporte_original.fechacreacion,
                        turno=request.POST['turno'],
                        perforista_id=request.POST['perforista'],
                        sonda_id=request.POST['sonda'],
                        sondajeCodigo_id=request.POST['sondajeCodigo'],
                        sondajeSerie=request.POST['sondajeSerie'],
                        sondajeEstado=sondaje_estado,
                        metroInicial=normalizar_decimal(request.POST['metroInicial']),
                        metroFinal=normalizar_decimal(request.POST['metroFinal']),
                        totalPerforado=normalizar_decimal(request.POST['totalPerforado']),
                        controlador=reporte_original.controlador, 
                        creador=request.user.first_name+' ' +request.user.last_name,
                        status=True,
                        id_checklist=request.POST['id'],
                        correlativo=reporte_original.correlativo,
                        progreso=reporte_original.progreso, 
                    )
                    
                    # --- HIJOS (Detalles, etc.) ---
                    
                    # 1. Detalles Perforación
                    try:
                        perforacion_count = len([key for key in request.POST.keys() if key.startswith('diametros_')])
                        for i in range(1, perforacion_count + 1):
                            DetallesPerforaciones.objects.create(
                                reporte=reporte_operacional,
                                diametros_id=request.POST.get(f'diametros_{i}'),
                                perforado=normalizar_decimal(request.POST.get(f'perforado_{i}')),
                                desde=normalizar_decimal(request.POST.get(f'desde_{i}')),
                                hasta=normalizar_decimal(request.POST.get(f'hasta_{i}')),
                                recuperacion=normalizar_decimal(request.POST.get(f'recuperacion_{i}')),
                                porcentajeRecuperacion=normalizar_decimal(request.POST.get(f'porcentajeRecuperacion_{i}')),
                                barra=int(request.POST.get(f'barra_{i}')) if request.POST.get(f'barra_{i}') and request.POST.get(f'barra_{i}').isdigit() else 0,
                                largoBarra=normalizar_decimal(request.POST.get(f'largoBarra_{i}')),
                                totalHtas=normalizar_decimal(request.POST.get(f'totalHtas_{i}')),
                                contra=normalizar_decimal(request.POST.get(f'contra_{i}')),
                                tipoTerreno_id=request.POST.get(f'tipoTerreno_{i}'),
                                orientacion_id=request.POST.get(f'orientacion_{i}'),
                                status=True,
                            )
                    except Exception as e:
                        raise Exception("Error en Detalle Perforación")

                    # 2. Control Horario
                    try:
                        control_horario_count = len([key for key in request.POST.keys() if key.startswith('inicio_')])
                        for i in range(1, control_horario_count + 1):
                            inicio_time = datetime.datetime.strptime(request.POST.get(f'inicio_{i}'), '%H:%M').time()
                            final_time = datetime.datetime.strptime(request.POST.get(f'final_{i}'), '%H:%M').time()
                            total_time = datetime.datetime.strptime(request.POST.get(f'total_{i}'), '%H:%M').time()
                            ControlesHorarios.objects.create(
                                reporte=reporte_operacional,
                                posicion=i, 
                                inicio=inicio_time,
                                final=final_time,
                                total=total_time,
                                detalleControlHorario_id=request.POST.get(f'detalleControlHorario_{i}'),
                                status=True,
                            )
                    except Exception as e:
                        raise Exception("Error en Control Horario")

                    # 3. Insumos
                    try:
                        Insumos.objects.create(
                            reporte=reporte_operacional,
                            corona=request.POST['corona'],
                            escareador=request.POST['escareador'],
                            cantidadAgua_id=request.POST['cantidadAgua'],
                            casing=normalizar_decimal(request.POST['casing']),
                            zapata=request.POST['zapata'],
                            status=True,
                        )
                    except Exception as e:
                        raise Exception("Error en Insumos")

                    # 4. Aditivos
                    try:
                        aditivo_count = len([key for key in request.POST.keys() if key.startswith('aditivo_')])
                        for i in range(1, aditivo_count + 1):
                            DetalleAditivos.objects.create(
                                reporte=reporte_operacional,
                                aditivo_id=request.POST.get(f'aditivo_{i}'),
                                cantidad=normalizar_entero(request.POST.get(f'cantidad_{i}')),
                                status=True,
                            )
                    except Exception as e:
                        raise Exception("Error en Aditivos")

                    # 5. Longitud Pozo
                    try:
                        LongitudPozos.objects.create(
                            reporte=reporte_operacional, 
                            largoBarril=normalizar_decimal(request.POST['largoBarril']),
                            puntoMuerto=normalizar_decimal(request.POST['puntoMuerto']),
                            restoBarra=normalizar_decimal(request.POST['restoBarra']),
                            numeroBarras=request.POST['numeroBarras'],
                            longitudPozo=normalizar_decimal(request.POST['longitudPozo']),
                            htaEnPozo=request.POST['htaEnPozo'],
                            mtsDeHta=normalizar_decimal(request.POST['mtsDeHta']),
                            profundidadHta=normalizar_decimal(request.POST['profundidadHta']),
                            status=True,
                        )
                    except Exception as e:
                        raise Exception("Error en Longitud de Pozo")

                    # 6. Observaciones
                    try:
                        ObservacionesReportes.objects.create(
                            reporte=reporte_operacional,
                            observaciones=request.POST['observaciones'],
                            status=True,
                        )
                    except Exception as e:
                        raise Exception("Error en Observaciones")

                except Exception as e:
                    # Este raise activa el rollback total
                    raise Exception(f"Error al guardar los nuevos datos: {str(e)}")

                # Redirecciones
                if request.POST['origin'] == 'my_report':
                    return redirect('manage_mis_reportes_digitales')
                elif request.POST['origin'] == 'check_report':
                    return redirect('manage_revisar_reportes_digitales')
                elif request.POST['origin'] == 'all_report':
                    return redirect('manage_todos_reportes_digitales')
                elif request.POST['origin'] == 'deleted_report':
                    return redirect('manage_todos_reportes_digitales_eliminados')
                else:
                    return redirect('dashboardSondaje')

        except Exception as e:
            # Captura final que asegura respuesta JSON
            return JsonResponse({'titleText': 'Error al Guardar', 'text': str(e)}, status=400)
    else:
        return redirect('dashboardSondaje')
    
@require_GET
@sondaje_admin_or_controlador_or_base_datos_or_supervisor_required
def obtener_estado_sonda(request):
    sonda = Sondas.objects.get(id=request.GET.get('sonda'))
    reporte = ReportesOperacionales.objects.filter(sonda=sonda, status=True).exclude(progreso="Aprobado").order_by('-fechacreacion').first()
    return JsonResponse({'reporte': reporte,}, status=200)

@login_required
def reporte_digital_pdf_view(request):
    if request.method == 'POST':
        try:
            checklist = ChecklistMaterialesSonda.objects.get(id=request.POST['id'])
            reporte = get_object_or_404(ReportesOperacionales, id_checklist=checklist.id_checklist, status=True)
            perforaciones = DetallesPerforaciones.objects.filter(reporte=reporte)
            control_horario = ControlesHorarios.objects.filter(reporte=reporte)
            aditivos = DetalleAditivos.objects.filter(reporte=reporte)
            insumos = get_object_or_404(Insumos, reporte=reporte)
            longitud_pozo = get_object_or_404(LongitudPozos, reporte=reporte)
            observaciones = get_object_or_404(ObservacionesReportes, reporte=reporte)
            current_datetime = datetime.datetime.now()
            context = {
                'reporte': reporte,
                'perforaciones': perforaciones,
                'control_horarios': control_horario,
                'aditivos': aditivos,
                'insumos': insumos,
                'longitud': longitud_pozo,
                'observaciones': observaciones,
                'current_datetime': current_datetime,
            }
            template_path = 'pages/pdfs/digital_report_pdf_template.html'
            template = get_template(template_path)
            html = template.render(context)
            
            sondaje_estado = reporte.get_sondajeEstado_display()
            if sondaje_estado:
                filename = f'{reporte.sondajeCodigo}-{reporte.sondajeSerie}-{sondaje_estado}'
            else:
                filename = f'{reporte.sondajeCodigo}-{reporte.sondajeSerie}'
            subfolder_name = f'{filename}'
            subfolder_path = os.path.join(settings.MEDIA_ROOT, 'reportes_digitales', subfolder_name)
            os.makedirs(subfolder_path, exist_ok=True)
            filename = f'{filename} {reporte.fechacreacion.strftime("%Y%m%d-%H%M%S")}.pdf'
            pdf_path = os.path.join(subfolder_path, filename)

            with open(pdf_path, 'wb') as pdf_file:
                pisa_status = pisa.CreatePDF(html, dest=pdf_file)
                if pisa_status.err:
                    return JsonResponse({'error': 'Error al generar el PDF'})

            pdf_url = os.path.join(settings.MEDIA_URL, 'reportes_digitales', subfolder_name, filename)

            return JsonResponse({'pdf_url': pdf_url, 'message': 'Documento creado con éxito'})
        except Exception as e:
            return JsonResponse({'error': f'Error inesperado: {str(e)}'}, status=500)
    else:
        return redirect('manage_mis_reportes_digitales')


# views.py

@login_required
@sondaje_admin_or_controlador_or_base_datos_or_supervisor_required
def manage_mis_reportes_digitales_eliminados(request): 
    usuario = UsuarioProfile.objects.get(user=request.user)

    # --- LÓGICA AJAX (DataTables Server-Side) ---
    if request.headers.get('x-requested-with') == 'XMLHttpRequest' and request.GET.get('draw'):
        
        # 1. Recuperar parámetros de DataTables
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search_value = request.GET.get('search[value]', '')
        
        # Parámetros de Ordenamiento
        order_column_index = int(request.GET.get('order[0][column]', 0))
        order_direction = request.GET.get('order[0][dir]', 'desc')

        # 2. Query Base (Filtrar solo los 'Eliminado')
        if usuario.faena.faena == 'SIN ASIGNAR' and request.user.role in ['ADMINISTRADOR', 'SUPERVISOR', 'BASE DATOS']:
            queryset = ChecklistMaterialesSonda.objects.filter(
                etapa="Entrada", 
                status=True, 
                progreso='Eliminado'
            )
        else:
            queryset = ChecklistMaterialesSonda.objects.filter(
                etapa="Entrada", 
                status=True, 
                sondajeCodigo__faena=usuario.faena, 
                progreso='Eliminado'
            )

        # 3. FILTRO DE BÚSQUEDA AVANZADO
        if search_value:
            queryset = queryset.annotate(
                # Campo virtual para buscar "Sondaje-Serie"
                sondaje_full=Concat(
                    'sondajeCodigo__sondaje', Value('-'), 
                    Cast('sondajeSerie', CharField()), 
                    output_field=CharField()
                )
            ).filter(
                Q(id_checklist__icontains=search_value) |
                Q(sondaje_full__icontains=search_value) |
                Q(sondajeCodigo__sondaje__icontains=search_value) |
                Q(sondajeSerie__icontains=search_value) |
                Q(sondajeEstado__icontains=search_value) |
                Q(sonda__sonda__icontains=search_value) |
                Q(creador__icontains=search_value) |
                Q(turno__icontains=search_value)
            )

        # 4. AGRUPACIÓN (Agrupar versiones, mostrar la última fecha de eliminación)
        grouped_queryset = queryset.values(
            'id_checklist', 
            'sondajeCodigo__sondaje', 
            'sondajeSerie', 
            'sondajeEstado',
            'sonda__sonda',
            'creador',
            'turno'
        ).annotate(max_fechacreacion=Max('fechacreacion'))

        # 5. ORDENAMIENTO (Mapping de columnas igual al HTML)
        column_mapping = {
            0: None,                             # Expander
            1: 'id_checklist',                   # #
            2: 'max_fechacreacion',              # Fecha
            3: 'sondajeSerie',                   # Sondaje
            4: 'sonda__sonda',                   # Sonda
            5: 'creador',                        # Controlador
            6: 'turno',                          # Turno
            7: None                              # Acciones
        }

        col_name = column_mapping.get(order_column_index)

        if col_name:
            if order_direction == 'desc':
                grouped_queryset = grouped_queryset.order_by(f'-{col_name}')
            else:
                grouped_queryset = grouped_queryset.order_by(col_name)
        else:
            grouped_queryset = grouped_queryset.order_by('-max_fechacreacion')

        # Conteo para Paginación
        total_records = queryset.values('id_checklist').distinct().count()
        records_filtered = grouped_queryset.count()

        # 6. PAGINACIÓN
        page_items = grouped_queryset[start:start+length]

        # 7. CONSTRUCCIÓN DE DATOS JSON
        data = []
        csrf_token = get_token(request)

        for item in page_items:
            # Recuperar objeto completo
            reporte_obj = (ChecklistMaterialesSonda.objects
                .filter(id_checklist=item['id_checklist'], fechacreacion=item['max_fechacreacion'])
                .select_related('sondajeCodigo', 'sonda')
                .first())

            if reporte_obj:
                # --- MODIFICACIÓN: Conversión a hora local ---
                fecha_local = timezone.localtime(reporte_obj.fecha_checklist)
                fecha_str = f"{fecha_local.strftime('%Y-%m-%d')} - {fecha_local.strftime('%H:%M')}"
                # ---------------------------------------------

                sondaje_str = f"{item['sondajeCodigo__sondaje']}-{item['sondajeSerie']}"
                if item['sondajeEstado']:
                    sondaje_str += f" {reporte_obj.get_sondajeEstado_display()}"

                botones_html = f"""
                <div class="form-inline-btn">
                    <form action="/editar_reporte_digital" method="POST">
                        <input type="hidden" name="csrfmiddlewaretoken" value="{csrf_token}">
                        <input type="hidden" name="id" value="{item['id_checklist']}">
                        <input type="hidden" name="origin" value="deleted_report">
                        <button type="submit" class="btn btn-info btn-size-edit">Ver</button>
                    </form>
                </div>
                """

                data.append({
                    "id_checklist": item['id_checklist'],
                    "fecha": fecha_str, # Usamos la fecha convertida
                    "sondaje": sondaje_str,
                    "sonda": item['sonda__sonda'],
                    "creador": item['creador'],
                    "turno": reporte_obj.get_turno_display(),
                    "acciones": botones_html
                })

        return JsonResponse({
            "draw": draw,
            "recordsTotal": total_records,
            "recordsFiltered": records_filtered,
            "data": data
        })

    # --- CARGA INICIAL (HTML Vacio) ---
    context = {
        'sidebar': 'manage_my_digital_reports_deleted',
        'sidebarmain': 'system_drilling', 
    }
    return render(request, 'pages/drilling/manage_mis_reportes_digitales_eliminados.html', context)

@login_required
@sondaje_admin_or_base_datos_or_supervisor_required
def eliminar_reporte_digital_action(request):
    if request.method == 'POST':
        # Detectamos si es una llamada AJAX (desde el script de la tabla)
        is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
        
        try:
            checklist_id = request.POST.get('id')
            checklist = ChecklistMaterialesSonda.objects.get(id=checklist_id)
            reporte = ReportesOperacionales.objects.filter(id_checklist=checklist.id_checklist, status=True).first()

            # --- VALIDACIÓN DE INTEGRIDAD HACIA ADELANTE (ELIMINAR) ---
            if reporte:
                # 1. Filtro Gemelo (Anti-Legacy)
                estado_gemelo = reporte.sondajeEstado
                if not estado_gemelo: 
                    filtro_gemelo = Q(sondajeEstado__isnull=True) | Q(sondajeEstado__exact='')
                else:
                    filtro_gemelo = Q(sondajeEstado=estado_gemelo)

                # 2. Buscar CUALQUIER reporte posterior válido para este mismo pozo
                posteriores = ReportesOperacionales.objects.filter(
                    filtro_gemelo,
                    sondajeCodigo=reporte.sondajeCodigo,
                    sondajeSerie=reporte.sondajeSerie,
                    fechacreacion__gt=reporte.fechacreacion, # Posterior en el tiempo
                    status=True,
                ).exclude(progreso='Eliminado')

                if posteriores.exists():
                    posterior_mas_cercano = posteriores.order_by('fechacreacion').first()
                    fecha_local = localtime(posterior_mas_cercano.fechacreacion)
                    
                    msg = f"No puedes eliminar este reporte intermedio. Existe un reporte posterior con fecha {fecha_local.strftime('%Y-%m-%d %H:%M')} que depende de la continuidad de este pozo."
                    
                    if is_ajax:
                        return JsonResponse({'status': 'error', 'message': msg}, status=400)
                    else:
                        messages.error(request, msg, extra_tags='Integridad de Datos')
                        return redirect(request.META.get('HTTP_REFERER', 'manage_revisar_reportes_digitales'))

            # --- EJECUCIÓN ATÓMICA ---
            with transaction.atomic():
                if reporte:
                    reporte.progreso = 'Eliminado'
                    reporte.save()
                
                ChecklistMaterialesSonda.objects.filter(id_checklist=checklist.id_checklist).update(progreso='Eliminado')

                try:
                    # Llamamos a la función que actualiza SOLO Sondajes (Diario y General)
                    # Esto refrescará el Excel quitando el reporte eliminado
                    task = threading.Thread(target=lambda: GenerateCsv().actualizar_solo_sondajes())
                    task.start()
                    print(f"Trigger Dashboard Sondajes (Eliminación) iniciado para reporte {checklist.id_checklist}")
                except Exception as e:
                    print(f"Error al disparar actualización Dashboard: {e}")

                msg = 'El reporte ha sido eliminado correctamente.'

                if is_ajax:
                    return JsonResponse({'status': 'success', 'message': msg})
                else:
                    messages.success(request, msg)

        except ChecklistMaterialesSonda.DoesNotExist:
             msg = 'El checklist seleccionado no existe.'
             if is_ajax: return JsonResponse({'status': 'error', 'message': msg}, status=404)
             messages.error(request, msg)

        except Exception as e:
            msg = f'Error al eliminar: {str(e)}'
            if is_ajax: return JsonResponse({'status': 'error', 'message': msg}, status=500)
            messages.error(request, msg)
            
        return redirect(request.META.get('HTTP_REFERER', 'manage_revisar_reportes_digitales'))
    else:
        return redirect('manage_revisar_reportes_digitales')
@login_required
@sondaje_admin_or_base_datos_or_supervisor_required
def export_data_view(request):
    if request.method == 'POST':
        form = ExportDataForm(request.POST)
        if form.is_valid():
            start_date = form.cleaned_data['fecha_inicio']
            end_date = form.cleaned_data['fecha_final']

            #Esto hace que el timezone sea consciente (Es decir que tomen en cuenta la hora pais)
            if timezone.is_naive(start_date):
                start_date = timezone.make_aware(start_date)
            if timezone.is_naive(end_date):
                end_date = timezone.make_aware(end_date)
            
            archivo_excel = Workbook() #se crea la instancia excel usando la libreria openpyxl
            if 'Sheet' in archivo_excel.sheetnames: #por defecto se crea con nombre "Sheet"
                del archivo_excel['Sheet'] #se borra para poder crear el nombre que le indiquemos
            
            lista_modelos = [
                (ReportesOperacionales, 'Reporte Digital Operacional'),
                (DetallesPerforaciones, 'Reporte Digital Perforaciones'),
                (ControlesHorarios, 'Reporte Digital Control Horario'),
                (Insumos, 'Reporte Digital Insumos'),
                (LongitudPozos, 'Reporte Digital Longitud'), 
                (DetalleAditivos, 'Reporte Digital Aditivos'), 
                (ObservacionesReportes, 'Reporte Digital Observaciones'), 
                (ChecklistMaterialesSonda, 'Checklist Materiales'), 
            ]
            
            for modelo, sheet_name in lista_modelos: 
                hoja = archivo_excel.create_sheet(title=sheet_name[:30])
                
                if modelo == ReportesOperacionales:
                    headers = [
                        'ID', 'Turno', 'Fecha Creacion', 'Controlador', 'Perforista', 
                        'Sonda', 'Sondaje - N° Sondaje - Gemelo', 'Metro Inicial', 'Metro Final', 
                        'Total Perforado', 'ID Checklist Materiales'
                    ]
                    hoja.append(headers)
                    
                    queryset = modelo.objects.filter(
                        fechacreacion__range=(start_date, end_date),
                        progreso='Aprobado',
                        status=True
                    ).select_related('controlador', 'sonda')

                    for obj in queryset:
                        try:
                            sondaje_compuesto = f"{obj.sondajeCodigo if obj.sondajeCodigo else ''}-{obj.sondajeSerie if obj.sondajeSerie is not None else ''}{obj.get_sondajeEstado_display() if obj.sondajeEstado else ''}"
                            fila = [
                                str(obj.id),
                                obj.get_turno_display() if obj.turno else "",
                                obj.fechacreacion.strftime("%Y-%m-%d %H:%M:%S") if obj.fechacreacion else "",
                                f"{obj.controlador.first_name} {obj.controlador.last_name}" if obj.controlador else "",
                                str(obj.perforista) if obj.perforista else "",
                                str(obj.sonda) if obj.sonda else "",
                                sondaje_compuesto,
                                str(obj.metroInicial) if obj.metroInicial is not None else "",
                                str(obj.metroFinal) if obj.metroFinal is not None else "",
                                str(obj.totalPerforado) if obj.totalPerforado is not None else "",
                                str(obj.id_checklist) if obj.id_checklist is not None else "",
                            ]
                            hoja.append(fila)
                        except Exception: continue

                elif modelo == DetallesPerforaciones:
                    headers = [
                        'ID Reporte Digital Operacional', 'Diametro', 'Perforado', 'Desde', 'Hasta', 
                        'Recuperacion', 'Porcentaje Recuperacion', 'Barra', 
                        'Largo Barra', 'Total Htas', 'Contra', 'Tipo Terreno', 
                        'Orientacion', 'Fecha Creacion'
                    ]
                    hoja.append(headers)
                    queryset = modelo.objects.filter(
                        reporte__fechacreacion__range=(start_date, end_date),
                        reporte__progreso='Aprobado',
                        status=True
                    ).select_related('reporte', 'diametros', 'tipoTerreno', 'orientacion')

                    for obj in queryset:
                        try:
                            fila = [
                                str(obj.reporte.id) if obj.reporte else "",
                                obj.diametros.diametro if obj.diametros else "",
                                str(obj.perforado) if obj.perforado is not None else "",
                                str(obj.desde) if obj.desde is not None else "",
                                str(obj.hasta) if obj.hasta is not None else "",
                                str(obj.recuperacion) if obj.recuperacion is not None else "",
                                f"{obj.porcentajeRecuperacion}%" if obj.porcentajeRecuperacion is not None else "",
                                str(obj.barra) if obj.barra is not None else "",
                                str(obj.largoBarra) if obj.largoBarra is not None else "",
                                str(obj.totalHtas) if obj.totalHtas is not None else "",
                                str(obj.contra) if obj.contra is not None else "",
                                obj.tipoTerreno.tipoTerreno if obj.tipoTerreno else "",
                                obj.orientacion.orientacion if obj.orientacion else "",
                                obj.fechacreacion.strftime("%Y-%m-%d %H:%M:%S") if obj.fechacreacion else "",
                            ]
                            hoja.append(fila)
                        except Exception: continue

                elif modelo == ControlesHorarios:
                    headers = ['ID Reporte Digital Operacional', 'Inicio', 'Final', 'Total', 'Detalle Control Horario', 'Fecha Creacion']
                    hoja.append(headers)
                    queryset = modelo.objects.filter(
                        reporte__fechacreacion__range=(start_date, end_date),
                        reporte__progreso='Aprobado',
                        status=True
                    ).select_related('reporte', 'detalleControlHorario')

                    for obj in queryset:
                        try:
                            fila = [
                                str(obj.reporte.id) if obj.reporte else "",
                                obj.inicio.strftime("%H:%M:%S") if obj.inicio else "",
                                obj.final.strftime("%H:%M:%S") if obj.final else "",
                                obj.total.strftime("%H:%M:%S") if obj.total else "",
                                obj.detalleControlHorario.detalle if obj.detalleControlHorario else "",
                                obj.fechacreacion.strftime("%Y-%m-%d %H:%M:%S") if obj.fechacreacion else "",
                            ]
                            hoja.append(fila)
                        except Exception: continue

                elif modelo == Insumos:
                    headers = ['ID Reporte Digital Operacional', 'Corona', 'Escariador', 'Cantidad Agua', 'Casing', 'Zapata']
                    hoja.append(headers)
                    queryset = modelo.objects.filter(
                        reporte__fechacreacion__range=(start_date, end_date),
                        reporte__progreso='Aprobado',
                        status=True
                    ).select_related('reporte', 'cantidadAgua')

                    for obj in queryset:
                        try:
                            fila = [
                                str(obj.reporte.id) if obj.reporte else "",
                                str(obj.corona) if obj.corona else "",
                                str(obj.escareador) if obj.escareador else "",
                                str(obj.cantidadAgua.cantidadAgua) if obj.cantidadAgua else "",
                                str(obj.casing) if obj.casing else "",
                                str(obj.zapata) if obj.zapata else "",
                            ]
                            hoja.append(fila)
                        except Exception: continue

                elif modelo == LongitudPozos:
                    headers = [
                        'ID Reporte Digital Operacional', 'Largo Barril', 'Punto Muerto', 'Resto Barra', 
                        'Numero Barras', 'Longitud Pozo', 'Hta en Pozo', 'Mts de Hta', 'Profundidad Hta'
                    ]
                    hoja.append(headers)
                    queryset = modelo.objects.filter(
                        reporte__fechacreacion__range=(start_date, end_date),
                        reporte__progreso='Aprobado',
                        status=True
                    ).select_related('reporte')

                    for obj in queryset:
                        try:
                            fila = [
                                str(obj.reporte.id) if obj.reporte else "",
                                str(obj.largoBarril) if obj.largoBarril is not None else "",
                                str(obj.puntoMuerto) if obj.puntoMuerto is not None else "",
                                str(obj.restoBarra) if obj.restoBarra is not None else "",
                                str(obj.numeroBarras) if obj.numeroBarras is not None else "",
                                str(obj.longitudPozo) if obj.longitudPozo is not None else "",
                                str(obj.htaEnPozo) if obj.htaEnPozo else "",
                                str(obj.mtsDeHta) if obj.mtsDeHta is not None else "",
                                str(obj.profundidadHta) if obj.profundidadHta is not None else "",
                            ]
                            hoja.append(fila)
                        except Exception: continue

                elif modelo == DetalleAditivos:
                    headers = ['ID Reporte Digital Operacional', 'Aditivo', 'Cantidad', 'Fecha Creacion']
                    hoja.append(headers)
                    queryset = modelo.objects.filter(
                        reporte__fechacreacion__range=(start_date, end_date),
                        reporte__progreso='Aprobado',
                        status=True
                    ).select_related('reporte', 'aditivo')

                    for obj in queryset:
                        try:
                            fila = [
                                str(obj.reporte.id) if obj.reporte else "",
                                obj.aditivo.aditivo if obj.aditivo else "",
                                str(obj.cantidad) if obj.cantidad is not None else "",
                                obj.fechacreacion.strftime("%Y-%m-%d %H:%M:%S") if obj.fechacreacion else "",
                            ]
                            hoja.append(fila)
                        except Exception: continue

                elif modelo == ObservacionesReportes:
                    headers = ['ID Reporte Digital Operacional', 'Fecha Creacion', 'Observaciones']
                    hoja.append(headers)
                    queryset = modelo.objects.filter(
                        reporte__fechacreacion__range=(start_date, end_date),
                        reporte__progreso='Aprobado',
                        status=True
                    ).select_related('reporte')

                    for obj in queryset:
                        try:
                            fila = [
                                str(obj.reporte.id) if obj.reporte else "",
                                obj.reporte.fechacreacion.strftime("%Y-%m-%d %H:%M:%S") if obj.reporte.fechacreacion else "",
                                str(obj.observaciones) if obj.observaciones else "",
                            ]
                            hoja.append(fila)
                        except Exception: continue

                elif modelo == ChecklistMaterialesSonda:
                    headers = [
                        'ID', 'Item', 'Cantidad', 'Creador', 'Jornada', 'Etapa', 
                        'Turno', 'Sonda', 'Sondaje - N° Sondaje - Gemelo', 'Fecha Creacion'
                    ]
                    hoja.append(headers)
                    
                    queryset = modelo.objects.filter(
                        fechacreacion__range=(start_date, end_date),
                        progreso='Aprobado',
                        status=True
                    ).select_related('sonda', 'item')

                    for obj in queryset:
                        try:
                            sondaje_compuesto = f"{obj.sondajeCodigo if obj.sondajeCodigo else ''}-{obj.sondajeSerie if obj.sondajeSerie is not None else ''}{obj.get_sondajeEstado_display() if obj.sondajeEstado else ''}"
                            fila = [
                                str(obj.id_checklist) if obj.id_checklist is not None else "",
                                obj.item.material if obj.item else "",
                                str(obj.cantidad) if obj.cantidad is not None else "",
                                str(obj.creador) if obj.creador else "",
                                obj.get_jornada_display() if obj.jornada else "",
                                str(obj.etapa) if obj.etapa else "",
                                obj.get_turno_display() if obj.turno else "",
                                obj.sonda.sonda if obj.sonda else "",
                                sondaje_compuesto,
                                obj.fechacreacion.strftime("%Y-%m-%d %H:%M:%S") if obj.fechacreacion else "",
                            ]
                            hoja.append(fila)
                        except Exception: continue

                else: 
                    campos = modelo._meta.fields 
                    headers = [campo.attname for campo in campos] 
                    hoja.append(headers)
                    
                    queryset = modelo.objects.all() 
                    if hasattr(modelo, 'fechacreacion'): 
                        queryset = queryset.filter(fechacreacion__range=(start_date, end_date))
                    elif hasattr(modelo, 'reporte'):
                        queryset = queryset.filter(reporte__fechacreacion__range=(start_date, end_date))
                    
                    for objeto in queryset: 
                        fila = []
                        for campo in campos:
                            valor = getattr(objeto, campo.attname) 
                            if valor is None:
                                valor = ""
                            else:
                                valor = str(valor)
                            fila.append(valor)
                        hoja.append(fila)
            
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=DrillingData.xlsx'
            archivo_excel.save(response)
            return response
    else:
        form = ExportDataForm()

    return render(request, 'pages/drilling/export_data.html', {'form': form, 'sidebar': 'export_data', 'sidebarmain': 'system_report_drilling'})

@login_required
@sondaje_admin_or_base_datos_or_supervisor_required
def get_ultima_sonda_por_pozo(request):
    pozo_string = request.GET.get('pozo_nombre')
    
    if pozo_string:
        try:
            partes = pozo_string.rsplit('-', 1)
            
            if len(partes) < 2:
                return JsonResponse({'sonda_id': None})

            codigo_sondaje = partes[0]
            serie_gemelo_str = partes[1]
            gemelo_letra = None
            serie_sondaje = serie_gemelo_str

            if serie_gemelo_str and not serie_gemelo_str[-1].isdigit():
                gemelo_letra = serie_gemelo_str[-1]
                serie_sondaje = serie_gemelo_str[:-1]

            filtros = {
                'sondajeCodigo__sondaje': codigo_sondaje,
                'sondajeSerie': serie_sondaje,
                'progreso': 'Aprobado'
            }
            
            if gemelo_letra:
                mapa_gemelo = {label: value for value, label in gemelo}
                gemelo_id = mapa_gemelo.get(gemelo_letra)
                
                if gemelo_id:
                    filtros['sondajeEstado'] = gemelo_id
                else:
                    return JsonResponse({'sonda_id': None})
            else:
                filtros['sondajeEstado__in'] = ['', None]

            ultimo_reporte = ReportesOperacionales.objects.filter(**filtros).order_by('-fechacreacion').first()
            
            if ultimo_reporte and ultimo_reporte.sonda:
                return JsonResponse({
                    'sonda_id': ultimo_reporte.sonda.id, 
                    'sonda_nombre': str(ultimo_reporte.sonda)
                })
                
        except Exception as e:
            print(f"Error parseando pozo: {e}")
            pass
            
    return JsonResponse({'sonda_id': None})