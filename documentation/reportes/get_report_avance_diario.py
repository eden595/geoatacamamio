from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import locale
import pytz
import requests 
from pathlib import Path
from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn
from collections import defaultdict
from openpyxl.drawing.image import Image
from openpyxl.formatting.rule import CellIsRule
from pprint import pprint
from openpyxl import load_workbook
from ..reportes import helpers
import copy

def agregar_datos_a_celdas(hoja, inicio_celda, datos, primer_color, segundo_color):
    fila_inicio = int(inicio_celda[1:])

    color_fondo_1 = PatternFill(start_color=primer_color, end_color=primer_color, fill_type="solid")
    color_fondo_2 = PatternFill(start_color=segundo_color, end_color=segundo_color, fill_type="solid")

    # Definir el borde 
    borde = Border(
        bottom=Side(border_style="thick", color="ed7d31")  # Borde inferior Remarcado
    )


    ultima_fila = fila_inicio
    sonda_actual = None

    mapeo_columnas = {
        "N°": "A",
        "RECOMENDACIÓN": "B",
        "SONDA": "C",
        "SONDAJE": "D",
        "TURNO": "E",
        "DESDE": "F",
        "HASTA": "G",
        "PERFORADO": "H",
        "OBSERVACIONES": "I"
    }

    last_dict = None
    for idx, diccionario in enumerate(datos):
        fila_actual = fila_inicio + idx
        ultima_fila = fila_actual
        color_fondo = color_fondo_1 if idx % 2 == 0 else color_fondo_2
        current_dict = f'{diccionario["RECOMENDACIÓN"]}{diccionario["SONDA"]}{diccionario["SONDAJE"]}'
        if current_dict != last_dict:
            last_dict = current_dict
        else:
            diccionario["RECOMENDACIÓN"] = ""
            diccionario["SONDA"] = ""
            diccionario["SONDAJE"] = ""

        for key, value in diccionario.items():
            
            if key in mapeo_columnas:
                columna = mapeo_columnas[key]
                celda = hoja[f"{columna}{fila_actual}"]
                celda.value = value
                celda.font = Font(name="Arial", size=8, bold=False)
                celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                celda.fill = color_fondo

                # Si idx es impar, agregar borde
                if idx % 2 == 1:
                    celda.border = borde


                hoja.row_dimensions[fila_actual].height = 45 

    return hoja, ultima_fila

def orden_clave(item):
    fecha = datetime.strptime(item['Fecha'], "%d/%m/%Y")
    turno_valor = 0 if item['TURNO'].lower() == 'A' else 1
    return (item['SONDA'], fecha, turno_valor, item['DESDE'])

def construir_data(reportes_agrupados,nombres_programas):
    resultado = []
    for reporte, detalles in reportes_agrupados.items():

        for detalle in detalles:

            if 'detalle_perforaciones' not in detalle:
                continue
            for perforacion in detalle['detalle_perforaciones']:

                desde = float(perforacion['DESDE'])
                hasta = float(perforacion['HASTA'])
                total_dia = hasta - desde

                resultado.append({
                    
                    "RECOMENDACIÓN": detalle['recomendacion']['recomendacion'],
                    "SONDA": detalle["nombre_sonda"],
                    "SONDAJE": detalle['nombre_sondaje'],
                    "TURNO": perforacion['TURNO'],
                    "DESDE": desde,
                    "HASTA": hasta,
                    "PERFORADO": total_dia,
                    "OBSERVACIONES": "SIN DATO",
                    "Fecha": perforacion['FECHA']
                })

    
    # Ordenar la lista
    datos_ordenados = sorted(resultado, key=orden_clave)

    tz_chile = pytz.timezone("America/Santiago")
    hoy_chile = datetime.now(tz_chile)
    fecha_hoy = f"{hoy_chile.day}/{hoy_chile.month}/{hoy_chile.year}"


    resultado_dia_actual = []
    for dato in datos_ordenados:
        # Copiar y eliminar 'Fecha' si coincide
        indice = 1
        dato_copia = dato.copy()
        if dato['Fecha'] == fecha_hoy:
            del dato_copia['Fecha']
            dato_copia['N°'] = indice

            resultado_dia_actual.append(dato_copia)
            indice += 1


    if not resultado_dia_actual:
         resultado_dia_actual = [
            {
            'N°': 1,
            'DESDE': 0.0,
            'Fecha': f'{fecha_hoy}',
            'HASTA': 0.0,
            'OBSERVACIONES': 'SIN DATOS PARA MOSTRAR',
            'PERFORADO': 0.0,
            'RECOMENDACIÓN': 'SIN DATO',
            'SONDA': 'SIN DATO ',
            'SONDAJE': 'SIN DATO',
            'TURNO': 'SIN DATO'},
        ]
    return resultado_dia_actual,hoy_chile.year
def run(libro,reportes_agrupados,nombres_programas):

    data, anio_actual = construir_data(reportes_agrupados,nombres_programas)

    # Agregar una nueva hoja 
    hoja = libro.create_sheet(title=f"AVANCE DIARIO_{anio_actual}")
    hoja = helpers.obtener_fecha_documento(hoja, "I1",None,"right")


    # Agregamos primer titulo
    hoja = helpers.agregar_titulo(hoja, "A3", "A3:F3", "c. PLANIFICACIÓN TOTAL VS AVANCE REAL CONTRATO ACTUAL 2023 - 2025")

    # Aplicar el color negro a las celdas de A4 a I35
    for fila in range(4, 36):  # Filas de 4 a 35
        for col in range(1, 10):  # Columnas A (1) a I (9)
            hoja.cell(row=fila, column=col).fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Agregamos Segundo titulo
    hoja = helpers.agregar_titulo(hoja, "A37", "A37:C37", "d. INFORMACIÓN DIARIA DE PERFORACIÓN")

    titulos = ["N°", "RECOMENDACIÓN", "SONDA", "SONDAJE", "TURNO", "DESDE", "HASTA", "PERFORADO", "OBSERVACIONES"]

    # Agregamos Headers a tabla
    hoja = helpers.agregar_titulos(hoja, "A", 38, titulos,color_fondo="ed7d31", alto_fila_factor=1, ancho_columna=15)
    # Agregamos data a la tabla
    hoja, ultima_fila = agregar_datos_a_celdas(hoja, "A39", data,"fbe4d5","b4c6e7")

    # Seteamos anchos de las columnas 
    hoja.column_dimensions["A"].width = 5
    hoja.column_dimensions["B"].width = 15
    hoja.column_dimensions["C"].width = 8
    hoja.column_dimensions["D"].width = 15 
    hoja.column_dimensions["E"].width = 8
    hoja.column_dimensions["F"].width = 10
    hoja.column_dimensions["G"].width = 8
    hoja.column_dimensions["H"].width = 10
    hoja.column_dimensions["I"].width = 80


    # Mover la nueva hoja a la primera posición
    libro._sheets.insert(0, libro._sheets.pop(-1))
    return libro, ultima_fila, f"AVANCE DIARIO_{anio_actual}"