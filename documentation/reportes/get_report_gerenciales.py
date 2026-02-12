from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import locale
import pytz
import requests 
from pathlib import Path
from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn
from collections import defaultdict
from openpyxl.drawing.image import Image
from openpyxl.formatting.rule import CellIsRule
from pprint import pprint
from openpyxl import load_workbook
from documentation.reportes import helpers
import copy

def agregar_datos_a_celdas_resumen(hoja, inicio_celda, datos, primer_color, segundo_color):
    # Obtener la fila de inicio desde la celda (Ej: "A1" -> 1)
    fila_inicio = int(inicio_celda[1:])

    # Definir los colores alternados
    colores = [PatternFill(start_color=primer_color, end_color=primer_color, fill_type="solid"),
               PatternFill(start_color=segundo_color, end_color=segundo_color, fill_type="solid")]

    # Mapeo de claves a columnas
    mapeo_columnas = {
        "PROGRAMA": "A",
        "REC": "B",
        "SONDA": "C",
        "SONDAJE": "D",
        "UBICACION": "E",
        "PROFUNDIDAD PROGRAMADA (m)": "F",
        "PERFORACIÓN AVANCE (m)": "G",
        "AVANCE DÍA (m)": "H",
        "POR PERFORAR (m)": "I",
        "% AVANCE": "J",
        "OBSERVACIÓN": "K"
    }

    if not datos:
        return hoja, fila_inicio - 1  # <- Manejo explícito del caso vacío
    
    # Iterar sobre los datos (lista de diccionarios)
    for idx, diccionario in enumerate(datos):
        fila_actual = fila_inicio + idx
        color_fondo = colores[idx % 2]  # Alternar colores

        for key, value in diccionario.items():
            columna = mapeo_columnas.get(key)  # Obtener la columna asignada
            if not columna:
                continue  # Ignorar claves no mapeadas
            
            celda = hoja[f"{columna}{fila_actual}"]
            celda.value = value

            # Si es número, aplicar formato de dos decimales
            if isinstance(value, (int, float)):
                celda.number_format = '#,##0.00'
                
            # Aplicar formato a la celda
            celda.font = Font(name="Arial", size=8, bold=True)
            celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            celda.fill = color_fondo

    return hoja, fila_actual  # Retorna la hoja y la última fila usada

def agregar_datos_a_celdas_avance_contrato(hoja, inicio_celda, datos, primer_color, segundo_color,mes_actual):
    # Obtener la fila de inicio desde la celda (Ej: "A1" -> 1)
    fila_inicio = int(inicio_celda[1:])

    # Definir los colores alternados
    colores = [
        PatternFill(start_color=primer_color, end_color=primer_color, fill_type="solid"),
        PatternFill(start_color=segundo_color, end_color=segundo_color, fill_type="solid")
    ]


    # Definir el borde 
    borde = Border(
        left=Side(border_style="thin", color="ed7d31"),  # Borde izquierdo
        right=Side(border_style="thin", color="ed7d31"),  # Borde derecho
        top=Side(border_style="thin", color="ed7d31"),  # Borde superior
        bottom=Side(border_style="thin", color="ed7d31")  # Borde inferior
    )

    fila_actual = fila_inicio  # Variable para rastrear la fila actual

    num_merge_cell = 0
    # Iterar sobre las campañas
    for idx, (campaña, datos_campaña) in enumerate(datos.items()):
        color_fondo = colores[idx % 2]  # Alternar colores según el índice

        # Primero hacemos merge pr el largo de los programas
        fin_merge = fila_actual + len(datos_campaña["Programas"]) - 1
        ## Merge de Campaña
        rango = f"A{fila_actual}:A{fin_merge}"
        hoja.merge_cells(rango)
        ## Merge de Metros Planificados
        rango = f"B{fila_actual}:B{fin_merge}"
        hoja.merge_cells(rango)
        ## Merge de Avance Campaña
        rango = f"F{fila_actual}:F{fin_merge}"
        hoja.merge_cells(rango)
        ## Merge de % Avance programa
        rango = f"G{fila_actual}:G{fin_merge}"
        hoja.merge_cells(rango)
        ## Merge de % Avance Campaña
        rango = f"H{fila_actual}:H{fin_merge}"
        hoja.merge_cells(rango)

        # Escribir nombre dela campaña
        hoja[f"A{fila_actual}"].value = campaña
        hoja[f"A{fila_actual}"].font = Font(name="Arial", size=9, bold=True)
        hoja[f"A{fila_actual}"].alignment = Alignment(horizontal="center", vertical="center")
        hoja[f"A{fila_actual}"].fill = color_fondo  # Aplica el color de fondo
        hoja[f"A{fila_actual}"].border = borde # Aplicar el borde a la celda

        # Escribir metros Planificados
        hoja[f"B{fila_actual}"].value = datos_campaña['Metros Planificados']
        hoja[f"B{fila_actual}"].font = Font(name="Arial", size=9, bold=True)
        hoja[f"B{fila_actual}"].alignment = Alignment(horizontal="center", vertical="center")
        hoja[f"B{fila_actual}"].fill = color_fondo  # Aplica el color de fondo
        hoja[f"B{fila_actual}"].border = borde # Aplicar el borde a la celda


        avance_campaña = 0
        metros_programados = 0
        current_row = fila_actual

        # Escribir los datos de cada programa dentro de la campaña
        for idx, programa in enumerate(datos_campaña["Programas"]):

            for j, (key, value) in enumerate(programa.items()):
                columna = "C"
                if key == "Nombre":
                    columna = "C"
                if key == f"Metros programados a {mes_actual}":
                    columna = "D"
                    metros_programados += value
                if key == "Metros de avance por programa":
                    columna = "E"
                    avance_campaña += value
                
                celda = hoja[f"{columna}{current_row}"]
                celda.value = value

                # Aplicar formato a la celda
                celda.font = Font(name="Arial", size=8, bold=True)
                celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                celda.fill = color_fondo
                celda.border = borde
            
            current_row += 1  # Avanzar a la siguiente fila

        # Escribir Avance de Campaña
        hoja[f"F{fila_actual}"].value = avance_campaña
        hoja[f"F{fila_actual}"].font = Font(name="Arial", size=9, bold=True)
        hoja[f"F{fila_actual}"].alignment = Alignment(horizontal="center", vertical="center")
        hoja[f"F{fila_actual}"].fill = color_fondo  # Aplica el color de fondo
        hoja[f"F{fila_actual}"].border = borde # Aplicar el borde a la celda

        # Escribir porcentaje Avance de Campaña

        if metros_programados != 0:

            g_resultado = (avance_campaña / metros_programados)*1

        else:
            g_resultado = float(0.00)


        hoja[f"G{fila_actual}"].value = g_resultado
        hoja[f"G{fila_actual}"].number_format = '0%' 
        hoja[f"G{fila_actual}"].font = Font(name="Arial", size=9, bold=True)
        hoja[f"G{fila_actual}"].alignment = Alignment(horizontal="center", vertical="center")
        hoja[f"G{fila_actual}"].fill = color_fondo  # Aplica el color de fondo
        hoja[f"G{fila_actual}"].border = borde # Aplicar el borde a la celda

        # Escribir porcentaje Avance de Campaña
        hoja[f"H{fila_actual}"].value = (avance_campaña / datos_campaña['Metros Planificados'])*1
        hoja[f"H{fila_actual}"].number_format = '0%' 
        hoja[f"H{fila_actual}"].font = Font(name="Arial", size=9, bold=True)
        hoja[f"H{fila_actual}"].alignment = Alignment(horizontal="center", vertical="center")
        hoja[f"H{fila_actual}"].fill = color_fondo  # Aplica el color de fondo
        hoja[f"H{fila_actual}"].border = borde # Aplicar el borde a la celda

        fila_actual = fin_merge + 1
    return hoja, fila_actual  # Retorna la hoja y la última fila usada

def maqueta_segunda_tabla(programas,campanas,planificacion_programas,anio_final, mes_final):
    meses = {
        1: 'Enero',
        2: 'Febrero',
        3: 'Marzo',
        4: 'Abril',
        5: 'Mayo',
        6: 'Junio',
        7: 'Julio',
        8: 'Agosto',
        9: 'Septiembre',
        10: 'Octubre',
        11: 'Noviembre',
        12: 'Diciembre'
    }
    segunda_tabla = {}

    mes_final = int(mes_final)
    for c in campanas:
        if c['campana'] not in segunda_tabla:

            todos_los_programas = {}
            for p in programas:

                if c['id'] == p['campana']:
                    
                    if p['programa'] not in todos_los_programas:

                        data_filtrada = [item for item in planificacion_programas if item["campana"] == c['id'] and item["programa"] == p['id']]
                        datos_filtrados = [
                            item for item in data_filtrada
                            if not (int(item['ano']) > int(anio_final) or (int(item['ano']) == int(anio_final) and int(item['mes']) > int(mes_final)))
                        ]
                        datos_ordenados = sorted(
                            datos_filtrados,
                            key=lambda x: (int(x['ano']), int(x['mes']))
                        )
                        suma_plan = sum(float(item['plan']) for item in datos_ordenados if item['plan'] is not None)

                        #falta agregar los fatos a la segunda tabla, datos para el acumulado para el mes actual
                        todos_los_programas[p['programa']] = {
                            "Nombre": p['programa'],
                            f"Metros programados a {meses[int(mes_final)]}": float(suma_plan),
                            "Metros de avance por programa": float(0.00)
                        }



            segunda_tabla[c['campana']] = {
                "Metros Planificados": c['metros'],
                "Programas": [todos_los_programas]
            }
     

    return segunda_tabla, meses[int(mes_final)]

def limpiar_segunda_table(segunda_tabla):
    # Transformación
    resultado = {}

    for area, contenido in segunda_tabla.items():
        metros_planificados = contenido['Metros Planificados']
        programas = []
        for prog_dict in contenido['Programas']:
            for _, prog_data in prog_dict.items():
                programas.append(prog_data)
        resultado[area] = {
            'Metros Planificados': metros_planificados,
            'Programas': programas
        }

    return resultado
def contruir_data_seguna_tabla(resultado,programas,campanas,planificacion_programas,anio_final, mes_final):

    segunda_tabla , mes_actual= maqueta_segunda_tabla(programas,campanas,planificacion_programas,anio_final, mes_final)


    for r in resultado:

        programas = []
        for programa in segunda_tabla[r['CAMPAÑA']]['Programas']:

            for p , detalle in programa.items():
                if p == r['PROGRAMA']:
                    detalle['Metros de avance por programa'] += float(r['PERFORACIÓN AVANCE (m)'])

            programas.append(programa)  

        segunda_tabla[r['CAMPAÑA']]['Programas'] = programas 

    segunda_tabla = limpiar_segunda_table(segunda_tabla)

    return segunda_tabla, mes_actual
def construir_data(reportes_agrupados,programas,campanas,planificacion_programas,anio_final, mes_final):

    resultado = []

    for reporte, detalles in reportes_agrupados.items():
        for detalle in detalles:
            campana = detalle['recomendacion']['campana']
            porcentaje_avance= float(0.00)
            observacion = 'Sin observación'
            programa = detalle['recomendacion']['programa']
            perforacion_avance = float(0.00)
            por_perforar = float(0.00)
            profundidad_programada = float(0.00)

            recomendacion = detalle['recomendacion']['recomendacion']
            sondaje = detalle['recomendacion']['pozo']
            sonda = detalle["nombre_sonda"]
            ubicacion = detalle['recomendacion']['sector']
            

            if 'detalle_perforaciones' in detalle:

                for perforacion in detalle['detalle_perforaciones']:
                    desde = float(perforacion['DESDE'])
                    hasta = float(perforacion['HASTA'])
                    total_dia =  hasta-desde 
                    largo_programado = float(detalle['recomendacion']['largo_programado'])
                    por_perforar = largo_programado - hasta
                    perforacion_avance = largo_programado - por_perforar

                    if por_perforar < 0:
                        avance_actual = float(100.00)

                    elif por_perforar != 0.0 and hasta != 0.0:
                        avance_actual = (hasta / por_perforar) * 100

                    else:
                        avance_actual = float(0.00)
                        
                    resultado.append({
                        'CAMPAÑA':campana,
                        '% AVANCE': avance_actual,
                        'AVANCE DÍA (m)': total_dia,
                        'OBSERVACIÓN': observacion,
                        'PERFORACIÓN AVANCE (m)': perforacion_avance,
                        'POR PERFORAR (m)': por_perforar,
                        'PROFUNDIDAD PROGRAMADA (m)': largo_programado,
                        'PROGRAMA': programa,
                        'REC': recomendacion,
                        'SONDA': sonda,
                        'SONDAJE': sondaje,
                        'UBICACION': ubicacion
                    })
            else:
                resultado.append({
                        'CAMPAÑA':campana,
                        '% AVANCE': porcentaje_avance,
                        'AVANCE DÍA (m)': float(0.00),
                        'OBSERVACIÓN': observacion,
                        'PERFORACIÓN AVANCE (m)': perforacion_avance,
                        'POR PERFORAR (m)': por_perforar,
                        'PROFUNDIDAD PROGRAMADA (m)': profundidad_programada,
                        'PROGRAMA': programa,
                        'REC': recomendacion,
                        'SONDA': sonda,
                        'SONDAJE': sondaje,
                        'UBICACION': ubicacion
                    })
                
    # Diccionario para guardar el registro con menor 'POR PERFORAR (m)' por REC
    min_por_perforar_por_rec = {}   

    for record in resultado:
        rec = record['REC']
        por_perforar = record['POR PERFORAR (m)']
        # Si el REC no está aún guardado o si encontramos un valor menor
        if rec not in min_por_perforar_por_rec or por_perforar < min_por_perforar_por_rec[rec]['POR PERFORAR (m)']:
            min_por_perforar_por_rec[rec] = record
    # Convertimos el resultado a una lista
    result = list(min_por_perforar_por_rec.values())

    data_seguna_tabla, mes_actual = contruir_data_seguna_tabla(result,programas,campanas,planificacion_programas,anio_final, mes_final)
    return result,data_seguna_tabla, mes_actual
def run(libro,reportes_agrupados,programas,campanas,planificacion_programas,anio_final, mes_final):

    data,data_seguna_tabla, mes_actual = construir_data(reportes_agrupados,programas,campanas,planificacion_programas,anio_final, mes_final)

    # Agregar una nueva hoja "INF. GERENCIAL (a)"

    hoja = libro.create_sheet(title="INF. GERENCIAL (a)")
 
    hoja = helpers.obtener_fecha_documento(hoja, "E1",merge_cells = "E1:G1")

    hoja = helpers.agregar_titulo(hoja, "A4", "A4:J4", "a. RESUMEN GERENCIAL")

    # PRIMERA TABLA
    titulos = ["PROGRAMA", "REC", "SONDA", "SONDAJE", "UBICACIÓN", "PROFUNDIDAD PROGRAMADA (m)","PERFORACIÓN AVANCE (m)",
    "AVANCE DÍA (m)","POR PERFORAR (m)","% AVANCE","OBSERVACIÓN"]
    hoja = helpers.agregar_titulos(hoja, "A", 6, titulos, "ed7d31", alto_fila_factor=2)

    hoja, ultima_fila = agregar_datos_a_celdas_resumen(hoja, "A7", data, "fbe4d5","b4c6e7")

    # SEGUNDA TABLA
    ## Valores necesarios para poder crear la segunda tabla  independiente de cuanto crezca la primera tabla
    celda_inicial_numero = int(ultima_fila) + 4
    celda_inicial_tabla = celda_inicial_numero+2
    celda_inicial = f'A{celda_inicial_numero}'

    hoja = helpers.agregar_titulo(hoja,celda_inicial,f'A{celda_inicial_numero}:B{celda_inicial_numero}',"b. AVANCE CONTRATO ACTUAL 2023/2025")

    titulos = ["Campaña", "Metros Planificados", "Programas", f"Metros programados a {mes_actual}", "Metros de avance por programa", "Avance campaña","% Avance programa","% Avance Campaña"]
    hoja = helpers.agregar_titulos(hoja, "A", int(celda_inicial_tabla), titulos, "ed7d31", alto_fila_factor=2)
    # Mover la nueva hoja a la primera posición
    hoja, ultima_fila = agregar_datos_a_celdas_avance_contrato(hoja, f"A{celda_inicial_tabla+1}", data_seguna_tabla, "fbe4d5","b4c6e7",mes_actual)
    
    
    libro._sheets.insert(0, libro._sheets.pop(-1))
    celda_inicial_tabla = f"A{celda_inicial_tabla+1}"


    return libro, ultima_fila, celda_inicial_tabla,data,data_seguna_tabla, mes_actual