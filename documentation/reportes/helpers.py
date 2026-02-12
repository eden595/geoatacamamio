from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side,PatternFill
from datetime import datetime
from openpyxl.utils import get_column_letter,column_index_from_string
import locale
import pytz

from pprint import pprint
from openpyxl.chart import LineChart, Reference, BarChart,PieChart
from collections import defaultdict
import calendar
from datetime import datetime
from collections import OrderedDict

def agregar_subtitulos(hoja, inicio_columna, fila, titulos,color_fondo="70AD47", alto_fila_factor=1, ancho_columna=8, tipo = None ):


    # Obtener el índice numérico de la columna de inicio
    col_index = ord(inicio_columna.upper()) - ord('A') + 1  

    # Ajustar el alto de la fila (doble del predeterminado de 15)
    hoja.row_dimensions[fila].height = 15 * alto_fila_factor

    # Estilos
    estilo_encabezado = Font(color="000000", size=8, bold=True)  # Texto Negro, negrita
    alineacion_centrada = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fondo_color = PatternFill(start_color=color_fondo, end_color=color_fondo, fill_type="solid")

    for i, titulo in enumerate(titulos):
        col_letra = get_column_letter(col_index + i)
        celda = hoja[f"{col_letra}{fila}"]
        celda.value = titulo
        celda.font = estilo_encabezado
        celda.alignment = alineacion_centrada
        celda.fill = fondo_color
        # Ajustar el ancho de la columna
        hoja.column_dimensions[col_letra].width = ancho_columna
    return hoja

def agregar_titulos(hoja, inicio_columna, fila, titulos,color_fondo="70AD47", alto_fila_factor=1, ancho_columna=15, tipo = None ,color_letras = None):


    # Obtener el índice numérico de la columna de inicio
    col_index = column_index_from_string(inicio_columna.upper())

    # Ajustar el alto de la fila (doble del predeterminado de 15)
    hoja.row_dimensions[fila].height = 15 * alto_fila_factor

    estilo_encabezado = Font(color="FFFFFF", size=8, bold=True)  # Texto blanco, negrita
    # Estilos
    if color_letras:
        estilo_encabezado = Font(size=8, bold=True)
    alineacion_centrada = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fondo_color = PatternFill(start_color=color_fondo, end_color=color_fondo, fill_type="solid")

    inicio_tabla = None
    termino_tabla = None
    if tipo == "avance_programa":
        total_titulos = len(titulos)  # No es necesario restar 1
        subtitulos = ["Plan","Real mensual","Acumulado plan","Acumulado real"]
        # Insertar los títulos
        last_cell = 2  # Esta variable almacenará la última celda procesada
        # Insertar los títulos
        for i, titulo in enumerate(titulos):
            if i == 0:  # El primer título en una celda
                col_letra = get_column_letter(col_index + i)
                celda = hoja[f"{col_letra}{fila}"]
                celda.value = titulo
                celda.font = estilo_encabezado
                celda.alignment = alineacion_centrada
                celda.fill = fondo_color
                hoja.column_dimensions[col_letra].width = ancho_columna
                agregar_subtitulos(hoja, col_letra, fila+1, [""], color_fondo="fcc000", alto_fila_factor=2)
                inicio_tabla = f"{col_letra}{fila}"

            else:
                col_letra_inicio = get_column_letter(last_cell)
                merge_inicio = f"{col_letra_inicio}{fila}"
                col_letra_final = get_column_letter(last_cell+3)
                merge_final = f"{col_letra_final}{fila}"
                # Realizar el merge
                hoja.merge_cells(f"{merge_inicio}:{merge_final}")
                last_cell += 4
                # se agregan datos
                celda = hoja[merge_inicio]
                celda.value = titulo
                celda.font = estilo_encabezado
                celda.alignment = alineacion_centrada
                celda.fill = fondo_color

                agregar_subtitulos(hoja,col_letra_inicio, fila+1, subtitulos, color_fondo="fcc000", alto_fila_factor=2)
                termino_tabla = merge_final




        return hoja, inicio_tabla, termino_tabla


    elif tipo == "detalle_sonda_dia":
        # Orden correcto esperado
        orden_esperado = [
            'Día/Sonda',
            'Cantidad de Sondas',
            'Mts. Día',
            'Mts. Proyectados a Perforar',
            'Rendimiento Día (m)',
            'Rendimiento Proyectado GEOTEC Día (m)',
            'Rendimiento por Hora (m)',
        ]

        # Titulos ordenados según el orden esperado
        titulos_ordenados = [titulo for titulo in orden_esperado if titulo in titulos]

        # Títulos desconocidos (no están en el orden esperado)
        titulos_desconocidos = list(set(titulos) - set(orden_esperado))

        if titulos_desconocidos:
            # 1. Separar listas
            titulos_con_parentesis = [t for t in titulos_desconocidos if "(" in t]
            titulos_sin_parentesis = [t for t in titulos_desconocidos if "(" not in t]
            # 2. Ordenar Z -> A la lista sin paréntesis
            titulos_sin_parentesis.sort(reverse=True)
            # 3. Buscar coincidencias 
            for titulo_sin in titulos_sin_parentesis:
                for titulo_con in titulos_con_parentesis:
                    if titulo_sin in titulo_con:
                        #print(f"'{titulo_sin}' → encontrado en: '{titulo_con}'")
                        if titulo_con not in titulos_ordenados:
                            titulos_ordenados.insert(1, titulo_con)
                            titulos_ordenados.insert(2, titulo_sin)  # Insertar en la posición 1
                        break
                else:
                    print(f"'{titulo_sin}' → NO encontrado")
                    exit()


        # Verificar si los t�tulos est�n en el orden esperado   

        for i, titulo in enumerate(titulos_ordenados):


 
            if ("-" in titulo and "(" not in titulo) or "/" in titulo:
                estilo_encabezado = Font(color="FFFFFF", size=8, bold=True)  # Texto blanco, negrita
            else:
                estilo_encabezado = Font(color="000000", size=8, bold=True)
            col_letra = get_column_letter(col_index + i)
            celda = hoja[f"{col_letra}{fila}"]
            celda.value = titulo
            celda.font = estilo_encabezado
            celda.alignment = alineacion_centrada
            celda.fill = fondo_color
            # Ajustar el ancho de la columna
            hoja.column_dimensions[col_letra].width = ancho_columna
        return hoja,titulos_ordenados
    # Insertar los títulos
   
    elif tipo == "muestrera":

        for i, titulo in enumerate(titulos):
            col_letra = get_column_letter(col_index + i)
            celda = hoja[f"{col_letra}{fila}"]


            celda.value = titulo
            celda.font = estilo_encabezado
            if col_letra in ["U", "V","W","X","AA","AB","AC","AD"]:   
                celda.font = Font(name="Arial", size=8, bold=True)
            celda.alignment = alineacion_centrada
            celda.fill = fondo_color
            # Ajustar el ancho de la columna
            hoja.column_dimensions[col_letra].width = ancho_columna
        return hoja
    else:
        for i, titulo in enumerate(titulos):
            col_letra = get_column_letter(col_index + i)
            celda = hoja[f"{col_letra}{fila}"]
            celda.value = titulo
            celda.font = estilo_encabezado
            celda.alignment = alineacion_centrada
            celda.fill = fondo_color
            # Ajustar el ancho de la columna
            hoja.column_dimensions[col_letra].width = ancho_columna
        return hoja

def agregar_titulo(hoja, celda_inicial, celdas_merge, value):

    hoja.merge_cells(celdas_merge)
    celda = hoja[celda_inicial]
    celda.value = value
    celda.font = Font(bold=True, size=8)
    celda.alignment = Alignment(horizontal="left", vertical="center")
    return hoja

def obtener_fecha_documento(hoja, value_cell,merge_cells = None,horizontal = None, background = None):
    # Establecer el idioma a español
    if not horizontal:
        horizontal = "left"
    locale.setlocale(locale.LC_TIME, "es_ES.UTF-8")
    # Obtener la hora actual en la zona horaria de Chile
    zona_chile = pytz.timezone("America/Santiago")
    fecha_actual = datetime.now(zona_chile)
    # Formatear la fecha actual
    fecha_formateada = fecha_actual.strftime("%A, %B %d, %Y").lower()

    if merge_cells:
        # Combinar las celdas de A1 a E1
        hoja.merge_cells(merge_cells)
    
    # Establecer el valor en la celda combinada
    celda = hoja[value_cell] # debe ser "E1" columna y fila
    celda.value = fecha_formateada

    # Aplicar formato: texto en negritas y tamaño 8
    celda.font = Font(bold=True, size=8)
    # Centrar el texto
    celda.alignment = Alignment(horizontal=horizontal, vertical="center")

    if background:
                # Crear un objeto de llenado con un color
        color_fondo = PatternFill(start_color=background, end_color=background, fill_type="solid")  # Amarillo
        celda.fill = color_fondo
    return hoja

def generar_grafico_lineal(
        hoja,
        titulo:str,
        eje_x:str,
        eje_y:str,
        celda_inicio_grafico:str,
        col_category:str, 
        col_data:str,
        min:int,
        max:int
        ):

    # Crear el gráfico de líneas
    chart = LineChart()

    # convertir letra de columna en numero
    data_col = column_index_from_string(col_data)
    categories_col = column_index_from_string(col_category)

    # Referenciar los datos
    data = Reference(hoja, min_col=data_col, min_row=min, max_row=max) 
    categories = Reference(hoja, min_col=categories_col, min_row=min+1, max_row=max) 

    # Agregar los datos al gráfico
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    # Configurar estilo de líneas rectas
    chart.style = 13  

    # Agregar un título al gráfico
    chart.title = titulo

    # Agregar títulos a los ejes
    chart.x_axis.title = eje_x
    chart.y_axis.title = eje_y

    # Definir tamaño del gráfico (en puntos)
    chart.width = 20  # Ancho en pulgadas
    chart.height = 16  # Alto en pulgadas

    # Insertar el gráfico en la hoja de Excel
    hoja.add_chart(chart, celda_inicio_grafico)

    return hoja

def generar_grafico_barras_agrupadas_desde_tabla(hoja, titulo, celda_inicio_grafico, tabla, start,hoja_aux):

    # Crear el gráfico de barras agrupadas
    chart = BarChart()
    chart.grouping = "clustered"  # Asegurar que sea un gráfico de barras agrupadas

    # Extraer las categorías y los datos de las series
    categorias = [fila[0] for fila in tabla[1:]]  # Categorías en la primera columna
    series_data = [fila[1:] for fila in tabla[1:]]  # Datos de las series (resto de columnas)

    # Escribir las categorías en la columna A (desde la fila 'start')
    for i, categoria in enumerate(categorias, start=start+1):
        celda = hoja_aux[f'A{i}']
        celda.value = categoria

    # Escribir los títulos de las series (fila antes de 'start')
    for j, nombre_serie in enumerate(tabla[0][1:], start=start+1):
        hoja_aux.cell(row=start, column=j, value=nombre_serie)

    # Escribir las series de datos a partir de la columna B (desde la fila 'start')
    for col, serie in enumerate(zip(*series_data), start=start+1):  # Transponer para iterar por columna
        for row, dato in enumerate(serie, start=start+1):
            celda = hoja_aux.cell(row=row, column=col, value=dato)

    # Referenciar las categorías (columna A)
    categories = Reference(hoja_aux, min_col=1, min_row=2, max_row=start + len(categorias) )

    # Referencia a los datos (desde fila 1, columnas B a ... )
    data_refs = Reference(hoja_aux,
                          min_col=2,
                          max_col=1 + len(tabla[0][1:]),
                          min_row=1,
                          max_row=start + len(categorias))
    
    # Agregar los datos al gráfico
    chart.add_data(data_refs, titles_from_data=True)
    
    # Establecer las categorías
    chart.set_categories(categories)

    # Configurar el estilo del gráfico
    chart.style = 13
    chart.title = titulo

    # Definir tamaño del gráfico
    chart.width = 35
    chart.height = 16

    # Insertar el gráfico en la hoja
    hoja.add_chart(chart, celda_inicio_grafico)

    return hoja, hoja_aux

def generar_grafico_barras(        
        hoja,
        titulo:str,
        celda_inicio_grafico:str,
        col_category:str, 
        label_category:str,
        col_data:str,
        min:int,
        max:int
        ):

    # Crear un gráfico de barras
    chart = BarChart()

    # Convertir letra de columna en número
    data_col = column_index_from_string(col_data)
    labels_col = column_index_from_string(label_category) 
    categories_col = column_index_from_string(col_category)

    # Referenciar los datos
    data = Reference(hoja, min_col=data_col, min_row=min-1, max_row=max) 
    labels = Reference(hoja, min_col=labels_col, min_row=min-1, max_row=max)
    categories = Reference(hoja, min_col=categories_col, min_row=min, max_row=max)  

    # Agregar el título
    chart.title = titulo

    # Agregar los datos al gráfico
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    # Definir tamaño del gráfico (en puntos)
    chart.width = 25  # Ancho en pulgadas
    chart.height = 16  # Alto en pulgadas

    # Insertar el gráfico en la hoja de Excel
    hoja.add_chart(chart, celda_inicio_grafico)

    return hoja

def generar_tabla_avance_programa(
        hoja, 
        titulo, 
        celda_inicio_grafico,
        columna_inicial, 
        inicio_filas, 
        termino_filas
        ):
    # Crear un gráfico de barras
    chart = BarChart()
    chart.grouping = "clustered"  # Asegurar que sea un gráfico de barras agrupadas

    # Referencia para los datos (asumiendo que los datos están en la columna "B")
    data_real = Reference(hoja, min_col=columna_inicial, min_row=inicio_filas-1, max_row=termino_filas)
    data_plan = Reference(hoja, min_col=columna_inicial+1, min_row=inicio_filas-1, max_row=termino_filas)
    # Referencia para las categorías (asumiendo que las categorías están en la columna "A")
    categories = Reference(hoja, min_col=1, min_row=inicio_filas, max_row=termino_filas)
    
    # Agregar datos y categorías al gráfico
    chart.add_data(data_real, titles_from_data=True)
    chart.add_data(data_plan, titles_from_data=True)
    chart.set_categories(categories)
    
    # Establecer el título del gráfico
    chart.title = titulo
    # Establecer el tamaño del gráfico
    chart.width = 10  # Ancho en pulgadas
    chart.height = 8 # Alto en pulgadas

    # Insertar el gráfico en la hoja de Excel
    hoja.add_chart(chart, celda_inicio_grafico)

    return hoja

# Función para obtener valores corregidos de celdas mergeadas
def obtener_valores_celdas_mergeadas(hoja, columna, inicio_fila, fin_fila):
    valores = []
    ultimo_valido = None
    for fila in range(inicio_fila, fin_fila + 1):
        valor = hoja[f"{columna}{fila}"].value
        if valor is not None:
            ultimo_valido = valor
        valores.append(ultimo_valido)
    return valores

def generar_grafico_barras_agrupadas_avance_diario(
        hoja_aux, 
        hoja_destino, 
        titulo, 
        celda_inicio_grafico, 
        tipo,
        tabla
        ):
    
    # Crear el gráfico de barras agrupadas
    chart = BarChart()
    chart.grouping = "clustered"  # Asegurar que sea un gráfico de barras agrupadas

    # Extraer las categorías y los datos de las series
    categorias = [fila[0] for fila in tabla[1:]]  # Categorías en la primera columna
    series_data = [fila[1:] for fila in tabla[1:]]  # Datos de las series (resto de columnas)


    # Escribir las categorías en la columna A (desde la fila 'start')
    for i, categoria in enumerate(categorias, start=2):
        celda = hoja_aux[f'A{i}']
        celda.value = categoria

    # Escribir los títulos de las series (fila antes de 'start')
    for j, nombre_serie in enumerate(tabla[0][1:], start=2):
        hoja_aux.cell(row=1, column=j, value=nombre_serie)

   # Escribir las series de datos a partir de la columna B (desde la fila 'start')
    for col, serie in enumerate(zip(*series_data), start=2):  # Transponer para iterar por columna
        for row, dato in enumerate(serie, start=2):
            celda = hoja_aux.cell(row=row, column=col, value=dato)

    # Referenciar las categorías (columna A)
    categories = Reference(hoja_aux, min_col=1, min_row=2, max_row=1 + len(categorias) )

    # Referencia a los datos (desde fila 1, columnas B a ... )
    data_refs = Reference(hoja_aux,
                          min_col=2,
                          max_col=1 + len(tabla[0][1:]),
                          min_row=1,
                          max_row=1 + len(categorias))

    # Agregar los datos al gráfico
    chart.add_data(data_refs, titles_from_data=True)
    
    # Establecer las categorías
    chart.set_categories(categories)

    # # Configurar el estilo del gráfico
    chart.style = 10
    chart.title = titulo

    # # Definir tamaño del gráfico
    chart.width = 10  # Ancho en pulgadas
    chart.height = 7  # Alto en pulgadas

    # # Insertar el gráfico en la hoja de destino (hoja_destino)
    hoja_destino.add_chart(chart, celda_inicio_grafico)

    return hoja_destino

def generar_grafico_torta(hoja_aux, hoja_destino, titulo, celda_inicio_grafico, tipo,tabla):
    # Crear el gráfico de torta
    pie_chart = PieChart()



    # Extraer las categorías y los datos de las series
    categorias = [fila[0] for fila in tabla[1:]]  # Categorías en la primera columna
    series_data = [fila[1:] for fila in tabla[1:]]  # Datos de las series (resto de columnas)


    # Escribir las categorías en la columna A (desde la fila 'start')
    for i, categoria in enumerate(categorias, start=2):
        celda = hoja_aux[f'A{i}']
        celda.value = categoria

    # Escribir los títulos de las series 
    for j, nombre_serie in enumerate(tabla[0], start=2):
        hoja_aux.cell(row=1, column=j, value=nombre_serie)

    # Escribir las series de datos a partir de la columna B (desde la fila 'start')
    for col, serie in enumerate(zip(*series_data), start=2):  # Transponer para iterar por columna
        for row, dato in enumerate(serie, start=2):
            celda = hoja_aux.cell(row=row, column=col, value=dato)

    # Los valores numéricos están en la columna B (columna 2)
    data = Reference(hoja_aux, min_col=2, min_row=2, max_row=1 + len(categorias))

    # Las etiquetas están en la columna A (columna 1)
    categories = Reference(hoja_aux, min_col=1, min_row=2, max_row=1 + len(categorias))
    from openpyxl.chart.label import DataLabelList

    # Definir tamaño del gráfico
    pie_chart.width = 10  # Ancho en pulgadas
    pie_chart.height = 7  # Alto en pulgadas
    # Agregar los datos y las categorías al gráfico de torta
    pie_chart.add_data(data, titles_from_data=False)
    pie_chart.set_categories(categories)

    # Mostrar porcentajes en el gráfico de torta
    pie_chart.dataLabels = DataLabelList()
    pie_chart.dataLabels.showVal = True        # Muestra el valor numérico
    pie_chart.dataLabels.showPercent = True    # Muestra el porcentaje
    pie_chart.dataLabels.showCatName = True    # (Opcional) Muestra el nombre de la categoría

    # Insertar el gráfico en la hoja de destino (hoja_destino)
    hoja_destino.add_chart(pie_chart, celda_inicio_grafico)

    return hoja_destino

def generar_grafico_lineal_agrupado(
    hoja,
    titulo:str,
    eje_x:str,
    eje_y:str,
    celda_inicio_grafico:str,
    col_category:str, 
    col_data_planificado:str,
    col_data_real:str,
    min:int,
    max:int
):
    # Crear el gráfico de líneas
    chart = LineChart()

    # Convertir letra de columna en número
    category_col = column_index_from_string(col_category)
    data_col_planificado = column_index_from_string(col_data_planificado)
    data_col_real = column_index_from_string(col_data_real)

    # Referenciar las categorías (eje X)
    categories = Reference(hoja, min_col=category_col, min_row=min, max_row=max)

    # Referenciar los datos planificado y real (eje Y)
    data_planificado = Reference(hoja, min_col=data_col_planificado, min_row=min, max_row=max)
    data_real = Reference(hoja, min_col=data_col_real, min_row=min, max_row=max)

    # Agregar los datos de ambas series al gráfico
    chart.add_data(data_planificado, titles_from_data=True)
    chart.add_data(data_real, titles_from_data=True)

    # Establecer las categorías en el gráfico
    chart.set_categories(categories)

    # Configurar estilo de líneas
    chart.style = 13  # Estilo del gráfico

    # Agregar título al gráfico
    chart.title = titulo

    # Agregar títulos a los ejes
    chart.x_axis.title = eje_x
    chart.y_axis.title = eje_y

    # Definir tamaño del gráfico
    chart.width = 10  # Ancho en pulgadas
    chart.height = 16  # Alto en pulgadas

    # Insertar el gráfico en la hoja de Excel
    hoja.add_chart(chart, celda_inicio_grafico)

    return hoja

def filtrar_reportes_aprobados(reportes, sondajes):
    """Filtra reportes con progreso 'Aprobado'."""
    return [
        reporte for reporte in reportes
        if reporte.get("progreso") == "Aprobado" and
        f"{sondajes.get(reporte['sondajeCodigo'], 'Unknown')}-{reporte['sondajeSerie']}".strip()
    ]

def agrupar_reportes_por_sondaje(reportes_aprobados, sondajes, estados):
    """Agrupa reportes aprobados por código de sondaje."""
    reportes_agrupados = defaultdict(list)

    for reporte in reportes_aprobados:
        # Verificar si reporte es una lista y tiene elementos
        if isinstance(reporte, list) and reporte:
            reporte_data = reporte[0]  # Extrae el primer elemento
        elif isinstance(reporte, dict):  # Si es un diccionario, úsalo directamente
            reporte_data = reporte
        else:
            print(f"Advertencia: formato inesperado en reporte: {reporte}")
            continue  # Ignorar este reporte si no tiene formato válido
    
        codigo_sondaje = f"{sondajes.get(reporte_data.get('sondajeCodigo', ''), 'Unknown')}-{reporte_data.get('sondajeSerie', '')}{estados.get(reporte_data.get('sondajeEstado', ''), '')}".strip()
        reportes_agrupados[codigo_sondaje].append(reporte)

    return reportes_agrupados

def generar_dict_rango_fechas(mes_inicio: int, anio_inicio: int, mes_fin: int, anio_fin: int) -> dict:
    meses_es = {
        1: 'ENERO', 2: 'FEBRERO', 3: 'MARZO', 4: 'ABRIL',
        5: 'MAYO', 6: 'JUNIO', 7: 'JULIO', 8: 'AGOSTO',
        9: 'SEPTIEMBRE', 10: 'OCTUBRE', 11: 'NOVIEMBRE', 12: 'DICIEMBRE'
    }

    resultado = OrderedDict()
    fecha_actual = datetime(anio_inicio, mes_inicio, 1)

    while fecha_actual.year < anio_fin or (fecha_actual.year == anio_fin and fecha_actual.month <= mes_fin):
        anio = fecha_actual.year
        mes = fecha_actual.month
        nombre_mes = meses_es.get(mes, "MES DESCONOCIDO")
        clave = f"{nombre_mes} {anio}"

        dias_en_mes = calendar.monthrange(anio, mes)[1]
        dias_lista = []

        for dia in range(1, dias_en_mes + 1):
            # Formato DD/MM/YYYY con ceros a la izquierda
            dia_str = f"{dia:02}"
            mes_str = f"{mes:02}"
            fecha_str = f"{dia_str}/{mes_str}/{anio}"

            dias_lista.append({
                "Día/Sonda": fecha_str,
                "Cantidad de Sondas": 0,
                "Mts. Día": 0.00,
                "Mts. Proyectados a Perforar": 0.00,
                "Rendimiento Día (m)": 0.00,
                "Rendimiento Proyectado GEOTEC Día (m)": 0.00,
                "Rendimiento por Hora (m)": 0.00
            })

        resultado[clave] = dias_lista

        # Avanzar al siguiente mes
        if mes == 12:
            fecha_actual = datetime(anio + 1, 1, 1)
        else:
            fecha_actual = datetime(anio, mes + 1, 1)

    return resultado