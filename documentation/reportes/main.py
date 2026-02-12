from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import locale
import pytz
import requests 
from pathlib import Path
from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn
from collections import defaultdict
from openpyxl.drawing.image import Image
from openpyxl.formatting.rule import CellIsRule
from pprint import pprint
from openpyxl import load_workbook
import argparse

from django.conf import settings
from documentation.reportes import get_report_sondajes as reportes_sondajes
from documentation.reportes import get_report_perf_vs_rec as reportes_perf_vs_rec
from documentation.reportes import get_report_avance_programa as reportes_avance_programa
from documentation.reportes import get_report_rendimiento_mensual as reportes_rendimiento_mensual
from documentation.reportes import get_report_programas as reportes_programas
from documentation.reportes import get_report_avance_diario as reportes_avance_diario
from documentation.reportes import get_report_gerenciales as reportes_gerenciales
from documentation.reportes import get_report_detalle_r_sonda_dia as reportes_detalle_r_sonda_dia
from documentation.reportes import get_report_recomendacion as reportes_recomendacion
from documentation.reportes import get_report_m_trayectoria as reportes_m_trayectoria
from documentation.reportes import get_report_avance_muestrera as reportes_avance_muestrera
from documentation.reportes import get_graficos_avance_diario as graficos_avance_diario
from documentation.reportes import helpers

from core.models import (
    Perforistas, Sondas, Sondajes, Diametros, TipoTerreno, Orientacion,
    CantidadAgua, Aditivos, DetalleControlHorario, MaterialesSonda, MaterialesCaseta, Recomendacion,Faena
)

class ObtenerReportes():
    def __init__(self,fecha_inicial,fecha_final,faena_id):
          
          self.fecha_inicial = fecha_inicial
          self.fecha_final = fecha_final
          # Separar el string por el guion "-"
          fecha_inicial_str = fecha_inicial.strftime('%Y-%m-%d')
          fecha_final_str = fecha_final.strftime('%Y-%m-%d')

          self.anio_inicial,self.mes_inicial, self.dia_inicial = fecha_inicial_str.split('-')
          self.anio_final, self.mes_final, self.dia_final = fecha_final_str.split('-')
          
          self.faena_id = faena_id

    def consumir_api(self,url,params):
        try:
            response = requests.get(url, params=params , timeout=10)
            response.raise_for_status()
            return response.json()
        except requests.Timeout:
            raise ValueError("La solicitud a la API ha excedido el tiempo de espera.")
        except requests.RequestException as e:
            raise ValueError(f"Error al consumir la API: {e}")

    def consumir_datos_api(self):

        endpoint = "/api/reporte_avance_campana/"
        
        url_api = f"{settings.BASE_API_URL.rstrip('/')}/{endpoint.lstrip('/')}"
        
        if settings.DEBUG:
            print(f"🚀 MODO DESARROLLO: Consumiendo reporte desde {url_api}")

        params = {
            'fecha_inicio': self.fecha_inicial,
            'fecha_final': self.fecha_final,
            'faena_id': self.faena_id
        }
        
        data = self.consumir_api(url_api, params)
        #https://sicgeoatacama.cl/api/reporte_avance_campana/?fecha_inicio=2025-03-10&fecha_final=2025-03-12&faena_id=2
        #http://127.0.0.1:8000/api/reporte_avance_campana/?fecha_inicio=2025-03-10&fecha_final=2025-03-12&faena_id=1


        if 'error' in data:
            return None,None, None, None, None, None,None, None, None, None,None,None,None,None,None,data['error']

        
        
        if not data["data_sondas"] or "sondajes" not in data["data_sondas"] or "gemelo" not in data["data_sondas"]:
            print("Error: La API devolvió datos en un formato inesperado o vacío.")
            return None,None, None, None, None, None,None, None, None, None,None,None,None, "Error: La API devolvió datos en un formato inesperado o vacío."

        sondajes = {item["id"]: item["sondaje"] for item in data["data_sondas"].get("sondajes", [])}
        estados = {item["value"]: item["display"] for item in data["data_sondas"].get("gemelo", [])}
        sondas = {item["id"]: item["sonda"] for item in data["data_sondas"].get("sondas", [])}
        reportes = data["data_sondas"].get("reportesOperacionales", [])
        reportes_controles_horarios = data["data_sondas"].get("reportesControlesHorarios", [])
        reportes_detalles = data["data_sondas"].get("reportesDetallesPerforaciones", [])
        diametros_data = {item["id"]: item["diametro"] for item in data["data_sondas"].get("diametros", [])}
        recomendaciones = data["data_sondas"].get("recomendaciones", [])
        observaciones = data["data_sondas"].get("reportesObservacionesReportes", [])


        if not reportes:
            print("No hay datos de reportes operacionales disponibles.")
            return None,None, None, None, None, None,None, None, None, None,None,None,None,"Error: No hay datos de reportes operacionales disponibles."

        all_ajustes = data["all_ajuste_recomendaciones"]
        all_final = data["all_final_recomendaciones"]

        if not all_ajustes:
            all_ajustes = []

        if not all_final:
            all_final = []

        
        return observaciones,sondajes, estados, sondas, reportes, reportes_controles_horarios,reportes_detalles, diametros_data, recomendaciones,data["all_recomendaciones"], data["data_campanas"],data["data_programas"],data["data_planificacion_programas"],all_ajustes,all_final, None

    def buscar_sondaje(self,sondaje_id,sondajes):

        if sondaje_id in sondajes:
            return sondajes[sondaje_id]
        else:
            return "SIN DATO - SONDAJE"
    def buscar_sonda(self,sonda_id,sondas):
        
        if sonda_id in sondas:
            return sondas[sonda_id]
        else:
            return "SIN DATO - SONDA"

    def buscar_campana(self,campana_id,campanas):

        for campana in campanas:
            if campana['id'] == campana_id:
                return campana['campana']

        return "SIN DATO - CAMPAÑA"
    def buscar_programa(self,programa_id,programas):

        for programa in programas:
            if programa['id'] == programa_id:
                return programa['programa']

        return "SIN DATO - PROGRAMA"
    def buscar_recomendacion(self,value , recomendaciones,nombre_sonda,campanas,programas):

        for rec in recomendaciones:

            if value == rec['pozo']:

                resultado = {
                    'campana': self.buscar_campana(rec['campana'],campanas),
                    'programa': self.buscar_programa(rec['programa'],programas),
                    # 'campana': rec.get('campana') or 'SIN CAMPAÑA',
                    # 'programa': rec.get('programa') or 'SIN PROGRAMA',
                    'azimut': int(rec['azimut']),
                    'cota': float(rec['cota']),
                    'creador': rec['creador'],
                    'este': float(rec['este']),
                    'fecha_inicio': datetime.strptime(rec['fecha_inicio'].split("T")[0], "%Y-%m-%d").strftime("%d/%m/%Y"),
                    'fechacreacion': datetime.strptime(rec['fechacreacion'].split("T")[0], "%Y-%m-%d").strftime("%d/%m/%Y"),
                    'fechaupdateestado': datetime.strptime(rec['fechaUpdateEstado'].split("T")[0], "%Y-%m-%d").strftime("%d/%m/%Y"),
                    'id': rec['id'],
                    'inclinacion': float(rec['inclinacion']),
                    'largo_programado': float(rec['largo_programado']),
                    'largo_real': float(rec['largo_real']),
                    'norte': float(rec['norte']),
                    'pozo': value,
                    'recomendacion': rec['recomendacion'],
                    'sector': rec['sector'],
                    'sonda': nombre_sonda,
                    'status': rec['status'],
                    'estado': rec['estado']
                    }

                return resultado

        return

    def data(self,reportes_agrupados, recomendaciones,sondas,sondajes,estados,campanas,programas):

        pozos_sin_recomendacion = []
        for nombre_pozo, detalles in reportes_agrupados.items():
            #recomendacion = self.buscar_recomendacion(nombre, recomendaciones,sondas,sondajes)
            # Agregamos recomendacion a cada uno de los detalles ( perforaciones)
            for detalle in detalles:

                nombre_sonda = self.buscar_sonda(detalle['sonda'],sondas)
                nombre_sondaje = self.buscar_sondaje(detalle['sondajeCodigo'],sondajes)

                recomendacion = self.buscar_recomendacion(nombre_pozo, recomendaciones,nombre_sonda,campanas,programas)

                if not recomendacion:
                    pozos_sin_recomendacion.append(nombre_pozo)
                    continue

                detalle["nombre_sonda"] = nombre_sonda
                detalle["nombre_sondaje"] = nombre_sondaje
                detalle["fechacreacion"] = datetime.strptime(detalle["fechacreacion"].split("T")[0], "%Y-%m-%d").strftime("%d/%m/%Y")
                detalle["fechaedicion"] = datetime.strptime(detalle["fechaedicion"].split("T")[0], "%Y-%m-%d").strftime("%d/%m/%Y")
                detalle["recomendacion"] = recomendacion

        return reportes_agrupados, pozos_sin_recomendacion

    def run(self):
        libro = Workbook()
        observaciones, sondajes, estados, sondas, reportes, reportes_controles_horarios, reportes_detalles, diametros_data, recomendaciones,all_recomendaciones, campanas, programas, planificacion_programas,all_recomendaciones_ajuste, all_recomendaciones_final, error= self.consumir_datos_api()

        if error:
            return libro, f'{error}'
        
        if not reportes:
            return libro, f"No existen reportes para esta busqueda"

        reportes_aprobados = helpers.filtrar_reportes_aprobados(reportes, sondajes)
        reportes_agrupados = helpers.agrupar_reportes_por_sondaje(reportes_aprobados, sondajes, estados)

        
        primera_hoja = True

        ## Se generan los reportes de cada sonda en el excel
        libro = reportes_sondajes.generar_excel(libro,primera_hoja,reportes_agrupados,sondas,reportes_controles_horarios,reportes_detalles,diametros_data,recomendaciones,observaciones)

        # Se generan los reportes detalles r sonda por dia en el excel
        sondajes_recomendaciones, pozos_sin_recomendacion = self.data(reportes_agrupados,all_recomendaciones,sondas,sondajes,estados,campanas,programas)
        pozos_sin_recomendacion = set(pozos_sin_recomendacion)

        if pozos_sin_recomendacion:
            
            if len(pozos_sin_recomendacion) > 10:
                return libro, f"Existe un total de {len(pozos_sin_recomendacion)} pozos sin recomendaciones asignadas"
            else:
                return libro, f"Existen pozos sin recomendaciones:\n {', '.join(repr(p) for p in pozos_sin_recomendacion)}"
  
        libro, datos_mensuales = reportes_detalle_r_sonda_dia.run(libro,sondajes_recomendaciones)

        # Se generan los reportes gerenciales en el excel
        libro,reportes_agrupados = reportes_perf_vs_rec.run(libro,reportes_agrupados, reportes_detalles,diametros_data)

        # Se generan los reportes recomendaciones en el excel
        libro = reportes_recomendacion.run(libro,reportes_agrupados,all_recomendaciones,sondas,all_recomendaciones_ajuste, all_recomendaciones_final)

        # Se generan los reportes gerenciales en el excel
        libro,nombres_programas = reportes_avance_programa.run(libro,campanas,programas,planificacion_programas,datos_mensuales,sondajes_recomendaciones)


        nombre_faena = Faena.objects.get(id=int(self.faena_id)).faena

        # Se generan los reportes m trayectoria en el excel
        libro = reportes_m_trayectoria.run(libro,reportes_agrupados,nombres_programas,self.anio_inicial,self.anio_final,nombre_faena)

        # Se generan los reportes de rendimiento mensual en el excel
        libro = reportes_rendimiento_mensual.run(libro,reportes_agrupados,nombres_programas,sondajes_recomendaciones)

        # Se generan los reportes avance muestrera en el excel
        libro = reportes_avance_muestrera.run(libro,reportes_agrupados,nombres_programas)

        # Se generan los reportes de programas en el excel
        libro, anio_inicial_avance, anio_final_avance = reportes_programas.run(libro,reportes_agrupados,programas,campanas,all_recomendaciones_final)

        # Se generan los reportes de avance diario en el excel
        libro, ultima_fila_avance_diario, nombre_hoja_avance_diario = reportes_avance_diario.run(libro,reportes_agrupados,nombres_programas)

        # Se generan los reportes gerenciales en el excel
        libro, ultima_fila_gerencial, celda_inicial_tabla_gerencial,data,data_seguna_tabla, mes_actual = reportes_gerenciales.run(libro,reportes_agrupados,programas,campanas,planificacion_programas,self.anio_final, self.mes_final)

        libro = graficos_avance_diario.run(libro, ultima_fila_avance_diario, ultima_fila_gerencial-1, celda_inicial_tabla_gerencial,nombre_hoja_avance_diario,data,data_seguna_tabla, mes_actual, anio_inicial_avance, anio_final_avance)

        return libro, None


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    
    parser.add_argument(
        "-fi",
        "--fecha_inicial",
        required=True,
    )
    parser.add_argument(
        "-ff",
        "--fecha_final",
        required=True,
    )
    parser.add_argument(
        "-faena",
        "--faena_id",
        required=True,
    )

    args = parser.parse_args()
    fecha_inicial = args.fecha_inicial
    fecha_final = args.fecha_final
    faena_id = args.faena_id

    ObtenerReportes(fecha_inicial,fecha_final, faena_id).run()