from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime, timedelta
import locale
import pytz
import requests 
from pathlib import Path
from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn
from collections import defaultdict
from openpyxl.drawing.image import Image
from openpyxl.formatting.rule import CellIsRule
from pprint import pprint
from openpyxl import load_workbook
from documentation.reportes import helpers
import re


def agregar_datos_a_celdas(hoja, inicio_celda, datos, primer_color, segundo_color):

    fila_inicio = inicio_celda

    # Definir los colores alternados para las filas
    color_fondo_1 = PatternFill(start_color=primer_color, end_color=primer_color, fill_type="solid")
    color_fondo_2 = PatternFill(start_color=segundo_color, end_color=segundo_color, fill_type="solid")

    ultima_fila = fila_inicio  # Variable para almacenar la última fila usada

    for idx,diccionario in enumerate(datos):
        # Determinar la fila donde se agregará este diccionario
        fila_actual = fila_inicio + idx
        ultima_fila = fila_actual  # Actualizar la última fila usada
        if idx % 2 == 0:
            color_fondo = color_fondo_1  # Verde para filas pares
        else:
            color_fondo = color_fondo_2  # Rojo para filas impares

        for  j, (key, value)in enumerate(diccionario.items()):


            celda = hoja[f"{key}{fila_actual}"]
            celda.value = value
            # Aplicar formato a la celda
            celda.font = Font(name="Arial", size=8, bold=False)
            celda.alignment = Alignment(horizontal="center", vertical="center")
            # Aplicar el color de fondo a la celda
            celda.fill = color_fondo

    return hoja, ultima_fila

def ordenar_data(data, mes):

    


    # Obtener valores para primer grafico de barras
    primera_tabla = [["Campañas", "Planificado", "Avance Campaña"]]
    segunda_tabla = [["Campañas", f"Metros Programados a {mes}", "Metros de avance por programa"]]
    tercera_tabla = [['Categoría', 'Valor']]

    total_primera_tabla = ["TOTAL", 0.00, 0.00]
    total_segunda_tabla = ["TOTAL", 0.00, 0.00]
    for campana, detalle in data.items():
        avance_real = 0.00
        metros_planificados_mes = 0.00
        for programa in detalle['Programas']:
            for key, value in programa.items():
                if key == "Metros de avance por programa":
                    avance_real += value
                if key == f"Metros programados a {mes}":
                    metros_planificados_mes += value


        total_primera_tabla[1] += detalle['Metros Planificados']
        total_primera_tabla[2] += avance_real
        primera_tabla.append([campana, detalle['Metros Planificados'], avance_real])

        total_segunda_tabla[1] += metros_planificados_mes
        total_segunda_tabla[2] += avance_real
        segunda_tabla.append([campana, metros_planificados_mes, avance_real])

        tercera_tabla.append([campana, detalle['Metros Planificados']])

    #primera_tabla.append(total_primera_tabla)
    #segunda_tabla.append(total_segunda_tabla)
    planificacion_total = round(float(total_primera_tabla[1]), 2)
    programado_al_mes = round(float(total_segunda_tabla[1]), 2)
    avance_actual = round(float(total_segunda_tabla[2]), 2)
    porcentaje_avance_planificacion = round((total_primera_tabla[2]*100)/total_primera_tabla[1], 2)
    porcentaje_avance_programacion = round((total_segunda_tabla[2]*100)/total_segunda_tabla[1], 2)


    return primera_tabla, segunda_tabla, tercera_tabla,planificacion_total, programado_al_mes, avance_actual, porcentaje_avance_planificacion, porcentaje_avance_programacion

def obtener_celdas_intermedias(rango):
    inicio, fin = rango.split(":")  # Separar inicio y fin
    col_inicio, fila_inicio = inicio[0], int(inicio[1:])  # Extraer columna y fila inicial
    col_fin, fila_fin = fin[0], int(fin[1:])  # Extraer columna y fila final
    
    cols = range(column_index_from_string(col_inicio), column_index_from_string(col_fin) + 1)
    return [f"{get_column_letter(c)}" for c in cols]

def asignar_columnas(diccionario, celdas_intermedias,graficos_actuales,total_data):
    fila_resultado = {}  # Diccionario para almacenar los resultados

    # La columna 'A' siempre debe contener el valor de 'mes'
    if 'mes' in diccionario:
        fila_resultado['A'] = diccionario['mes']
    
    columnas_disponibles = celdas_intermedias[1:]  # Excluimos 'A' porque es para 'mes'
    indice_columna = 0  # Índice para recorrer la lista de columnas disponibles

    # Iterar sobre cada categoría en el diccionario
    for categoria, subdiccionario in diccionario.items():
        if categoria == 'mes':  # Saltar 'mes' porque ya fue asignado a 'A'
            continue
        if categoria not in graficos_actuales:
            graficos_actuales[categoria] = {"plan":{}, "acumulado_plan":{}}
        
        # Agregar los valores al gráfico
        graficos_actuales[categoria]["plan"]["columna_inicial"]=indice_columna+2
        graficos_actuales[categoria]["plan"]["inicio_filas"]=5
        graficos_actuales[categoria]["plan"]["termino_filas"]=5 + total_data
        graficos_actuales[categoria]["acumulado_plan"]["columna_inicial"]=indice_columna+2
        graficos_actuales[categoria]["acumulado_plan"]["inicio_filas"]=5
        graficos_actuales[categoria]["acumulado_plan"]["termino_filas"]=5 + total_data

        if isinstance(subdiccionario, dict):  # Verificar si es un diccionario anidado
            for subkey, subvalue in subdiccionario.items():
                if indice_columna < len(columnas_disponibles):  # Asegurar que no excedamos las columnas

                    columna = columnas_disponibles[indice_columna]
                    fila_resultado[columna] = subvalue
                    indice_columna += 1  # Avanzar a la siguiente columna

    return fila_resultado, graficos_actuales
def run(libro, ultima_fila_avance_diario, ultima_fila_gerencial, celda_inicial_tabla_gerencial,nombre_hoja_avance_diario,data,data_seguna_tabla, mes_actual, anio_inicial_avance, anio_final_avance):

    hoja_avance_diario = libro[nombre_hoja_avance_diario]

    primera_tabla, segunda_tabla, tercera_tabla,planificacion_total, programado_al_mes, avance_actual, porcentaje_avance_planificacion, porcentaje_avance_programacion = ordenar_data(data_seguna_tabla, mes_actual)

    # Se genera hoja auxiliar
    hoja_aux_plan_avance_real = libro.create_sheet(title="AUX_AVANCE_PLAN_REAL")

    # Generar el gráfico de barras agrupadas desde la tabla
    hoja_avance_diario = helpers.generar_grafico_barras_agrupadas_avance_diario(
        hoja_aux = hoja_aux_plan_avance_real,
        hoja_destino = hoja_avance_diario,
        titulo = "Planificación total vs avance real",
        celda_inicio_grafico = "B5",
        tipo = "planificacion",
        tabla = primera_tabla
        )

    # Se genera hoja auxiliar
    hoja_aux_perforacion_real = libro.create_sheet(title="AUX_PERFORACION_REAL")
    # Generar el gráfico de barras agrupadas desde la tabla
    hoja_avance_diario = helpers.generar_grafico_barras_agrupadas_avance_diario(
        hoja_aux = hoja_aux_perforacion_real,
        hoja_destino = hoja_avance_diario,
        titulo = f"Planificación a {mes_actual} vs perforación real",
        celda_inicio_grafico = "B19",
        tipo = "perforacion",
        tabla= segunda_tabla
        )

    # Se genera hoja auxiliar
    hoja_aux_grafico_circular = libro.create_sheet(title="AUX_AVANCE_DIARIO")  
    hoja_avance_diario = helpers.generar_grafico_torta(
        hoja_aux = hoja_aux_grafico_circular,
        hoja_destino = hoja_avance_diario,
        titulo = "Planificación total vs avance real",
        celda_inicio_grafico = "I19",
        tipo = "perforacion",
        tabla= tercera_tabla
        )
  
    # Fusionar celdas y asignar formato a la celda superior izquierda
    celda = hoja_avance_diario['I5']  # Celda superior izquierda del rango fusionado
    celda.value =  f"Planificación total    {planificacion_total}                      "
    celda.font = Font( size=12, bold=True)
    celda.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)

    # Fusionar celdas y asignar formato a la celda superior izquierda
    celda = hoja_avance_diario['I7']  # Celda superior izquierda del rango fusionado
    celda.value = f"Programado a {mes_actual}    {programado_al_mes}                      "
    celda.font = Font( size=12, bold=True)
    celda.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    # Fusionar celdas y asignar formato a la celda superior izquierda
    celda = hoja_avance_diario['I10']  # Celda superior izquierda del rango fusionado
    celda.value = f"Avance actual    {avance_actual}                      "
    celda.font = Font( size=12, bold=True)
    celda.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    # Fusionar celdas y asignar formato a la celda superior izquierda
    celda = hoja_avance_diario['I13']  # Celda superior izquierda del rango fusionado
    celda.value = f"%Avance respecto a Planificación total       {porcentaje_avance_planificacion}%                      "
    celda.font = Font( size=12, bold=True)
    celda.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    # Fusionar celdas y asignar formato a la celda superior izquierda
    celda = hoja_avance_diario['I16']  # Celda superior izquierda del rango fusionado
    celda.value = f"%Avance respecto a Programación total       {porcentaje_avance_programacion}%                      "
    celda.font = Font( size=12, bold=True)
    celda.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    

    return libro