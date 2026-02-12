from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime, timedelta
import locale
import pytz
import requests 
from pathlib import Path
from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn
from collections import defaultdict
from openpyxl.drawing.image import Image
from openpyxl.formatting.rule import CellIsRule
from pprint import pprint
from openpyxl import load_workbook
from documentation.reportes import helpers
import re


def aplicar_borde_a_rango(hoja, celda_inicio, celda_fin):
    # Extraer la parte numérica de la celda
    fila_inicio = int(re.search(r'\d+', celda_inicio).group())
    fila_fin = int(re.search(r'\d+', celda_fin).group())

    columna_inicio = re.search(r'[A-Z]+', celda_inicio).group()

    for fila in range(fila_inicio, fila_fin + 1):
        celda = hoja[f"{columna_inicio}{fila}"]
        celda.border = Border(left=Side(border_style="thin", color="FF000000"))
    
    return hoja

def agregar_datos_a_celdas(hoja, inicio_celda, datos, primer_color, segundo_color):
    # Obtener la fila de inicio desde la celda (ej. "A1" -> 1)
    fila_inicio = int(inicio_celda[1:])

    # Definir los colores alternados para las filas
    color_fondo_1 = PatternFill(start_color=primer_color, end_color=primer_color, fill_type="solid")
    color_fondo_2 = PatternFill(start_color=segundo_color, end_color=segundo_color, fill_type="solid")

    ultima_fila = fila_inicio  # Variable para almacenar la última fila usada

    # Mapeo actualizado de claves del JSON a columnas de Excel
    mapeo_columnas = {
        "ID": "B", 
        "PROGRAMA": "C", 
        "Azimut": "D", 
        "Dip": "E", 
        "Fondo Teórico": "F", 
        "Fondo Real": "G",
        "Falta Perforar": "H",
        "Estado": "I",
        "Desde": "J",
        "Hasta": "K",
        "Falta": "L",
        "Fecha Medición": "M",
        "Operador": "N",
        "Unidad de Medición": "O",
        "Placa": "P",
        "Giroscopio": "Q",
        "Certificado": "R",
        "% Desviación": "S",
        "Multa": "T",
        "Observación": "U",
        "Desde2": "V",
        "Hasta2": "W",
        "Falta2": "X",
        "Fecha": "Y",
        "Operador2": "Z",
        "Unidad de Medición2": "AA",
        "Reporte": "AB",
        "Teleview": "AC"
    }

    # Iterar sobre los datos
    for idx, diccionario in enumerate(datos):
        fila_actual = fila_inicio + idx

        # Agregamos el indice de cada ID
        celda = hoja[f"A{fila_actual}"]
        celda.value = idx+1
        celda.font = Font(name="Arial", size=8, bold=True)
        celda.alignment = Alignment(horizontal="center", vertical="center")

        ultima_fila = fila_actual  # Actualizar la última fila usada
        color_fondo = color_fondo_1 if idx % 2 == 0 else color_fondo_2  # Alternar colores

        # Iterar sobre cada clave y asignar los valores en la hoja
        for key, value in diccionario.items():
            if key in mapeo_columnas:
                columna = mapeo_columnas[key]
                celda = hoja[f"{columna}{fila_actual}"]
                celda.value = value

                # Aplicar formato a la celda
                celda.font = Font(name="Arial", size=8, bold=False)
                if columna == "B":
                    celda.font = Font(name="Arial", size=8, bold=True)
                celda.alignment = Alignment(horizontal="center", vertical="center")
                celda.fill = color_fondo  # Aplicar color de fondo

                if columna =="E" or columna  =="F" or columna  =="G" or columna  =="H" or columna  =="J" or columna  =="K" or columna  =="L" :
                    celda.number_format = '0.00'

    return hoja, ultima_fila

def agregar_titulos_tablas(hoja, anio_inicial, anio_final,nombre_faena):
    
    # Agregar TITULOS
    hoja.merge_cells("B2:I2")
    hoja = helpers.agregar_titulos(hoja, "B", 2, [f"Medición Desviación de sondajes - {nombre_faena.upper()} {anio_inicial}_{anio_final}"],"c5e0b3", alto_fila_factor=2, ancho_columna=8,color_letras= "negro")

    hoja.merge_cells("J2:U2")
    hoja = helpers.agregar_titulos(hoja, "J", 2, ["Medidos"],"FFFF00", alto_fila_factor=2, ancho_columna=8,color_letras= "negro")

    hoja.merge_cells("V2:AC2")
    hoja = helpers.agregar_titulos(hoja, "V", 2, ["Filmados"],"ffc000", alto_fila_factor=2, ancho_columna=8,color_letras= "negro")
    
    titulos= [
        "ID", 
        "PROGRAMA", 
        "Azimut", 
        "Dip", 
        "Fondo Teórico", 
        "Fondo Real",
        "Falta Perforar",
        "Estado",
        "Desde",
        "Hasta",
        "Falta",
        "Fecha Medición",
        "Operador",
        "Unidad de Medición",
        "Placa",
        "Giroscopio",
        "Certificado",
        "% Desviación",
        "Multa",
        "Observación",
        "Desde2",
        "Hasta2",
        "Falta2",
        "Fecha",
        "Operador2",
        "Unidad de Medición2",
        "Reporte",
        "Teleview"
    ]
    hoja= helpers.agregar_titulos(hoja, "B", 3, titulos,"70ad47", alto_fila_factor=2, ancho_columna=8,color_letras= "negro")

    return hoja

def construir_data(reportes_agrupados,nombres_programas):

    opciones_estado = {
            '1':'Abortado',
            '2':'En Avance',
            '3':'En Espera',
            '4':'Finalizado',
            }
    
    resultado = []
    for reporte, detalles in reportes_agrupados.items():

        detalles_ordenados = sorted(
            detalles,
            key=lambda x: (float(x['metroFinal']), -float(x['metroInicial']))
        )
        datos_finales = detalles_ordenados[-1]

        fondo_teorico = float(datos_finales['recomendacion']['largo_programado'])
        fondo_real = float(datos_finales['totalPerforado'])

        id_estado = datos_finales['recomendacion'].get('estado')

        if id_estado:
    
            estado = opciones_estado[datos_finales['recomendacion']['estado']]
            
        else:

            estado = "SIN ESTADO"

        resultado.append({
            "ID" :reporte,
            "PROGRAMA": datos_finales['recomendacion']['programa'],
            "Azimut": datos_finales['recomendacion']['azimut'],
            "Dip": datos_finales['recomendacion']['inclinacion'],
            "Fondo Teórico": fondo_teorico,
            "Fondo Real": fondo_real,
            "Falta Perforar": fondo_teorico-fondo_real,
            "Estado": estado,
            #EStos datos no estan agregados
            "Desde": 0.00,
            "Hasta": 0.00,
            "Falta": 0.00,
            "Fecha Medición": "SIN DATO",
            "Operador": "SIN DATO",
            "Unidad de Medición": "SIN DATO",
            "Placa": "SIN DATO",
            "Giroscopio": "SIN DATO",
            "Certificado": "SIN DATO",
            "% Desviación": "SIN DATO",
            "Multa": "SIN DATO",
            "Observación": "SIN DATO",
            "Desde2": "SIN DATO",
            "Hasta2": "SIN DATO",
            "Falta2": "SIN DATO",
            "Fecha": "SIN DATO",
            "Operador2": "SIN DATO",
            "Unidad de Medición2": "SIN DATO",
            "Reporte": "SIN DATO",
            "Teleview": "SIN DATO"
            })
    return resultado

def run(libro,reportes_agrupados,nombres_programas, anio_inicial, anio_final,nombre_faena):


    data = construir_data(reportes_agrupados,nombres_programas)
    
    # Agregar una nueva hoja 
    hoja = libro.create_sheet(title="M. DE TRAYECTORIA")

    hoja = agregar_titulos_tablas(hoja, anio_inicial, anio_final,nombre_faena)

    hoja , ultima_fila = agregar_datos_a_celdas(hoja, "B4", data, "e2efd9", "b4c6e7")
    hoja = aplicar_borde_a_rango(hoja,"J2",f'J{ultima_fila}')
    hoja = aplicar_borde_a_rango(hoja,"V2",f'V{ultima_fila}')
    hoja = aplicar_borde_a_rango(hoja,"AD2",f'AD{ultima_fila}')

    # Ajustar el ancho de las columnas
    hoja.column_dimensions["A"].width = 2
    hoja.column_dimensions["B"].width = 13
    hoja.column_dimensions["C"].width = 13
    hoja.column_dimensions["D"].width = 6
    hoja.column_dimensions["E"].width = 6
    hoja.column_dimensions["F"].width = 10
    hoja.column_dimensions["G"].width = 8
    hoja.column_dimensions["H"].width = 8
    hoja.column_dimensions["I"].width = 8
    hoja.column_dimensions["J"].width = 8
    hoja.column_dimensions["K"].width = 8
    hoja.column_dimensions["L"].width = 8
    hoja.column_dimensions["M"].width = 8
    hoja.column_dimensions["N"].width = 40
    hoja.column_dimensions["O"].width = 8
    hoja.column_dimensions["P"].width = 8
    hoja.column_dimensions["Q"].width = 8
    hoja.column_dimensions["R"].width = 8
    hoja.column_dimensions["S"].width = 8
    hoja.column_dimensions["T"].width = 8
    hoja.column_dimensions["U"].width = 25
    hoja.column_dimensions["V"].width = 10
    hoja.column_dimensions["W"].width = 10
    hoja.column_dimensions["X"].width = 10
    hoja.column_dimensions["Y"].width = 10
    hoja.column_dimensions["Z"].width = 25
    hoja.column_dimensions["AA"].width = 10
    hoja.column_dimensions["AB"].width = 10
    hoja.column_dimensions["AC"].width = 10

    hoja = helpers.generar_grafico_barras(
        hoja,
        titulo = f"Sondajes con medición trayectoria vs fondo real - {nombre_faena.upper()} {anio_inicial}_{anio_final}",
        celda_inicio_grafico =f"A{ultima_fila+8}",
        col_category="B",
        label_category="G",
        col_data="K",
        min=4,
        max=ultima_fila,
        )
    libro._sheets.insert(0, libro._sheets.pop(-1))

    return libro