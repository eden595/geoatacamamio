from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime, timedelta
import locale
import pytz
import requests 
from pathlib import Path
from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn
from collections import defaultdict
from openpyxl.drawing.image import Image
from openpyxl.formatting.rule import CellIsRule
from pprint import pprint
from openpyxl import load_workbook
from documentation.reportes import helpers
import re

def agregar_datos_a_celdas(hoja, inicio_celda, datos, primer_color, segundo_color):
    # Obtener la fila de inicio desde la celda (ej. "A1" -> 1)
    fila_inicio = int(inicio_celda[1:])

    # Definir los colores alternados para las filas
    color_fondo_1 = PatternFill(start_color=primer_color, end_color=primer_color, fill_type="solid")
    color_fondo_2 = PatternFill(start_color=segundo_color, end_color=segundo_color, fill_type="solid")

    ultima_fila = fila_inicio  # Variable para almacenar la última fila usada

    # Mapeo actualizado de claves del JSON a columnas de Excel
    mapeo_columnas = {
        "Rec": "A",
        "Programa": "B",
        "Sondaje": "C",
        "Estado de Pago": "D",
        "Sonda": "E",
        "Levantamiento de Collar": "F",
        "Azimut": "G",
        "Inclinación": "H",
        "Inicio": "I",
        "Término": "J",
        "Estado": "K",
        "Largo Programado": "L",
        "Fondo Final": "M",
        "Nº Bandejas": "N",
        "Faltante": "O",
        "% Perforado": "P",
        "Medición": "Q",
        "Desde": "R",
        "Hasta": "S",
        "Certificado Medición de Trayectoria": "T",
        "% Recuperación de Sondajes": "U",
        "% Rendimiento Sondajes": "V",
        "% Desviación": "W",
        "Tricono": "X",
        "Fotografía": "Y",
        "Corte": "Z",
        "Desde3": "AA",
        "Hasta3": "AB",
        "Desde2": "AC",
        "Hasta2": "AD"
    }


    # Iterar sobre los datos
    for idx, diccionario in enumerate(datos):
        fila_actual = fila_inicio + idx

        # Agregamos el indice de cada ID
        celda = hoja[f"A{fila_actual}"]
        celda.value = idx+1
        celda.font = Font(name="Arial", size=8, bold=True)
        celda.alignment = Alignment(horizontal="center", vertical="center")

        ultima_fila = fila_actual  # Actualizar la última fila usada
        color_fondo = color_fondo_1 if idx % 2 == 0 else color_fondo_2  # Alternar colores

        # Iterar sobre cada clave y asignar los valores en la hoja
        for key, value in diccionario.items():
            if key in mapeo_columnas:
                columna = mapeo_columnas[key]
                celda = hoja[f"{columna}{fila_actual}"]
                celda.value = value
                celda.font = Font(name="Arial", size=8, bold=True)
                celda.alignment = Alignment(horizontal="center", vertical="center")
                celda.fill = color_fondo  # Aplicar color de fondo

                if columna =="H" or columna  =="L" or columna == "M" or columna == "O" or columna == "P" or columna == "R" or columna == "S" or columna == "U"or columna == "W" or columna == "X" or columna == "AA" or columna == "AB" or columna == "AC" or columna == "AD":
                    celda.number_format = '0.00'

    return hoja, ultima_fila

def aplicar_borde_a_rango(hoja, celda_inicio, celda_fin, header=None):
    # Extraer la parte numérica de la celda
    fila_inicio = int(re.search(r'\d+', celda_inicio).group())
    fila_fin = int(re.search(r'\d+', celda_fin).group())
    columna_inicio = re.search(r'[A-Z]+', celda_inicio).group()

    # Definir los bordes
    borde_normal = Border(left=Side(border_style="thin", color="FF000000"))
    borde_completo = Border(
        left=Side(border_style="thin", color="FF000000"),
        top=Side(border_style="thin", color="FF000000"),
        right=Side(border_style="thin", color="FF000000"),

    )

    for fila in range(fila_inicio, fila_fin + 1):
        celda = hoja[f"{columna_inicio}{fila}"]
        
        if header and fila == fila_inicio:  # Si es la fila del header
            celda.border = borde_completo
        else:
            celda.border = borde_normal  # Bordes normales para otras filas
    
    return hoja

def construir_data(reportes_agrupados,nombres_programas):
    opciones_estado = {
            '1':'Abortado',
            '2':'En Avance',
            '3':'En Espera',
            '4':'Finalizado',
            }
    
    resultado = []
    for reporte, detalles in reportes_agrupados.items():

        detalles_ordenados = sorted(
            detalles,
            key=lambda x: (float(x['metroFinal']), -float(x['metroInicial']))
        )
        datos_finales = detalles_ordenados[-1]

        fondo_teorico = float(datos_finales['recomendacion']['largo_programado'])
        fondo_real = float(datos_finales['totalPerforado'])
        faltante = fondo_teorico-fondo_real

        if fondo_teorico == 0 or fondo_real == 0:
            porcentaje_perforado = 'SIN TEORICO ASIGNADO'
        else:   
            porcentaje_perforado = float((fondo_real*100)/fondo_teorico)

        id_estado = datos_finales['recomendacion'].get('estado')

        if id_estado:
    
            estado = opciones_estado[datos_finales['recomendacion']['estado']]
            
        else:

            estado = "SIN ESTADO"

        resultado.append({
            "Rec": datos_finales['recomendacion']['recomendacion'],
            "Programa": datos_finales['recomendacion']['programa'],
            "Sondaje": datos_finales['recomendacion']['pozo'],
            "Estado de Pago": "SIN EP'S",
            "Sonda": datos_finales['nombre_sonda'],
            "Levantamiento de Collar": "SIN LEVANTAMIENTO",
            "Azimut": datos_finales['recomendacion']['azimut'],
            "Inclinación": datos_finales['recomendacion']['inclinacion'],
            "Inicio": datos_finales['recomendacion']['fecha_inicio'],
            "Término": "SIN TÉRMINO",
            "Estado": estado,
            "Largo Programado": fondo_teorico,
            "Fondo Final": fondo_real,
            "Nº Bandejas": "SIN DATO",
            "Faltante": fondo_teorico-fondo_real,
            "% Perforado": porcentaje_perforado,
            "Medición": "SIN DATO",
            "Desde": "SIN DATO",
            "Hasta": "SIN DATO",
            "Certificado Medición de Trayectoria": "SIN DATO",
            "% Recuperación de Sondajes": "SIN DATO",
            "% Rendimiento Sondajes": "SIN DATO",
            "% Desviación": "SIN DATO",
            "Tricono": "SIN DATO",
            "Fotografía": "SIN DATO",
            "Corte": "SIN DATO",
            "Desde3": "SIN DATO",
            "Hasta3": "SIN DATO",
            "Desde2": "SIN DATO",
            "Hasta2": "SIN DATO"
        })

    return resultado
def run(libro,reportes_agrupados,nombres_programas):

    data = construir_data(reportes_agrupados,nombres_programas)
    #data = consumir_datos_api()
    #data = None

#     if not data:
#         print("No hay datos disponibles en REPORTE AVANCE MUESTRERA.")
#     data = [
#     {
#         "Rec": "GM29_04",
#         "Programa": "GEOMETALÚRGICO",
#         "Sondaje": "DDH3866",
#         "Estado de Pago": "Octubre / Noviembre / Diciembre",
#         "Sonda": "CH1-99/77",
#         "Levantamiento de Collar": "SI",
#         "Azimut": 270,
#         "Inclinación": -52,
#         "Inicio": "9/23/2023",
#         "Término": "12/13/2023",
#         "Estado": "Finalizado",
#         "Largo Programado": 592.00,
#         "Fondo Final": 100.00,
#         "Nº Bandejas": 100.00,
#         "Faltante": 100.00,
#         "% Perforado": 100.00,
#         "Medición": "Si",
#         "Desde": 0.00,
#         "Hasta": 427.00,
#         "Certificado Medición de Trayectoria": "SI",
#         "% Recuperación de Sondajes": "99%",
#         "% Rendimiento Sondajes": "",
#         "% Desviación": 6.59,
#         "Tricono": 4.50,
#         "Fotografía": 100.00,
#         "Corte": 100.00,
#         "Desde3": 4.50,
#         "Hasta3": 100.00,
#         "Desde2": 4.50,
#         "Hasta2": 100.00
#     },
#     {
#         "Rec": "GM29_65",
#         "Programa": "GEOMETALÚRGICO",
#         "Sondaje": "DDH3867",
#         "Estado de Pago": "Octubre",
#         "Sonda": "CH1-129",
#         "Levantamiento de Collar": "SI",
#         "Azimut": 270,
#         "Inclinación": -59,
#         "Inicio": "9/23/2023",
#         "Término": "10/18/2023",
#         "Estado": "Finalizado",
#         "Largo Programado": 458.00,
#         "Fondo Final": 100.00,
#         "Nº Bandejas": 184,
#         "Faltante": 100.00,
#         "% Perforado": 100.00,
#         "Medición": "Si",
#         "Desde": 0.00,
#         "Hasta": 454.00,
#         "Certificado Medición de Trayectoria": "SI",
#         "% Recuperación de Sondajes": "97%",
#         "% Rendimiento Sondajes": "",
#         "% Desviación": 4.65,
#         "Tricono": 36.00,
#         "Fotografía": 100.00,
#         "Corte": 100.00,
#         "Desde3": 36.00,
#         "Hasta3": 100.00,
#         "Desde2": 36.00,
#         "Hasta2": 100.00
#     },
#     {
#         "Rec": "GM29_03B",
#         "Programa": "GEOMETALÚRGICO",
#         "Sondaje": "DDH3868",
#         "Estado de Pago": "Octubre / Noviembre",
#         "Sonda": "CH1-130",
#         "Levantamiento de Collar": "SI",
#         "Azimut": 270,
#         "Inclinación": -50,
#         "Inicio": "26/09/2023",
#         "Término": "7/11/2023",
#         "Estado": "Finalizado",
#         "Largo Programado": 599.00,
#         "Fondo Final": 100.00,
#         "Nº Bandejas": 100.00,
#         "Faltante": 100.00,
#         "% Perforado": 100.00,
#         "Medición": "Si",
#         "Desde": 0.00,
#         "Hasta": 570.00,
#         "Certificado Medición de Trayectoria": "SI",
#         "% Recuperación de Sondajes": "99%",
#         "% Rendimiento Sondajes": "",
#         "% Desviación": 6.13,
#         "Tricono": 6.70,
#         "Fotografía": 100.00,
#         "Corte": 100.00,
#         "Desde3": 6.70,
#         "Hasta3": 100.00,
#         "Desde2": 6.70,
#         "Hasta2": 100.00
#     }
# ]
#         # Obtener todas las claves distintas de "mes" en la data 
    
    # Agregar una nueva hoja 
    hoja = libro.create_sheet(title="AVANCE MUESTRERA")

    hoja.merge_cells("AA1:AB1")
    helpers.agregar_titulos(hoja, "AA", 1, ["Mapeo Geológico"],"70ad47", 1,15,"","letras_negras")
    hoja.merge_cells("AC1:AD1")
    helpers.agregar_titulos(hoja, "AC", 1, ["Envío a Preparación Mécanica"],"FFFF00", 1,15,"","letras_negras")
    titulos = ["Rec", "Programa", "Sondaje", "Estado de Pago", "Sonda", "Levantamiento de Collar", "Azimut", "Inclinación", "Inicio", "Término", "Estado", "Largo Programado", "Fondo Final", "Nº Bandejas", "Faltante", "% Perforado", "Medición", "Desde", "Hasta", "Certificado Medición de Trayectoria", "% Recuperación de Sondajes", "% Rendimiento Sondajes", "% Desviación", "Tricono", "Fotografía", "Corte", "Desde3", "Hasta3", "Desde2", "Hasta2"]
    helpers.agregar_titulos(hoja, "A", 2, titulos,"70ad47", 2,15,"muestrera")

    hoja , ultima_fila = agregar_datos_a_celdas(hoja, "A3", data, "e2efd9", "b4c6e7")

    hoja = aplicar_borde_a_rango(hoja,"Y2","Y2","header")
    hoja = aplicar_borde_a_rango(hoja,"Z2","Z2","header")
    hoja = aplicar_borde_a_rango(hoja,"Y3",f'Y{ultima_fila}')
    hoja = aplicar_borde_a_rango(hoja,"Z3",f'Z{ultima_fila}')
    hoja = aplicar_borde_a_rango(hoja,"AA1",f'AA{ultima_fila}')
    hoja = aplicar_borde_a_rango(hoja,"AC1",f'AC{ultima_fila}')
    hoja = aplicar_borde_a_rango(hoja,"AE1",f'AE{ultima_fila}')
    
    libro._sheets.insert(0, libro._sheets.pop(-1))


    return libro