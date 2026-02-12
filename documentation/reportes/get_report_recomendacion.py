from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime, timedelta
import locale
import pytz
import requests 
from pathlib import Path
from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn
from collections import defaultdict
from openpyxl.drawing.image import Image
from openpyxl.formatting.rule import CellIsRule
from pprint import pprint
from openpyxl import load_workbook
from documentation.reportes import helpers


def agregar_titulos_tablas(hoja):
    
    # Agregar TITULOS
    hoja.merge_cells("D1:H1")
    hoja = helpers.agregar_titulos(hoja, "D", 1, ["RECOMENDACIÓN PROGRAMADA"],"FFFF00", alto_fila_factor=2, ancho_columna=8,color_letras= "negro")

    hoja.merge_cells("I1:M1")
    hoja = helpers.agregar_titulos(hoja, "I", 1, ["AJUSTE DE RECOMENDACIÓN"],"70ad47", alto_fila_factor=2, ancho_columna=8)

    hoja.merge_cells("N1:Q1")
    hoja = helpers.agregar_titulos(hoja, "N", 1, ["RECOMENDACIÓN FINAL (CERTIFICADO)"],"FFFF00", alto_fila_factor=2, ancho_columna=8,color_letras= "negro")
    
    hoja.merge_cells("S1:U1")
    hoja = helpers.agregar_titulos(hoja, "S", 1, ["VALIDADOR (NO MAYOR A 20 MTS,) REC_PROGRAMADA VS, REC_ FINAL"],"70ad47", alto_fila_factor=2, ancho_columna=8)
    
    titulos_primera_tabla = [
        "SONDAJE","RECOM","SECTOR",
        "ESTE LOCAL","NORTE LOCAL","COTA (m,s,n,m)","AZIMUT (°)","MANTEO (°)",
        "ESTE LOCAL2","NORTE LOCAL2","COTA (m,s,n,m)2","AZIMUT (°)2","MANTEO (°)2",
        "ESTE","NORTE","COTA","FECHA"]
    hoja = helpers.agregar_titulos(hoja, "A", 2, titulos_primera_tabla,"70ad47", alto_fila_factor=2, ancho_columna=8,color_letras= "negro")

    titulos_segunda_tabla = ["ESTE","NORTE","COTA"]
    hoja = helpers.agregar_titulos(hoja, "S", 2, titulos_segunda_tabla,"70ad47", alto_fila_factor=2, ancho_columna=8)

    return hoja

def agregar_datos_a_celdas(hoja, inicio_celda, datos, primer_color, segundo_color,segunda_tabla = None):
    # Obtener la fila de inicio desde la celda (ej. "A1" -> 1)
    fila_inicio = int(inicio_celda[1:])

    # Definir los colores alternados para las filas
    color_fondo_1 = PatternFill(start_color=primer_color, end_color=primer_color, fill_type="solid")  # Verde
    color_fondo_2 = PatternFill(start_color=segundo_color, end_color=segundo_color, fill_type="solid")  # Rojo
    ultima_fila = fila_inicio  # Variable para almacenar la última fila usada

    # Mapeo de claves del JSON a columnas de Excel


    mapeo_columnas = {
        "SONDAJE": "A",
        "RECOM": "B",
        "SECTOR": "C",
        "ESTE LOCAL": "D",
        "NORTE LOCAL": "E",
        "COTA (m,s,n,m)": "F",
        "AZIMUT (°)": "G",
        "MANTEO (°)": "H",
        "ESTE LOCAL2": "I",
        "NORTE LOCAL2": "J",
        "COTA (m,s,n,m)2": "K",
        "AZIMUT (°)2": "L",
        "MANTEO (°)2": "M",
        "ESTE": "N",
        "NORTE": "O",
        "COTA": "P",
        "FECHA": "Q"
    }

    if segunda_tabla:
            mapeo_columnas = {
                "ESTE": "S",
                "NORTE": "T",
                "COTA": "U",
            }
    # Iterar sobre los datos
    for idx, diccionario in enumerate(datos):
        fila_actual = fila_inicio + idx
        ultima_fila = fila_actual  # Actualizar la última fila usada
        color_fondo = color_fondo_1 if idx % 2 == 0 else color_fondo_2  # Alternar colores

        # Iterar sobre cada clave y asignar los valores en la hoja
        for key, value in diccionario.items():

            if key in mapeo_columnas:
                columna = mapeo_columnas[key]

                celda = hoja[f"{columna}{fila_actual}"]
                celda.value = value

                # Aplicar formato a la celda
                celda.font = Font(name="Arial", size=8, bold=False)
                celda.alignment = Alignment(horizontal="center", vertical="center")
                if segunda_tabla and (value <= -20 or value >= 20):
                    celda.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Aplicar color de rojo
                else:
                    celda.fill = color_fondo  # Aplicar color de fondo



    return hoja, ultima_fila

def aplicar_borde_a_rango(hoja, celda_inicio, celda_fin, color_borde="00b050"):

    # Convertir las referencias de celdas a índices
    fila_inicio = int(celda_inicio[1:])  # La fila es el número después de la letra
    col_inicio = column_index_from_string(celda_inicio[0])  # La columna es el índice de la letra

    fila_fin = int(celda_fin[1:])  # La fila es el número después de la letra
    col_fin = column_index_from_string(celda_fin[0])  # La columna es el índice de la letra

    # Definir el borde
    borde = Border(
        left=Side(border_style="thick", color=color_borde),  # Borde izquierdo
        right=Side(border_style="thick", color=color_borde),  # Borde derecho
        top=Side(border_style="thick", color=color_borde),  # Borde superior
        bottom=Side(border_style="thick", color=color_borde)  # Borde inferior
    )

    # Aplicar los bordes solo al perímetro (exterior) del rango
    for fila in range(fila_inicio, fila_fin + 1):
        for columna in range(col_inicio, col_fin + 1):
            celda = hoja.cell(row=fila, column=columna)
            
            # Si estamos en la primera o última fila o columna, aplicar el borde
            if columna == col_inicio:
                celda.border = Border(left=Side(border_style="thin", color=color_borde))
            elif columna == col_fin:
                celda.border = Border(right=Side(border_style="thin", color=color_borde))

    
    return hoja

def obtener_nombre_sondas(reportes_agrupados):

    nombres_sondas = []
    for sonda, detalle in reportes_agrupados.items():

        nombres_sondas.append(sonda)

    return nombres_sondas

def obtener_ajustes(all_recomendaciones_ajuste,id):
    este_local_2 = float(0.00)
    norte_local_2 = float(0.00)
    cota_2 = float(0.00)
    azimut_2 = int(0)
    manteo_2 = int(0)

    for rec in all_recomendaciones_ajuste:

        if 'recomendacionAjuste' in rec and rec['recomendacionAjuste'] == id:
            este_local_2 = float(rec['esteAjuste'])
            norte_local_2 = float(rec['norteAjuste'])
            cota_2 = float(rec['cotaAjuste'])
            azimut_2 = int(rec['azimutAjuste'])
            manteo_2 = float(rec['manteoAjuste'])

            return este_local_2,norte_local_2,cota_2,azimut_2,manteo_2
        
    return este_local_2,norte_local_2,cota_2,azimut_2,manteo_2

def obtener_final(all_recomendaciones_final,id):
    este_final = float(0.00)
    norte_final = float(0.00)
    cota_final = float(0.00)
    fecha_final = "SIN DATO"

    for rec in all_recomendaciones_final:

        if 'recomendacionFinal' in rec and rec['recomendacionFinal'] == id:
            este_final = float(rec['esteFinal'])
            norte_final = float(rec['norteFinal'])
            cota_final = float(rec['cotaFinal'])
            fecha_final_str = str(rec['fechaFinal']).split('T')[0]
            fecha_final = datetime.strptime(fecha_final_str, '%Y-%m-%d').strftime('%d-%m-%Y')

            return este_final,norte_final,cota_final,fecha_final

    return este_final,norte_final,cota_final,fecha_final
def construir_data(reportes_agrupados,all_recomendaciones,sondas,all_recomendaciones_ajuste, all_recomendaciones_final):
    primera_tabla=[]
    segunda_tabla=[]

    nombres_sondajes = obtener_nombre_sondas(reportes_agrupados)

    for sondaje in nombres_sondajes:

        rec = "SIN RECOMENDACION"
        sector = "SIN SECTOR"

        este_local = float(0.00)
        norte_local = float(0.00)
        cota_local = float(0.00)
        azimut_local = int(0)
        manteo_local = int(0)

        este_local_2 = float(0.00)
        norte_local_2 = float(0.00)
        cota_2 = float(0.00)
        azimut_2 = int(0)
        manteo_2 = int(0)

        este_final = float(0.00)
        norte_final = float(0.00)
        cota_final = float(0.00)
        fecha_final = "SIN DATO"
        
        for recomendacion in all_recomendaciones:

            if recomendacion["pozo"] == sondaje:

                rec = recomendacion['recomendacion']
                sector = recomendacion['sector']
                este_local = float(recomendacion['este'])
                norte_local = float(recomendacion['norte'])
                cota_local = float(recomendacion['cota'])
                azimut_local = int(recomendacion['azimut'])
                manteo_local = float(recomendacion['inclinacion'])

                if all_recomendaciones_ajuste:
                    este_local_2,norte_local_2,cota_2,azimut_2,manteo_2 = obtener_ajustes(all_recomendaciones_ajuste,recomendacion['id'])

                if all_recomendaciones_final:

                    este_final,norte_final,cota_final,fecha_final = obtener_final(all_recomendaciones_final,recomendacion['id'])


                break

        primera_tabla.append({
                        "SONDAJE": sondaje,
                        "RECOM": rec,
                        "SECTOR": sector,
                        # recomendacion programada
                        "ESTE LOCAL": este_local,
                        "NORTE LOCAL": norte_local,
                        "COTA (m,s,n,m)": cota_local,
                        "AZIMUT (°)": azimut_local,
                        "MANTEO (°)": manteo_local,
                        # ajuste de recomendacion
                        "ESTE LOCAL2": este_local_2,
                        "NORTE LOCAL2": norte_local_2,
                        "COTA (m,s,n,m)2": cota_2,
                        "AZIMUT (°)2": azimut_2,
                        "MANTEO (°)2": manteo_2,
                        # recomendacion final
                        "ESTE": este_final,
                        "NORTE": norte_final,
                        "COTA": cota_final,
                        "FECHA": fecha_final
                    })

        segunda_tabla.append({
                        "ESTE": este_final-este_local_2,
                        "NORTE": norte_final-norte_local_2,
                        "COTA": cota_final-cota_2
                    })

    return primera_tabla,segunda_tabla
    
def run(libro,reportes_agrupados,all_recomendaciones,sondas,all_recomendaciones_ajuste, all_recomendaciones_final):

    primera_tabla, segunda_tabla = construir_data(reportes_agrupados,all_recomendaciones,sondas, all_recomendaciones_ajuste, all_recomendaciones_final)

    # Obtener todas las claves distintas de "mes" en la data 
    
    # Agregar una nueva hoja 
    hoja = libro.create_sheet(title="RECOMENDACIÓN")

    hoja = agregar_titulos_tablas(hoja)

    hoja,ultima_fila = agregar_datos_a_celdas(hoja,"A3",primera_tabla,"e2efd9", "b4c6e7")
    hoja,ultima_fila = agregar_datos_a_celdas(hoja,"S3",segunda_tabla,"e2efd9", "b4c6e7","seguna_tabla")

    hoja = aplicar_borde_a_rango(hoja,"D1",f'H{ultima_fila}')
    hoja = aplicar_borde_a_rango(hoja,"I1",f'M{ultima_fila}')
    hoja = aplicar_borde_a_rango(hoja,"N1",f'Q{ultima_fila}')
    hoja = aplicar_borde_a_rango(hoja,"S1",f'U{ultima_fila}')

    # Ajustar el ancho de las columnas
    hoja.column_dimensions["A"].width = 8
    hoja.column_dimensions["B"].width = 10
    hoja.column_dimensions["C"].width = 13
    hoja.column_dimensions["D"].width = 6
    hoja.column_dimensions["E"].width = 6
    hoja.column_dimensions["F"].width = 6
    hoja.column_dimensions["G"].width = 10
    hoja.column_dimensions["H"].width = 10
    hoja.column_dimensions["I"].width = 6
    hoja.column_dimensions["J"].width = 6
    hoja.column_dimensions["K"].width = 8
    hoja.column_dimensions["L"].width = 8
    hoja.column_dimensions["M"].width = 8
    hoja.column_dimensions["N"].width = 7
    hoja.column_dimensions["O"].width = 7
    hoja.column_dimensions["P"].width = 7
    hoja.column_dimensions["Q"].width = 13
    hoja.column_dimensions["R"].width = 1
    hoja.column_dimensions["S"].width = 10
    hoja.column_dimensions["T"].width = 10
    hoja.column_dimensions["U"].width = 10


    libro._sheets.insert(0, libro._sheets.pop(-1))

    return libro