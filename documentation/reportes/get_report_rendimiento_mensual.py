from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import locale
import pytz
import requests 
from pathlib import Path
from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn
from collections import defaultdict
from openpyxl.drawing.image import Image
from openpyxl.formatting.rule import CellIsRule
from pprint import pprint
from openpyxl import load_workbook
from ..reportes import helpers
import copy


titulos = ["Fecha","Turno","Sonda","Rec","Sondaje","Largo Programado","Diametro","Desde (m)","Hasta (m)","Total Día (m)","Faltante","Programa"]

def agregar_datos_a_celdas(hoja, inicio_celda, datos, primer_color, segundo_color):
    # Obtener la fila de inicio desde la celda (ej. "A1" -> 1)
    fila_inicio = int(inicio_celda[1:])

    # Definir los colores alternados para las filas
    color_fondo_1 = PatternFill(start_color=primer_color, end_color=primer_color, fill_type="solid")  # Verde
    color_fondo_2 = PatternFill(start_color=segundo_color, end_color=segundo_color, fill_type="solid")  # Rojo

    ultima_fila = fila_inicio  # Variable para almacenar la última fila usada
    sonda_actual = None

    # Mapeo de claves del JSON a columnas de Excel
    mapeo_columnas = {
        "Fecha": "B",
        "Turno": "C",
        "Sonda": "D",
        "Rec": "E",
        "Sondaje": "F",
        "Largo Programado": "G",
        "Diametro": "H",
        "Desde (m)": "I",
        "Hasta (m)": "J",
        "Total Día (m)": "K",
        "Faltante": "L",
        "Programa": "M"
    }

    # Iterar sobre los datos
    for idx, diccionario in enumerate(datos):
        fila_actual = fila_inicio + idx
        ultima_fila = fila_actual  # Actualizar la última fila usada
        color_fondo = color_fondo_1 if idx % 2 == 0 else color_fondo_2  # Alternar colores

        # Iterar sobre cada clave y asignar los valores en la hoja
        for key, value in diccionario.items():
            if key in mapeo_columnas:
                columna = mapeo_columnas[key]
                celda = hoja[f"{columna}{fila_actual}"]
                celda.value = value

                # Aplicar formato a la celda
                celda.font = Font(name="Arial", size=8, bold=False)
                celda.alignment = Alignment(horizontal="center", vertical="center")
                celda.fill = color_fondo  # Aplicar color de fondo

            if sonda_actual != diccionario["Sonda"]:
                sonda_actual = diccionario["Sonda"]
                celda_o = hoja[f"O{fila_actual}"]
                celda_o.value = "Inicio"
                celda_o.font = Font(name="Arial", size=8, bold=True)
                celda_o.alignment = Alignment(horizontal="center", vertical="center")
                celda_o.fill = PatternFill(start_color="ffc000", end_color="ffc000", fill_type="solid")

                if idx > 0:
                    celda_o = hoja[f"O{fila_actual-1}"]
                    celda_o.value = "Finalizado"
                    celda_o.font = Font(name="Arial", size=8, bold=True)
                    celda_o.alignment = Alignment(horizontal="center", vertical="center")
                    celda_o.fill = PatternFill(start_color="66ff33", end_color="66ff33", fill_type="solid")

            if columna =="G" or columna  =="I" or columna  =="J" or columna  =="K" or columna  =="L" :
                    celda.number_format = '0.00'

    return hoja, ultima_fila


# Función de ordenamiento
def orden_clave(item):
    fecha = datetime.strptime(item['Fecha'], "%d/%m/%Y")
    turno_valor = 0 if item['Turno'].lower() == 'día' else 1
    return (item['Sonda'], fecha, turno_valor, item['Desde (m)'])
def construir_data(reportes_agrupados,nombres_programas,sondajes_recomendaciones):

    resultado = []



    for reporte, detalles in reportes_agrupados.items():

        for detalle in detalles:

            if 'detalle_perforaciones' not in detalle:
                continue
            for perforacion in detalle['detalle_perforaciones']:
                fecha_detalle = perforacion['FECHA']
                fecha = datetime.strptime(fecha_detalle, '%d/%m/%Y')
                # Obtener el mes actual
                mes_actual = datetime.now().month

                if fecha.month != mes_actual:
                    continue
                
                if perforacion['TURNO'] == "A":
                    turno = "Día"
                else:
                    turno = "Noche"

                desde = float(perforacion['DESDE'])
                hasta = float(perforacion['HASTA'])
                total_dia = desde + hasta
                largo_programado = float(detalle['recomendacion']['largo_programado'])

                resultado.append({
                    "Fecha": perforacion['FECHA'],
                    "Turno": turno,
                    "Sonda": detalle["nombre_sonda"],
                    "Rec": detalle['recomendacion']['recomendacion'],
                    "Sondaje": detalle['recomendacion']['pozo'],
                    "Largo Programado": largo_programado ,
                    "Diametro": perforacion['DIÁMETRO'],
                    "Desde (m)": desde,
                    "Hasta (m)": hasta,
                    "Total Día (m)": total_dia,
                    "Faltante": largo_programado - hasta,
                    "Programa": detalle['recomendacion']['programa'],
                })

    # Ordenar la lista
    datos_ordenados = sorted(resultado, key=orden_clave)

    return datos_ordenados




def run(libro,reportes_agrupados,nombres_programas,sondajes_recomendaciones):


    data = construir_data(reportes_agrupados,nombres_programas,sondajes_recomendaciones)



#     if not data:
#         print("No hay datos disponibles en REPORTE RENDIMIENTO MENSUAL.")
#         data = [
#     {
#         "Fecha": "23/09/2023",
#         "Turno": "Día",
#         "Sonda": "DEMO-99",
#         "Rec": "DEMO_04",
#         "Sondaje": "DDH-3866",
#         "Largo Programado": 592.00,
#         "Diametro": "",
#         "Desde (m)": 0.00,
#         "Hasta (m)": 0.00,
#         "Total Día (m)": 0.00,
#         "Faltante": 0.00,
#         "Programa": "GEOMETALÚRGICO",
        
#     },
#     {
#         "Fecha": "23/09/2023",
#         "Turno": "Noche",
#         "Sonda": "DEMO-99",
#         "Rec": "DEMO_04",
#         "Sondaje": "DDH-3866",
#         "Largo Programado": 592.00,
#         "Diametro": "",
#         "Desde (m)": 0.00,
#         "Hasta (m)": 0.00,
#         "Total Día (m)": 0.00,
#         "Faltante": 0.00,
#         "Programa": "GEOMETALÚRGICO"
#     },
#     {
#         "Fecha": "24/09/2023",
#         "Turno": "Día",
#         "Sonda": "DEMO-100",
#         "Rec": "DEMO_100",
#         "Sondaje": "DDH-3866",
#         "Largo Programado": 592.00,
#         "Diametro": "",
#         "Desde (m)": 0.00,
#         "Hasta (m)": 0.00,
#         "Total Día (m)": 0.00,
#         "Faltante": 0.00,
#         "Programa": "GEOMETALÚRGICO"
#     },
#     {
#         "Fecha": "24/09/2023",
#         "Turno": "Noche",
#         "Sonda": "DEMO-100",
#         "Rec": "DEMO_100",
#         "Sondaje": "DDH-3866",
#         "Largo Programado": 592.00,
#         "Diametro": "Tricono",
#         "Desde (m)": 0.00,
#         "Hasta (m)": 4.50,
#         "Total Día (m)": 0.00,
#         "Faltante": 0.00,
#         "Programa": "GEOMETALÚRGICO"
#     }
# ]

    # Agregar una nueva hoja 
    hoja = libro.create_sheet(title="Rendimiento Mensual")

    hoja = helpers.agregar_titulos(hoja, "B", 2, titulos, "70ad47",2,14)

    # Ajustar el ancho de las columnas
    hoja.column_dimensions["A"].width = 2
    hoja.column_dimensions["N"].width = 2

    hoja = agregar_datos_a_celdas(hoja, "B3", data, "e2efd9", "b4c6e7")
    # Mover la nueva hoja a la primera posición
    libro._sheets.insert(0, libro._sheets.pop(-1))

    return libro