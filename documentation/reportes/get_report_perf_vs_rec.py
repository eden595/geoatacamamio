from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import locale
import pytz
import requests 
from pathlib import Path
from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn
from collections import defaultdict
from openpyxl.drawing.image import Image
from openpyxl.formatting.rule import CellIsRule
from pprint import pprint
from openpyxl import load_workbook
from ..reportes import helpers
import copy
from drilling.models import ReportesOperacionales

def agregar_datos_a_celdas(hoja, inicio_celda, datos, primer_color,segundo_color):
    # Obtener la fila y columna de la celda de inicio
    fila_inicio = int(inicio_celda[1:])  # Extraer la fila desde la celda (A1 -> 1)

    # Definir los colores alternados para las filas
    color_fondo_1 = PatternFill(start_color=primer_color, end_color=primer_color, fill_type="solid")  # Verde
    color_fondo_2 = PatternFill(start_color=segundo_color, end_color=segundo_color, fill_type="solid")  # Rojo

    ultima_fila = fila_inicio  # Variable para almacenar la última fila usada

    pozo_actual = None
    # Iterar sobre los diccionarios en los datos (lista de diccionarios)
    for idx, diccionario in enumerate(datos):
        # Determinar la fila donde se agregará este diccionario
        fila_actual = fila_inicio + idx
        ultima_fila = fila_actual  # Actualizar la última fila usada
        if idx % 2 == 0:
            color_fondo = color_fondo_1  # Verde para filas pares
        else:
            color_fondo = color_fondo_2  # Rojo para filas impares
        # Iterar sobre las claves del diccionario y asignar los valores a las celdas
        for j, (key, value) in enumerate(diccionario.items()):
            if key == "POZO":
                columna = "B" 
            if key == "DESDE":
                columna = "C"
            if key == "HASTA":
                columna = "D"
            if key == "PERFORADO":
                columna = "E"
            if key == "RECUPERADO":
                columna = "F"
            if key == "% RECUPERADO":
                columna = "G"
            if key == "Descuentos":
                columna = "H"
            if key == "FECHA":
                columna = "I"
            if key == "TURNO":
                columna = "J"
            if key == "DIÁMETRO":
                columna = "K"

            celda = hoja[f"{columna}{fila_actual}"]  # Obtener la celda
            celda.value = value  # Asignar el valor del diccionario

            # Aplicar formato a la celda
            celda.font = Font(name="Arial", size=8, bold=False)
            celda.alignment = Alignment(horizontal="center", vertical="center")
            # Aplicar el color de fondo a la celda
            celda.fill = color_fondo

            if key == "DESDE" or key == "HASTA" or key == "PERFORADO" or key == "RECUPERADO":
                celda.number_format = '#,##0.00'
            if key == "Descuentos":
                celda.number_format = '0.00%'
                
        if pozo_actual != diccionario["POZO"]:
            pozo_actual = diccionario["POZO"]
            celda_m = hoja[f"A{fila_actual}"]
            celda_m.value = "Inicio"
            celda_m.font = Font(name="Arial", size=8, bold=True)
            celda_m.alignment = Alignment(horizontal="center", vertical="center")
            celda_m.fill = PatternFill(start_color="ffc000", end_color="ffc000", fill_type="solid")

        
        # Agregar el valor 2 en la columna M

        celda_m = hoja[f"M{fila_actual}"]
        celda_m.value = diccionario["PERFORADO"]-(diccionario["PERFORADO"]*diccionario["Descuentos"])
        celda_m.font = Font(name="Arial", size=8, bold=False)
        celda_m.alignment = Alignment(horizontal="center", vertical="center")
        celda_m.number_format = '#,##0.00'
        


    return hoja, ultima_fila

def buscar_detalle_perforaciones(id, detalles_perforaciones,pozo,turno,diametros_data):

    resultado =[]
    for detalle in detalles_perforaciones:

        if detalle['reporte'] == id:

            if detalle['diametros'] in diametros_data:
                diametro  = diametros_data[detalle['diametros']]

            else:
                diametro = "Desconocido"

            fecha_iso = detalle['fechacreacion']
            # Parseamos la fecha
            fecha_obj = datetime.fromisoformat(fecha_iso)
            # Formateamos como día/mes/año sin ceros a la izquierda
            fecha_formateada = f"{fecha_obj.day}/{fecha_obj.month}/{fecha_obj.year}"
            
            if float(detalle['porcentajeRecuperacion']) < 85.00:
                descuento = 100
            elif 90.00 > float(detalle['porcentajeRecuperacion']) >= 85.00:
                descuento = 10
            elif 95.00 > float(detalle['porcentajeRecuperacion']) >= 90.00:
                descuento = 7
            elif float(detalle['porcentajeRecuperacion']) >= 95.00:
                descuento = 0

            resultado.append({
                "POZO": pozo,
                "DESDE": float(detalle['desde']),
                "HASTA": float(detalle['hasta']),
                "PERFORADO": float(detalle['perforado']),
                "RECUPERADO": float(detalle['recuperacion']),  
                "% RECUPERADO": float(detalle['porcentajeRecuperacion']),
                "Descuentos": descuento/100,
                "FECHA": fecha_formateada,
                "TURNO": turno,
                "DIÁMETRO": diametro
            })

    return resultado           
def contruir_data(datos, detalles_perforaciones,diametros_data):


    # obtengo todos los detalles de perforacion para cada 
    data_construida = []
    for pozo, detalles_generales in datos.items():

        resultados_pozo = []

        for detalle in detalles_generales:

            # se buscara los id de cada reporte operacional, para obtener  el match con su detalle de perforacion

            if detalle['turno'] == "1":
                turno = "A"
            else:
                turno = "B"

            resultados = buscar_detalle_perforaciones(detalle['id'], detalles_perforaciones,pozo, turno,diametros_data)

            if resultados:
                resultados_pozo.extend(resultados)

        #resultados_pozo_ordenado = sorted(resultados_pozo, key=lambda d: float(d['DESDE']))

        data_construida.extend(resultados_pozo)

        detalle['detalle_perforaciones']= resultados_pozo
    return data_construida, datos

def run(libro,reportes_agrupados, reportes_detalles,diametros_data):

    data,reportes_agrupados= contruir_data(reportes_agrupados, reportes_detalles,diametros_data)

    # Agregar una nueva hoja "PERF. VS REC."
    hoja = libro.create_sheet(title="PERF. VS REC.")

    titulos = ["POZO", "DESDE", "HASTA", "PERFORADO", "RECUPERADO", "% RECUPERADO","Descuentos","FECHA", "TURNO", "DIÁMETRO"]

    hoja = helpers.agregar_titulos(hoja, "B", 1, titulos, "70ad47",1,12)

    celda = hoja[f"M1"] 
    celda.value = "Valor con descuento" 

    hoja, ultima_fila = agregar_datos_a_celdas(hoja, "B2", data, "e2efd9", "b4c6e7")

    ###### ********** FALTA AGREGAR LOS TOTALES ABAJO DE LA TABLA
    
    # Mover la nueva hoja a la primera posición
    libro._sheets.insert(0, libro._sheets.pop(-1))

    return libro,reportes_agrupados