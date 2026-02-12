from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime, timedelta
import locale
import pytz
import requests 
from pathlib import Path
from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn
from collections import defaultdict
from openpyxl.drawing.image import Image
from openpyxl.formatting.rule import CellIsRule
from pprint import pprint
from openpyxl import load_workbook
from dateutil.relativedelta import relativedelta
from collections import OrderedDict
import copy
from ..reportes import helpers

dict_meses = {
        "1": "ene",
        "2": "feb",
        "3": "mar",
        "4": "abr",
        "5": "may",
        "6": "jun",
        "7": "jul",
        "8": "ago",
        "9": "sep",
        "10": "oct",
        "11": "nov",
        "12": "dic",
    }
def agregar_datos_a_celdas(hoja, inicio_celda, datos, primer_color, segundo_color):

    fila_inicio = inicio_celda

    # Definir los colores alternados para las filas
    color_fondo_1 = PatternFill(start_color=primer_color, end_color=primer_color, fill_type="solid")
    color_fondo_2 = PatternFill(start_color=segundo_color, end_color=segundo_color, fill_type="solid")

    ultima_fila = fila_inicio  # Variable para almacenar la última fila usada

    for idx,diccionario in enumerate(datos):
        # Determinar la fila donde se agregará este diccionario
        fila_actual = fila_inicio + idx
        ultima_fila = fila_actual  # Actualizar la última fila usada
        if idx % 2 == 0:
            color_fondo = color_fondo_1  # Verde para filas pares
        else:
            color_fondo = color_fondo_2  # Rojo para filas impares

        for  j, (key, value)in enumerate(diccionario.items()):


            celda = hoja[f"{key}{fila_actual}"]
            celda.value = value
            # Aplicar formato a la celda
            celda.font = Font(name="Arial", size=8, bold=False)
            celda.alignment = Alignment(horizontal="center", vertical="center")
            # Aplicar el color de fondo a la celda
            celda.fill = color_fondo

    return hoja, ultima_fila

def ordenar_data(data, orden_claves):

    data_ordenada = []
    for item in data:
        nuevo_item = {"mes": item["mes"]}  # Mantiene la clave "mes"
        for clave in orden_claves:
            if clave in item:
                nuevo_item[clave] = item[clave]
        data_ordenada.append(nuevo_item)

    return data_ordenada

def obtener_celdas_intermedias(rango):
    inicio, fin = rango.split(":")  # Separar inicio y fin
    col_inicio, fila_inicio = inicio[0], int(inicio[1:])  # Extraer columna y fila inicial
    col_fin, fila_fin = fin[0], int(fin[1:])  # Extraer columna y fila final
    
    cols = range(column_index_from_string(col_inicio), column_index_from_string(col_fin) + 1)
    return [f"{get_column_letter(c)}" for c in cols]

def asignar_columnas(diccionario, celdas_intermedias,graficos_actuales,total_data):
    fila_resultado = {}  # Diccionario para almacenar los resultados

    # La columna 'A' siempre debe contener el valor de 'mes'
    if 'mes' in diccionario:
        fila_resultado['A'] = diccionario['mes']
    
    columnas_disponibles = celdas_intermedias[1:]  # Excluimos 'A' porque es para 'mes'
    indice_columna = 0  # Índice para recorrer la lista de columnas disponibles

    # Iterar sobre cada categoría en el diccionario
    for categoria, subdiccionario in diccionario.items():
        if categoria == 'mes':  # Saltar 'mes' porque ya fue asignado a 'A'
            continue
        if categoria not in graficos_actuales:
            graficos_actuales[categoria] = {"plan":{}, "acumulado_plan":{}}
        
        # Agregar los valores al gráfico
        graficos_actuales[categoria]["plan"]["columna_inicial"]=indice_columna+2
        graficos_actuales[categoria]["plan"]["inicio_filas"]=5
        graficos_actuales[categoria]["plan"]["termino_filas"]=5 + total_data
        graficos_actuales[categoria]["acumulado_plan"]["columna_inicial"]=indice_columna+2
        graficos_actuales[categoria]["acumulado_plan"]["inicio_filas"]=5
        graficos_actuales[categoria]["acumulado_plan"]["termino_filas"]=5 + total_data

        if isinstance(subdiccionario, dict):  # Verificar si es un diccionario anidado
            for subkey, subvalue in subdiccionario.items():
                if indice_columna < len(columnas_disponibles):  # Asegurar que no excedamos las columnas

                    columna = columnas_disponibles[indice_columna]
                    fila_resultado[columna] = subvalue
                    indice_columna += 1  # Avanzar a la siguiente columna

    return fila_resultado, graficos_actuales

def obtener_maqueta_programas(programas):

    nombres_programas = {}
    maqueta_programas = {}

    for programa in programas:
        if programa['programa'] not in maqueta_programas:
            maqueta_programas[programa['programa'].upper()]= {
                "Plan": float(0.00),
                "Real mensual": float(0.00),
                "Acumulado plan": float(0.00),
                "Acumulado real": float(0.00)
            }
            nombres_programas[f'{programa["id"]}'] = programa['programa'].upper()

    return maqueta_programas,nombres_programas

def obtener_maqueta_tabla(maqueta_programas, mes_inicio,anio_inicio,mes_termino,anio_termino):
    abreviaciones_es = ["ene", "feb", "mar", "abr", "may", "jun", 
                        "jul", "ago", "sep", "oct", "nov", "dic"]

    fecha_inicio = datetime(anio_inicio, mes_inicio, 1)
    fecha_termino = datetime(anio_termino, mes_termino, 1)

    resultado = []

    while fecha_inicio <= fecha_termino:
        mes_str = abreviaciones_es[fecha_inicio.month - 1]
        anio_str = str(fecha_inicio.year)[-2:]

        current_dict = OrderedDict()
        current_dict["mes"] = f"{mes_str}-{anio_str}"
        for k, v in maqueta_programas.items():
            current_dict[k] = copy.deepcopy(v)  

        resultado.append(current_dict)
        fecha_inicio += relativedelta(months=1)

    return resultado

def agregamos_planificacion_maqueta(fecha_programa,nombre_programa,maqueta,plan):

    if plan == None:
        plan = float(0.00)
    
    for m in maqueta:
        if m['mes'] == fecha_programa:
            m[nombre_programa]["Plan"] = float(plan)
            return maqueta

def obtener_planificacion_programas(planificacion_programas,nombres_programas,maqueta):
    abreviaciones = ["ene", "feb", "mar", "abr", "may", "jun", 
                        "jul", "ago", "sep", "oct", "nov", "dic"]
    

    for programa in planificacion_programas:

        mes_programa = abreviaciones[int(programa["mes"])-1]
        anio_programa = str(programa["ano"])[-2:]

        fecha_programa = f"{mes_programa}-{anio_programa}"
        nombre_programa = nombres_programas[f'{programa["programa"]}']
        
        maqueta = agregamos_planificacion_maqueta(fecha_programa,nombre_programa,maqueta,programa["plan"])

    return maqueta

def obtener_programa(programas,sonda_actual,sondajes_recomendaciones):

    for pozo, detalles in sondajes_recomendaciones.items():

        for detalle in detalles:

            if detalle['nombre_sonda'] == sonda_actual:
                return detalle['recomendacion']['programa']
                
def agregamos_avence_real_maqueta(fecha_programa,nombre_programa,maqueta,real):

    if real == None:
        real = float(0.00)
    
    for m in maqueta:

        if m['mes'] == fecha_programa['mes']:

            m[nombre_programa.upper()]["Real mensual"] += float(real)

            return maqueta

def obtener_avance_real_programas(datos_mensuales,programas,maqueta,sondajes_recomendaciones):
    abreviaciones = ["ene", "feb", "mar", "abr", "may", "jun", 
                        "jul", "ago", "sep", "oct", "nov", "dic"]
    

    for fecha, dias in datos_mensuales.items():
        nombre_mes = fecha[:3].lower()    
        ultimos_digitos_anio = fecha[-2:] 
        fecha_actual = {
            "mes": f"{nombre_mes}-{ultimos_digitos_anio}",
        }

        for d in dias:

            for key , value in d.items():

                if "-" in key and "(" not in key:

                    programa = obtener_programa(programas,key,sondajes_recomendaciones)

                    if programa:
                        
                        maqueta = agregamos_avence_real_maqueta(fecha_actual,programa,maqueta,value)

    return maqueta

def obtener_acumulados(maqueta):
    acumulados = {}

    for mes in maqueta:

        for key, values in mes.items():
            if key == "mes":
                continue

            if key not in acumulados:
                acumulados[key] = {
                    "Acumulado plan": values["Plan"],
                    "Acumulado real": values["Real mensual"]
                }
            else:
                acumulados[key]["Acumulado plan"] += values["Plan"]
                acumulados[key]["Acumulado real"] += values["Real mensual"]

            # Asignar los acumulados al programa en ese mes
            values["Acumulado plan"] = acumulados[key]["Acumulado plan"]
            values["Acumulado real"] = acumulados[key]["Acumulado real"]

    return maqueta

def construir_data(mes_inicio,anio_inicio,mes_termino,anio_termino,programas,
                   planificacion_programas,datos_mensuales,sondajes_recomendaciones):

    maqueta_programas,nombres_programas = obtener_maqueta_programas(programas)

    maqueta = obtener_maqueta_tabla(maqueta_programas, mes_inicio,anio_inicio,mes_termino,anio_termino)

    maqueta = obtener_planificacion_programas(planificacion_programas,nombres_programas,maqueta)


    maqueta = obtener_avance_real_programas(datos_mensuales,programas,maqueta,sondajes_recomendaciones)

    maqueta = obtener_acumulados(maqueta)

    return  maqueta,nombres_programas


def run(libro,campanas,programas,planificacion_programas,datos_mensuales,sondajes_recomendaciones):
    
    mes_inicio = 1
    anio_inicio = 2025
    mes_termino = 12
    anio_termino = 2025


    # se busca encontrar el año mas bajo de las campañas para poder obtener los años de inicio y final
    for c in campanas:
        if int(c['anoInicial']) < anio_inicio:
            anio_inicio = int(c['anoInicial'])

        if int(c['anoFinal']) > anio_termino:
            anio_termino = int(c['anoFinal'])


    data,nombres_programas = construir_data(mes_inicio,anio_inicio,mes_termino,anio_termino,programas,
                          planificacion_programas,datos_mensuales,sondajes_recomendaciones)
    #data = consumir_datos_api()

    # if not data:
    #     print("No hay datos disponibles en REPORTE AVANCE PROGRAMAS.")
    #     data = [
    #     {
    #         "mes": "sep-23",
    #         "GEOLOGÍA": {
    #             "Plan": 10.00,
    #             "Real mensual": 10.00,
    #             "Acumulado plan": 10.00,
    #             "Acumulado real": 10.00
    #         },
    #         "GEOTÉCNICO": {
    #             "Plan": 20.00,
    #             "Real mensual": 20.00,
    #             "Acumulado plan": 20.00,
    #             "Acumulado real": 20.00
    #         },
    #         "HIDROGEOLÓGICO - EVU": {
    #             "Plan": 0.00,
    #             "Real mensual": 0.00,
    #             "Acumulado plan": 0.00,
    #             "Acumulado real": 0.00
    #         }
    #     },
    #     {
    #         "mes": "oct-23",
    #         "GEOLOGÍA": {
    #             "Plan": 20.00,
    #             "Real mensual": 20.00,
    #             "Acumulado plan": 20.00,
    #             "Acumulado real": 20.00
    #         },
    #         "GEOTÉCNICO": {
    #             "Plan": 30.00,
    #             "Real mensual": 30.00,
    #             "Acumulado plan": 30.00,
    #             "Acumulado real": 30.00
    #         },
    #         "HIDROGEOLÓGICO - EVU": {
    #             "Plan": 0.00,
    #             "Real mensual": 0.00,
    #             "Acumulado plan": 0.00,
    #             "Acumulado real": 0.00
    #         }
    #     },
    #     {
    #         "mes": "nov-23",
    #         "GEOLOGÍA": {
    #             "Plan": 0.00,
    #             "Real mensual": 0.00,
    #             "Acumulado plan": 0.00,
    #             "Acumulado real": 0.00
    #         },
    #         "GEOTÉCNICO": {
    #             "Plan": 0.00,
    #             "Real mensual": 0.00,
    #             "Acumulado plan": 0.00,
    #             "Acumulado real": 0.00
    #         },
    #         "HIDROGEOLÓGICO - EVU": {
    #             "Plan": 0.00,
    #             "Real mensual": 0.00,
    #             "Acumulado plan": 0.00,
    #             "Acumulado real": 0.00
    #         }
    #     },
    #     {
    #         "mes": "dic-23",
    #         "GEOLOGÍA": {
    #             "Plan": 0.00,
    #             "Real mensual": 0.00,
    #             "Acumulado plan": 0.00,
    #             "Acumulado real": 0.00
    #         },
    #         "GEOTÉCNICO": {
    #             "Plan": 0.00,
    #             "Real mensual": 0.00,
    #             "Acumulado plan": 0.00,
    #             "Acumulado real": 0.00
    #         },
    #         "HIDROGEOLÓGICO - EVU": {
    #             "Plan": 0.00,
    #             "Real mensual": 0.00,
    #             "Acumulado plan": 0.00,
    #             "Acumulado real": 0.00
    #         }
    #     }
    # ]
    
    #     # Obtener todas las claves distintas de "mes" en la data 
    
    # Agregar una nueva hoja 
    hoja = libro.create_sheet(title="AVANCE PROGRAMA")

    hoja = helpers.obtener_fecha_documento(hoja, "A1","A1:F1")


    keys_unicas = set()
    for item in data:
        keys_unicas.update(item.keys())

    # Eliminar la clave "mes" si está presente y genera lsita de titulos
    keys_unicas.discard("mes")
    keys_unicas = sorted(list(keys_unicas))
    titulos = ["Programa"] + keys_unicas

    hoja, inicio_tabla, termino_tabla = helpers.agregar_titulos(hoja, "A", 3, titulos,"f2900e", alto_fila_factor=1, ancho_columna=8, tipo = "avance_programa")
    # Mover la nueva hoja a la primera posición

    data_ordenada = ordenar_data(data, keys_unicas)

    celdas_intermedias = obtener_celdas_intermedias(f'{inicio_tabla}:{termino_tabla}')
    

    filas_para_tabla = []
    graficos_actuales = {} 


    total_data = len(data_ordenada) -1 # se le resta 1 para poder calzar el incremento de filas

    for data in data_ordenada:
        fila_resultado,graficos_actuales = asignar_columnas(data, celdas_intermedias,graficos_actuales,total_data)
        filas_para_tabla.append(fila_resultado,)
    
    hoja, ultima_fila= agregar_datos_a_celdas(hoja,5, filas_para_tabla, "ffffff", "ffe598")

    ultima_fila = ultima_fila + 2
    # Iterar por las categorías
    for categoria, valores in graficos_actuales.items():
        titulo_plan = f'{categoria.upper()} PLAN vs REAL'
        titulo_acumulado = f'{categoria.upper()} ACUMULADO'
        # Iterar por cada subdiccionario dentro de la categoría
        for key, value in valores.items():

            if key == "plan":
                celda_inicio_grafico = f"A{ultima_fila}"
                titulo = titulo_plan
                columna_inicial = value["columna_inicial"]
            if key == "acumulado_plan":
                celda_inicio_grafico = f"L{ultima_fila}"
                titulo = titulo_acumulado
                columna_inicial = value["columna_inicial"]+2

            hoja = helpers.generar_tabla_avance_programa(
                hoja,
                titulo = titulo,
                celda_inicio_grafico =celda_inicio_grafico,
                columna_inicial = columna_inicial,
                inicio_filas = value["inicio_filas"],
                termino_filas = value["termino_filas"]
                )

            
        ultima_fila = ultima_fila + 17


    libro._sheets.insert(0, libro._sheets.pop(-1))

    return libro,nombres_programas