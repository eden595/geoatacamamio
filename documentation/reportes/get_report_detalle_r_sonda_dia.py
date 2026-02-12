from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime, timedelta
import locale
import pytz
import requests 
from pathlib import Path
from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn
from collections import defaultdict
from openpyxl.drawing.image import Image
from openpyxl.formatting.rule import CellIsRule
from pprint import pprint
from openpyxl import load_workbook
from documentation.reportes import helpers
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter, column_index_from_string
from collections import OrderedDict
import calendar

def titulo_tabla(hoja,start_merge,merge, color, text):
    
    hoja.merge_cells(merge)

    cell = hoja[start_merge]

    # Aplicar formato

    cell.alignment = Alignment(text_rotation=90, vertical='center', horizontal='center')
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.value = text

    return hoja

def agregar_datos_a_celdas(hoja, inicio_celda, datos, primer_color, segundo_color,mes,titulos_ordenados=None):

    fila_inicio = inicio_celda
    # Definir los colores alternados para las filas
    color_fondo_1 = PatternFill(start_color=primer_color, end_color=primer_color, fill_type="solid")
    color_fondo_2 = PatternFill(start_color=segundo_color, end_color=segundo_color, fill_type="solid")
    color_fondo_rojo = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
    color_fondo_verde = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid") 

    header_col_map = {header: get_column_letter(i + 2) for i, header in enumerate(titulos_ordenados)}

    numero_inicio_grafico = fila_inicio
    # columna_inicio_grafico = None
    # columna_dia_sonda = "B"
    # columna_mts_dia = None
    columna_mts_dia = header_col_map.get("Mts. Día")
    columna_dia_sonda = header_col_map.get("Día/Sonda")
    columna_rendimiento_dia = header_col_map.get("Rendimiento Día (m)")
    columna_inicio_grafico = None
    totales={}
    for idx,diccionario in enumerate(datos):

        # Determinar la fila donde se agregará este diccionario
        fila_actual = fila_inicio + idx
        ultima_fila = fila_actual  # Actualizar la última fila usada
        if idx % 2 == 0:
            color_fondo = color_fondo_1  # Verde para filas pares
        else:
            color_fondo = color_fondo_2  # Rojo para filas impares

        
        for j, key in enumerate(titulos_ordenados):  # Iteramos sobre los títulos ordenados
            columna = header_col_map.get(key, get_column_letter(j + 2))
            celda = hoja[f"{columna}{fila_actual}"]
            valor = diccionario.get(key, 0.00)  # Si la clave no existe, usa 0
            celda.value = valor
            celda.font = Font(name="Arial", size=8, bold=False)
            celda.alignment = Alignment(horizontal="center", vertical="center")
            if "-" in key and "(" not in key:

                if valor < 25:
                    celda.font = Font(name="Arial", size=8, bold=False, color="FF0000")  # Rojo
                    celda.fill = color_fondo_rojo
                else:
                    celda.font = Font(name="Arial", size=8, bold=False, color="00FF00")  # Verde
                    celda.fill = color_fondo_verde
            else:
                celda.fill = color_fondo
            if key != "Día/Sonda":
                celda.number_format = '#,##0.00'
            columna_inicio_grafico = get_column_letter(j + 4)
            

            if key != "Día/Sonda" and key != "Cantidad de Sondas" :

                if key not in totales:
                    totales[key]={
                        "columna":columna,
                        "total":float(valor)
                    }
                else:
                    totales[key]["total"] += float(valor)

    # Se agregan totales abajo de las tablas
    total_filas=len(datos)
    for key, total in totales.items():
        if "-" not in key:
            celda = hoja[f"{total['columna']}{fila_actual+1}"]
            celda.value = round(float(total['total']), 2)
            celda.font = Font(name="Arial", size=8, bold=True)
            celda.number_format = '#,##0.00'
            celda.alignment = Alignment(horizontal="center", vertical="center")
        
        if "-" in key and "(" not in key:
            celda = hoja[f"{total['columna']}{fila_actual+1}"]
            celda.value = round(float(total['total']), 2)
            celda.font = Font(name="Arial", size=8, bold=True)
            celda.number_format = '#,##0.00'
            celda.alignment = Alignment(horizontal="center", vertical="center")

            celda = hoja[f"{total['columna']}{fila_actual+2}"]
            celda.value = round(float(total['total']/total_filas), 2)
            celda.font = Font(name="Arial", size=8, bold=True)
            celda.number_format = '#,##0.00'
            celda.alignment = Alignment(horizontal="center", vertical="center")


    minimo = numero_inicio_grafico-1
    maximo = len(datos)+numero_inicio_grafico -1
    celda_inicio_grafico = f'{columna_inicio_grafico}{numero_inicio_grafico}'
    

    hoja = helpers.generar_grafico_lineal(
        hoja,
        f"Metros Diarios Perforados, {mes}",
        "DÌAS","METROS",
        celda_inicio_grafico,
        columna_dia_sonda,
        columna_mts_dia,
        minimo,
        maximo
    )

    indice_columna = column_index_from_string(columna_inicio_grafico) +10
    # Convertir índice de vuelta a letra
    columna_inicio_grafico = get_column_letter(indice_columna)
    celda_inicio_grafico = f'{columna_inicio_grafico}{numero_inicio_grafico}'

    hoja = helpers.generar_grafico_lineal(
        hoja,
        f"Rendimientos Día (m), {mes}",
        "DÌAS","METROS",
        celda_inicio_grafico,
        columna_dia_sonda,
        columna_rendimiento_dia,
        minimo,
        maximo
    )

    return hoja

def agregar_tabla_resumen(hoja,data, inicio_tabla_resumen):

    total_programas = 0

    estilo= Font(size=10, bold=True)  
    for categoria, valor in data.items():
        
        celda = hoja[f"B{inicio_tabla_resumen}"]
        celda.value = valor
        celda.font = estilo
        total_programas += valor

        hoja.merge_cells(f"C{inicio_tabla_resumen}:H{inicio_tabla_resumen}")
        celda = hoja[f"C{inicio_tabla_resumen}"]
        celda.value = f"ACUMULADO MTS. PERFORADOS ({categoria})"
        celda.font = estilo

        inicio_tabla_resumen+=1

    
    celda = hoja[f"B{inicio_tabla_resumen}"]
    celda.value = total_programas
    celda.font = estilo

    hoja.merge_cells(f"C{inicio_tabla_resumen}:H{inicio_tabla_resumen}")
    celda = hoja[f"C{inicio_tabla_resumen}"]
    celda.value = f"TOTAL AVANCE PROGRAMA 2021/2023"
    celda.font = estilo
    inicio_tabla_resumen+=1

    return hoja, inicio_tabla_resumen

def formatear_fecha_mes_anio(fecha_str):
    meses_es = {
        1: 'ENERO', 2: 'FEBRERO', 3: 'MARZO', 4: 'ABRIL',
        5: 'MAYO', 6: 'JUNIO', 7: 'JULIO', 8: 'AGOSTO',
        9: 'SEPTIEMBRE', 10: 'OCTUBRE', 11: 'NOVIEMBRE', 12: 'DICIEMBRE'
    }

    if not fecha_str:
        return "NO DISPONIBLE"

    
    try:
        partes = fecha_str.split('/')
        # formato esperado: DD/MM/YYYY
        dia = partes[0]
        mes = int(partes[1])
        anio = partes[2]
        return f"{meses_es.get(mes, 'MES DESCONOCIDO')} {anio}"
    except Exception:

        return "FORMATO INVÁLIDO"
    
def insertar_data_mensual(datos_mensuales,sondajes_recomendaciones):

          
    segunda_tabla= {}

    titulos_tablas_por_meses = []

    for sondaje , detalles in sondajes_recomendaciones.items():
        titulos_mes=[
                'Día/Sonda','Cantidad de Sondas','Mts. Día',
                'Mts. Proyectados a Perforar',
                'Rendimiento Día (m)',
                'Rendimiento Proyectado GEOTEC Día (m)',
                'Rendimiento por Hora (m)'
                ]
        
        for detalle in detalles:
            if detalle['recomendacion']["programa"] not in segunda_tabla:
                segunda_tabla[detalle['recomendacion']["programa"]] = 0.00

            fecha_dict = formatear_fecha_mes_anio(detalle['fechacreacion'])

            # Creacion de data para tabla mensual
            for dia in datos_mensuales[fecha_dict]:
                if dia['Día/Sonda'] == detalle['fechacreacion']:
                    avance_diario = float(detalle['metroFinal']) -float(detalle['metroInicial'])
                    sonda = detalle["nombre_sonda"] 
                    dia[f'Proyección GEOTEC ({sonda})'] = 25.00
                    dia[sonda] = avance_diario
                    dia['Cantidad de Sondas'] += 1
                    dia['Mts. Día'] += avance_diario
                    dia['Mts. Proyectados a Perforar'] = dia['Mts. Día']* 25
                    dia['Rendimiento Día (m)'] = dia['Mts. Día']/dia['Cantidad de Sondas']
                    dia['Rendimiento Proyectado GEOTEC Día (m)'] = dia['Mts. Proyectados a Perforar']/dia['Cantidad de Sondas']
                    dia['Rendimiento por Hora (m)'] = dia['Rendimiento Día (m)']/24

                    segunda_tabla[detalle['recomendacion']["programa"]] += avance_diario

                    if f'Proyección GEOTEC ({sonda})' not in titulos_mes:
                        titulos_mes.insert(1, f'Proyección GEOTEC ({sonda})')
                        titulos_mes.insert(2, sonda)



        titulos_tablas_por_meses.append({
            fecha_dict: titulos_mes
        })
            
    return datos_mensuales, segunda_tabla,titulos_tablas_por_meses

def separar_mes_anio(texto: str):
    partes = texto.strip().split()
    if len(partes) != 2:
        return None, None  # o lanzar una excepción si prefieres
    nombre_mes, anio = partes
    return nombre_mes, int(anio)

def ordenar_por_mes(diccionario):
    # Orden correcto de los meses
    orden_meses = [
        'ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO',
        'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE'
    ]

    nuevo_diccionario = {}
    for anio, datos in diccionario.items():
        ordenado = OrderedDict()
        for mes in orden_meses:
            if mes in datos:
                ordenado[mes] = datos[mes]
        if anio in datos:
            ordenado[anio] = datos[anio]
        nuevo_diccionario[anio] = ordenado

    return nuevo_diccionario

def convertir_todos_los_anios_a_tabla(datos_agrupados):

    tabla = [["Mes", "Metros Reales", "Metros Proyectados a la fecha"]]

    for anio in sorted(datos_agrupados.keys()):
        datos_mes = datos_agrupados[anio]
        
        for mes, valores in datos_mes.items():
            if isinstance(mes, int):
                mes = str(mes)
            fila = [mes.capitalize(), valores[1], valores[2]]
            tabla.append(fila)

    return tabla

def obtener_datos_tabla_meses_agrupados(datos_mensuales):

    total_por_anio ={}
    
    for dato, detalle in datos_mensuales.items():
        mes, anio = separar_mes_anio(dato)
        if anio not in total_por_anio:
            total_por_anio[anio] = {
                anio: [anio,0,0],
                }
        
        if mes not in total_por_anio[anio]:
            total_por_anio[anio][mes] = [mes,0,0]

        for d in detalle:
            total_por_anio[anio][mes][1] += d['Mts. Día']
            total_por_anio[anio][anio][1] += d['Mts. Día']

            total_por_anio[anio][mes][2] += d['Mts. Proyectados a Perforar']
            total_por_anio[anio][anio][2] += d['Mts. Proyectados a Perforar']

    # Se ordenan los meses por los años
    total_por_anio = ordenar_por_mes(total_por_anio)   

    # se genera tabla final 
    tabla = convertir_todos_los_anios_a_tabla(total_por_anio)

    return tabla

def obtener_total_dias(mes_anio: str) -> int:
    meses_es = {
        'ENERO': 1, 'FEBRERO': 2, 'MARZO': 3, 'ABRIL': 4,
        'MAYO': 5, 'JUNIO': 6, 'JULIO': 7, 'AGOSTO': 8,
        'SEPTIEMBRE': 9, 'OCTUBRE': 10, 'NOVIEMBRE': 11, 'DICIEMBRE': 12
    }

    mes_str, anio_str = mes_anio.upper().split()
    mes = meses_es[mes_str]
    anio = int(anio_str)

    _, total_dias = calendar.monthrange(anio, mes)
    return total_dias

def run(libro,sondajes_recomendaciones):

    #data = None
    datos_mensuales = helpers.generar_dict_rango_fechas(8, 2023, 9, 2035)

    datos_mensuales, resumen_programas,titulos_tablas_por_meses= insertar_data_mensual(datos_mensuales,sondajes_recomendaciones)

    tabla_meses_agrupados = obtener_datos_tabla_meses_agrupados(datos_mensuales)

    # if not data:
    #     print("No hay datos disponibles en REPORTE DETALLE R SONDA DIA.")
    #     # data = {
    #     #     "datos_mensuales": [
    #     #         {
    #     #             "AGOSTO 2023": [
    #     #                 {
    #     #                 "Día/Sonda": "8/1/2023",
    #     #                 "Proyección GEOTEC (CH1-99)": 20.00,
    #     #                 "CH1-99": 10.00,
    #     #                 "Proyección GEOTEC (CH1-129)": 20.00,
    #     #                 "CH1-129": 5.00,
    #     #                 "Proyección GEOTEC (CH1-130)": 20.00,
    #     #                 "CH1-130": 8.00,
    #     #                 "Cantidad de Sondas": 2,
    #     #                 "Mts. Día": 23.00,
    #     #                 "Mts. Proyectados a Perforar": 60.00,
    #     #                 "Rendimiento Día (m)": 11.50,
    #     #                 "Rendimiento Proyectado GEOTEC Día (m)": 20.00,
    #     #                 "Rendimiento por Hora (m)": 1.50
    #     #                 },
    #     #                 {
    #     #                 "Día/Sonda": "8/2/2023",
    #     #                 "Proyección GEOTEC (CH1-99)": 20.00,
    #     #                 "CH1-99": 12.00,
    #     #                 "Proyección GEOTEC (CH1-129)": 20.00,
    #     #                 "CH1-129": 7.00,
    #     #                 "Proyección GEOTEC (CH1-130)": 20.00,
    #     #                 "CH1-130": 9.00,
    #     #                 "Cantidad de Sondas": 2,
    #     #                 "Mts. Día": 28.00,
    #     #                 "Mts. Proyectados a Perforar": 60.00,
    #     #                 "Rendimiento Día (m)": 14.00,
    #     #                 "Rendimiento Proyectado GEOTEC Día (m)": 20.00,
    #     #                 "Rendimiento por Hora (m)": 1.75
    #     #                 },
    #     #                 {
    #     #                 "Día/Sonda": "8/3/2023",
    #     #                 "Proyección GEOTEC (CH1-99)": 20.00,
    #     #                 "CH1-99": 15.00,
    #     #                 "Proyección GEOTEC (CH1-129)": 20.00,
    #     #                 "CH1-129": 10.00,
    #     #                 "Proyección GEOTEC (CH1-130)": 20.00,
    #     #                 "CH1-130": 12.00,
    #     #                 "Cantidad de Sondas": 3,
    #     #                 "Mts. Día": 37.00,
    #     #                 "Mts. Proyectados a Perforar": 60.00,
    #     #                 "Rendimiento Día (m)": 12.33,
    #     #                 "Rendimiento Proyectado GEOTEC Día (m)": 20.00,
    #     #                 "Rendimiento por Hora (m)": 1.54
    #     #                 }
    #     #             ],
    #     #             "SEPTIEMBRE 2023": [
    #     #                 {
    #     #                 "Día/Sonda": "9/1/2023",
    #     #                 "Proyección GEOTEC (CH1-99)": 25.00,
    #     #                 "CH1-99": 15.00,
    #     #                 "Proyección GEOTEC (CH1-129)": 25.00,
    #     #                 "CH1-129": 36.00,
    #     #                 "Cantidad de Sondas": 2,
    #     #                 "Mts. Día": 402,
    #     #                 "Mts. Proyectados a Perforar": 60.00,
    #     #                 "Rendimiento Día (m)": 12.33,
    #     #                 "Rendimiento Proyectado GEOTEC Día (m)": 20.00,
    #     #                 "Rendimiento por Hora (m)": 1.54
    #     #                 },
    #     #                     {
    #     #                 "Día/Sonda": "9/1/2023",
    #     #                 "Proyección GEOTEC (CH1-99)": 25.00,
    #     #                 "CH1-99": 15.00,
    #     #                 "Proyección GEOTEC (CH1-129)": 25.00,
    #     #                 "CH1-129": 36.00,
    #     #                 "Cantidad de Sondas": 2,
    #     #                 "Mts. Día": 741,
    #     #                 "Mts. Proyectados a Perforar": 60.00,
    #     #                 "Rendimiento Día (m)": 12.33,
    #     #                 "Rendimiento Proyectado GEOTEC Día (m)": 20.00,
    #     #                 "Rendimiento por Hora (m)": 1.54
    #     #                     },
    #     #                 ],

    #     #             },
    #     #         ],
    #     #     "resumen_programas":{
    #     #             "Categorización": 10,
    #     #             "Hidrogeología": 10,
    #     #             "Hidrogeología AR": 10,
    #     #             "Geotécnico": 10,
    #     #             "Geometalúrgico": 10,
    #     #             "EVU": 10,
    #     #             "Condenación": 10,
    #     #         },
    #     #     "tabla_meses_agrupados":[
    #     #         ["Mes", "Metros Reales", "Metros Proyectados a la fecha"],  # Encabezados
    #     #         ["Enero", 100, 120],
    #     #         ["Febrero", 150, 180],
    #     #         ["Marzo", 200, 220],
    #     #         ["Abril", 250, 270],
    #     #         ["Total",1000,1000]
    #     #     ]}

    data = {
        "datos_mensuales":[datos_mensuales],
        "resumen_programas":resumen_programas,
        "tabla_meses_agrupados":tabla_meses_agrupados}

    # Agregar una nueva hoja 
    hoja = libro.create_sheet(title="DETALLE_R.SONDAXDÍA")

    hoja = helpers.obtener_fecha_documento(hoja, "A1","A1:C1","center")

    inicio_tabla_resumen = None
    
    # Proceso para agregar tablas por meses
    for entry in data["datos_mensuales"]:
        start_table = 4
        for mes, registros in entry.items():
            dias = obtener_total_dias(mes)
            hoja = titulo_tabla(hoja, f"A{start_table}", f"A{start_table}:A{start_table+dias}","FFFF00",mes)
            #Valor de inicio  tabla resumen
            inicio_tabla_resumen = start_table + 34 
            # Obtener las claves únicas del primer elemento de la Lista

            keys = set().union(*(d.keys() for d in registros))
            if mes in titulos_tablas_por_meses:
                keys = titulos_tablas_por_meses[mes]

            hoja,titulos_ordenados = helpers.agregar_titulos(hoja, "B", start_table, keys,"70ad47", 2,15,"detalle_sonda_dia")
            hoja = agregar_datos_a_celdas(hoja, start_table+1, registros,"e2efd9", "b4c6e7",mes,titulos_ordenados)
            start_table += 36 

    hoja,inicio_tabla_resumen = agregar_tabla_resumen(hoja,data["resumen_programas"],inicio_tabla_resumen )
    
    hoja = helpers.obtener_fecha_documento(
        hoja,
        f"B{inicio_tabla_resumen}",
        f"B{inicio_tabla_resumen}:H{inicio_tabla_resumen}",
        "center",
        "f7caac"
        )
    
    inicio_tabla_resumen+=4

    hoja_aux = libro.create_sheet(title="AUX_DETALLE_R.SONDAXDÍA")
    # Generar el gráfico de barras agrupadas desde la tabla
    hoja, hoja_aux = helpers.generar_grafico_barras_agrupadas_desde_tabla(
        hoja,
        "Metros reales vs Metros proyectados, agrupados por meses",
        f"A{inicio_tabla_resumen + 1}",
        data['tabla_meses_agrupados'],
        1,
        hoja_aux
        )

    # Asegurarse de que hoja sea la primera
    libro._sheets.insert(0, libro._sheets.pop(libro._sheets.index(hoja)))

    # Asegurarse de que hoja_aux sea la última
    libro._sheets.append(libro._sheets.pop(libro._sheets.index(hoja_aux)))

    return libro,datos_mensuales