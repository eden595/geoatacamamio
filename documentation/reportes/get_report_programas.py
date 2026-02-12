from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import locale
import pytz
import requests 
from pathlib import Path
from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn
from collections import defaultdict
from openpyxl.drawing.image import Image
from openpyxl.formatting.rule import CellIsRule
from pprint import pprint
from openpyxl import load_workbook
from collections import defaultdict
from ..reportes import helpers
import copy

def fusionar_celdas_con_formato(hoja, celda_inicio, celda_fin, valor, color_fondo="FFC000"):
    # Fusionar celdas
    rango = f"{celda_inicio}:{celda_fin}"
    hoja.merge_cells(rango)

    # Obtener la celda de inicio para aplicar formato
    celda = hoja[celda_inicio]
    celda.value = valor  # Asignar el valor

    # Aplicar formato
    celda.fill = PatternFill(start_color=color_fondo, end_color=color_fondo, fill_type="solid")
    celda.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")  # Texto Blancas
    celda.alignment = Alignment(horizontal="center", vertical="center")

    return hoja

def agregar_datos_a_celdas(hoja, inicio_celda, datos, primer_color, segundo_color):
    # Obtener la fila de inicio desde la celda (ej. "A1" -> 1)
    fila_inicio = int(inicio_celda[1:])

    # Definir los colores alternados para las filas
    color_fondo_1 = PatternFill(start_color=primer_color, end_color=primer_color, fill_type="solid")  # Verde
    color_fondo_2 = PatternFill(start_color=segundo_color, end_color=segundo_color, fill_type="solid")  # Rojo

    ultima_fila = fila_inicio  # Variable para almacenar la última fila usada

    # Mapeo de claves del JSON a columnas de Excel

    mapeo_columnas = {
        "ID": "B",
        "Tipo de Sondaje": "C",
        "Sondaje": "D",
        "Sector": "E",
        "Este": "F",
        "Norte": "G",
        "Cota": "H",
        "Azimut": "I",
        "Inclinación": "J",
        "Largo (m)": "K",
        "Fecha Inicio": "L",
        "Fecha Termino": "M",
        "Por Perforar (m)": "N",
        "Avance Actual (m)": "O",
        "Mts. Faltantes": "P",
        "%Avance": "Q",
        "Estatus Perforación (m)": "R",
        "Largo Final (m)": "S",
        "Certificado Collar": "T",
        "Fecha Medición de Trayectoria": "U",
        "Observación": "V"
    }
    # Iterar sobre los datos
    for idx, diccionario in enumerate(datos):
        fila_actual = fila_inicio + idx
        ultima_fila = fila_actual  # Actualizar la última fila usada
        color_fondo = color_fondo_1 if idx % 2 == 0 else color_fondo_2  # Alternar colores

        # Iterar sobre cada clave y asignar los valores en la hoja
        for key, value in diccionario.items():
            if key in mapeo_columnas:
                columna = mapeo_columnas[key]
                celda = hoja[f"{columna}{fila_actual}"]
                celda.value = value

                # Aplicar formato a la celda
                celda.font = Font(name="Arial", size=8, bold=False)
                celda.alignment = Alignment(horizontal="center", vertical="center")
                celda.fill = color_fondo  # Aplicar color de fondo


    return hoja, ultima_fila

def obtener_collar(id_recomendacion,all_recomendaciones_final):

    for rec in all_recomendaciones_final:

        if rec['recomendacionFinal'] == id_recomendacion:

            return "SI"

    return "NO"

def construir_data(reportes_agrupados,all_recomendaciones_final):

    opciones_estado = {
            '1':'Abortado',
            '2':'En Avance',
            '3':'En Espera',
            '4':'Finalizado',
            }
    resultado = defaultdict(list)

    for reporte, detalles in reportes_agrupados.items():
        ##################################
        # 1. calcular el máximo id
        max_id = max(d["id"] for d in detalles)

        # 2. obtener directamente el dict con ese id
        ultimo_registro = next(d for d in detalles if d["id"] == max_id)

        if 'detalle_perforaciones' in ultimo_registro:
            programa = ultimo_registro['recomendacion']['programa']

            if programa not in resultado:
                    resultado[programa]=[]

            # obtengo el ultimo de la lista
            ultimo_detalle_perforacion = ultimo_registro['detalle_perforaciones'][-1]
            
            desde = float(ultimo_detalle_perforacion['DESDE'])
            hasta = float(ultimo_detalle_perforacion['HASTA'])
            total_dia = desde + hasta
            largo_programado = float(ultimo_registro['recomendacion']['largo_programado'])

            id_estado = ultimo_registro['recomendacion'].get('estado')

            
            if id_estado:
                estado = opciones_estado[ultimo_registro['recomendacion']['estado']]
                if estado == "Finalizado":
                    fecha_termino = ultimo_registro['recomendacion']['fechaupdateestado']
                    
                else:
                    fecha_termino = " "
                
            else:
                estado = "SIN ESTADO"

            collar = obtener_collar(ultimo_registro['recomendacion']['id'],all_recomendaciones_final)

            por_perforar = largo_programado - hasta
            
            if por_perforar != float(0.00):
                avance_actual = (hasta / por_perforar) * 100
            else:
                avance_actual = float(0.00)


            resultado[programa].append({
                "ID": ultimo_registro['recomendacion']['programa'],
                "Tipo de Sondaje": "SIN DATO",
                "Sondaje": ultimo_registro['recomendacion']['pozo'],
                "Sector": ultimo_registro['recomendacion']['sector'],
                "Este": ultimo_registro['recomendacion']['este'],
                "Norte": ultimo_registro['recomendacion']['norte'],
                "Cota": ultimo_registro['recomendacion']['cota'],
                "Azimut": ultimo_registro['recomendacion']['azimut'],
                "Inclinación": ultimo_registro['recomendacion']['inclinacion'],
                "Largo (m)": largo_programado,
                "Fecha Inicio": ultimo_registro['recomendacion']['fecha_inicio'],
                "Fecha Termino": fecha_termino,
                "Por Perforar (m)": largo_programado - hasta,
                "Avance Actual (m)": hasta,
                "Mts. Faltantes": largo_programado - hasta,
                "%Avance": round(avance_actual,2),
                "Estatus Perforación (m)": estado,
                "Largo Final (m)": largo_programado,
                "Certificado Collar": collar,
                "Fecha Medición de Trayectoria": "SIN DATO",
                "Observación": "SIN DATO",
            })


        ##################################
        # for detalle in detalles:

        #     if 'detalle_perforaciones' not in detalle:
        #         continue

        #     programa = detalle['recomendacion']['programa']

        #     if programa not in resultado:
        #             resultado[programa]=[]

        #     for perforacion in detalle['detalle_perforaciones']:


        #         desde = float(perforacion['DESDE'])
        #         hasta = float(perforacion['HASTA'])
        #         total_dia = desde + hasta
        #         largo_programado = float(detalle['recomendacion']['largo_programado'])

        #         id_estado = detalle['recomendacion'].get('estado')

                
        #         if id_estado:
        #             estado = opciones_estado[detalle['recomendacion']['estado']]
        #             if estado == "Finalizado":
        #                 fecha_termino = detalle['recomendacion']['fechaupdateestado']
                        
        #             else:
        #                 fecha_termino = " "
                    
        #         else:
        #             estado = "SIN ESTADO"

        #         collar = obtener_collar(detalle['recomendacion']['id'],all_recomendaciones_final)

        #         por_perforar = largo_programado - hasta
                
        #         if por_perforar != float(0.00):
        #             avance_actual = (hasta / por_perforar) * 100
        #         else:
        #             avance_actual = float(0.00)


        #         resultado[programa].append({
        #             "ID": detalle['recomendacion']['programa'],
        #             "Tipo de Sondaje": "SIN DATO",
        #             "Sondaje": detalle['recomendacion']['pozo'],
        #             "Sector": detalle['recomendacion']['sector'],
        #             "Este": detalle['recomendacion']['este'],
        #             "Norte": detalle['recomendacion']['norte'],
        #             "Cota": detalle['recomendacion']['cota'],
        #             "Azimut": detalle['recomendacion']['azimut'],
        #             "Inclinación": detalle['recomendacion']['inclinacion'],
        #             "Largo (m)": largo_programado,
        #             "Fecha Inicio": detalle['recomendacion']['fecha_inicio'],
        #             "Fecha Termino": fecha_termino,
        #             "Por Perforar (m)": largo_programado - hasta,
        #             "Avance Actual (m)": hasta,
        #             "Mts. Faltantes": largo_programado - hasta,
        #             "%Avance": avance_actual,
        #             "Estatus Perforación (m)": estado,
        #             "Largo Final (m)": largo_programado,
        #             "Certificado Collar": collar,
        #             "Fecha Medición de Trayectoria": "SIN DATO",
        #             "Observación": "SIN DATO",
        #         })


    return resultado

def detalle_campanas(campanas,campana_actual):
    anio_inicial ="_SIN DATO"
    anio_final ="_SIN DATO"

    for c in campanas:
        if c['id'] == campana_actual:
            anio_inicial = f"_{c['anoInicial']}"
            anio_final = f"_{c['anoFinal']}"
            return anio_inicial,anio_final

    return anio_inicial,anio_final
def detalle_programas(programas, programa_actual,campanas):
    anio_inicial ="_SIN DATO"
    anio_final ="_SIN DATO"
    for p in programas:


        if p['programa'].upper() == programa_actual.upper():
            anio_inicial, anio_final = detalle_campanas(campanas,p['campana'])
            return anio_inicial, anio_final

    return anio_inicial,anio_final
def run(libro,reportes_agrupados,programas,campanas,all_recomendaciones_final):


    data = construir_data(reportes_agrupados,all_recomendaciones_final)

    lista_titulos_esperados = []

    anio_inicial_avance = 0
    anio_final_avance = 0
    for p in programas:
        anio_inicial,anio_final = detalle_programas(programas, p["programa"], campanas)

        if anio_inicial_avance == 0:
            anio_inicial_avance = anio_inicial
            anio_final_avance = anio_final

        if anio_inicial < anio_inicial_avance :  
            anio_inicial_avance = anio_inicial
  
        if anio_final > anio_final_avance:  
            anio_final_avance = anio_final  

        lista_titulos_esperados.append(f"{p['programa'].upper()}{anio_inicial}{anio_final}")
 

    titulos = ["ID","Tipo de Sondaje","Sondaje","Sector","Este","Norte","Cota","Azimut","Inclinación","Largo (m)","Fecha Inicio","Fecha Termino","Por Perforar (m)","Avance Actual (m)","Mts. Faltantes","%Avance","Estatus Perforación (m)","Largo Final (m)","Certificado Collar","Fecha Medición de Trayectoria","Observación"]
    
    for programa, data in data.items():
        anio_inicial,anio_final = detalle_programas(programas, programa, campanas)

        titulo = f"{programa.upper()}{anio_inicial}{anio_final}"
        if titulo in lista_titulos_esperados:
            lista_titulos_esperados.remove(titulo)

        hoja = libro.create_sheet(title=titulo)

        hoja = helpers.obtener_fecha_documento(hoja, "B1","B1:E1")

        hoja = fusionar_celdas_con_formato(hoja,"F3","J3","COORDENADAS", "70ad47")

        hoja = helpers.agregar_titulos(hoja, "B", 4, titulos, "70ad47",3,12)

        hoja, ultima_fila = agregar_datos_a_celdas(hoja, "B5", data, "e2efd9", "b4c6e7")
        # Ajustar el ancho de las columnas
        hoja.column_dimensions["A"].width = 1
        hoja.column_dimensions["B"].width = 13
        hoja.column_dimensions["C"].width = 8
        hoja.column_dimensions["D"].width = 8
        hoja.column_dimensions["E"].width = 8
        hoja.column_dimensions["F"].width = 8
        hoja.column_dimensions["G"].width = 8
        hoja.column_dimensions["H"].width = 8
        hoja.column_dimensions["I"].width = 8
        hoja.column_dimensions["J"].width = 8
        hoja.column_dimensions["K"].width = 8
        hoja.column_dimensions["L"].width = 8
        hoja.column_dimensions["M"].width = 8
        hoja.column_dimensions["N"].width = 8
        hoja.column_dimensions["O"].width = 8
        hoja.column_dimensions["P"].width = 8
        hoja.column_dimensions["Q"].width = 8
        hoja.column_dimensions["R"].width = 8
        hoja.column_dimensions["S"].width = 8
        hoja.column_dimensions["T"].width = 8
        hoja.column_dimensions["U"].width = 8
        hoja.column_dimensions["V"].width = 20


        # Mover la nueva hoja a la primera posición
        libro._sheets.insert(0, libro._sheets.pop(-1))

    for titulo_esperado in lista_titulos_esperados:


        hoja = libro.create_sheet(title=titulo_esperado)

        hoja = helpers.obtener_fecha_documento(hoja, "B1","B1:E1")

        hoja = fusionar_celdas_con_formato(hoja,"F3","J3","COORDENADAS", "70ad47")

        hoja = helpers.agregar_titulos(hoja, "B", 4, titulos, "70ad47",3,12)
        # Ajustar el ancho de las columnas
        hoja.column_dimensions["A"].width = 1
        hoja.column_dimensions["B"].width = 13
        hoja.column_dimensions["C"].width = 8
        hoja.column_dimensions["D"].width = 8
        hoja.column_dimensions["E"].width = 8
        hoja.column_dimensions["F"].width = 8
        hoja.column_dimensions["G"].width = 8
        hoja.column_dimensions["H"].width = 8
        hoja.column_dimensions["I"].width = 8
        hoja.column_dimensions["J"].width = 8
        hoja.column_dimensions["K"].width = 8
        hoja.column_dimensions["L"].width = 8
        hoja.column_dimensions["M"].width = 8
        hoja.column_dimensions["N"].width = 8
        hoja.column_dimensions["O"].width = 8
        hoja.column_dimensions["P"].width = 8
        hoja.column_dimensions["Q"].width = 8
        hoja.column_dimensions["R"].width = 8
        hoja.column_dimensions["S"].width = 8
        hoja.column_dimensions["T"].width = 8
        hoja.column_dimensions["U"].width = 8
        hoja.column_dimensions["V"].width = 20


        # Mover la nueva hoja a la primera posición
        libro._sheets.insert(0, libro._sheets.pop(-1))
    return libro, anio_inicial_avance, anio_final_avance