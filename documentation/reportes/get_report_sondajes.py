from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import locale
import pytz
import requests
from pathlib import Path
from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn
from collections import defaultdict
from openpyxl.drawing.image import Image
from openpyxl.formatting.rule import CellIsRule
from pathlib import Path
from openpyxl.drawing.image import Image

# ================================
# FUNCIONES PRINCIPALES PARA CREACIÓN DEL EXCEL
# ================================

def obtener_detalle_general(hoja):
    """
    Formatea la hoja con títulos, bordes, y colores.
    """

    # Aplicar bordes punteados a celdas específicas (E5:E9, H5:H8, L5:L7)
    borde_punteado = Border(left=Side(style="hair"), right=Side(style="hair"), top=Side(style="hair"), bottom=Side(style="hair"))
    fuente_negrita = Font(name="Arial", size=8, bold=True)
    wrap_text = Alignment(wrap_text=True, horizontal="center", vertical="center")

    # Aplicar color de fuente negra y fondo amarillo a celdas específicas (W, AE, AL)
    columnas_negras = ["W", "AE", "AL"]
    fondo_amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    estilo_moneda = NamedStyle(name="moneda")
    estilo_moneda.number_format = "$0"

    # Ajuste de ancho de columnas
    hoja.column_dimensions["A"].width = 2.0  # Columna A (1)
    hoja.column_dimensions["D"].width = 10.5 # Columna D (4)
    hoja.column_dimensions["G"].width = 17.5 # Columna G (7)
    hoja.column_dimensions["E"].width = 17.5 # Columna E (5)
    hoja.column_dimensions["AN"].width = 20.44  # Columna AN (40)
    hoja.column_dimensions["AO"].width = 56.51  # Columna AO (41)

    # Contenido y formato en D5:D9
    contenido_celdas = ["Rec", "Pozo", "Sonda", "Fecha Inicio", "Sector"]
    for fila, texto in enumerate(contenido_celdas, start=5):
        celda = hoja[f"D{fila}"]
        celda.value = texto
        celda.font = fuente_negrita
        hoja[f"E{fila}"].border = borde_punteado

    # Contenido y formato en G5:G8
    contenido_celdas = ["Azimut", "Inclinación", "Largo Programado", "Largo real"]
    for fila, texto in enumerate(contenido_celdas, start=5):
        celda = hoja[f"G{fila}"]
        celda.value = texto
        celda.font = fuente_negrita
        hoja[f"I{fila}"].border = borde_punteado

    # Contenido y formato en K5:K7
    contenido_celdas = ["Este", "Norte", "Cota"]
    for fila, texto in enumerate(contenido_celdas, start=5):
        celda = hoja[f"K{fila}"]
        celda.value = texto
        celda.font = fuente_negrita
        hoja[f"L{fila}"].border = borde_punteado

    # Agregar título en D3:L3
    hoja.merge_cells("D3:L3")
    celda = hoja["D3"]
    celda.value = "CAMPAÑA SONDAJE 2021/2025"
    celda.font = Font(bold=True, size=8)
    celda.alignment = Alignment(horizontal="center", vertical="center")

    # Aplicar bordes delgados a celdas combinadas D3:L3
    borde = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    for fila in range(3, 4):
        for columna in range(4, 13):
            hoja.cell(row=fila, column=columna).border = borde

    # Sección CLIENTE en L9:AO9
    hoja.merge_cells("L9:AO9")
    celda = hoja["L9"]
    celda.value = "CLIENTE"
    celda.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    celda.font = Font(color="FFFFFF", size=8, bold=True)
    celda.alignment = Alignment(horizontal="center", vertical="center")

    # Aplicar bordes punteados a L9:AO9
    for fila in range(9, 10):
        for columna in range(12, 42):  # Columnas L (12) a AO (41)
            hoja.cell(row=fila, column=columna).border = borde_punteado
    
    # Aplicar fondo a Columnas AM, AN, AO en la Fila 11
    for columna in range(39, 42):  # AM (39) hasta AO (41)
        hoja.cell(row=11, column=columna).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

    # Sección GEOTEC en AP9:BZ9
    hoja.merge_cells("AP9:BZ9")
    celda = hoja["AP9"]
    celda.value = "GEOTEC"
    celda.fill = PatternFill(start_color="339933", end_color="339933", fill_type="solid")
    celda.font = Font(color="FFFFFF", size=8, bold=True)
    celda.alignment = Alignment(horizontal="center", vertical="center")

    # Aplicar bordes punteados a AP9:BZ9
    for fila in range(9, 10):
        for columna in range(42, 79):  # Columnas AP (42) a BZ (78)
            hoja.cell(row=fila, column=columna).border = borde_punteado

    # Sección PRECIO CON EQUIPO OPERANDO en M11:V11
    hoja.merge_cells("M11:V11")
    celda = hoja["M11"]
    celda.value = "I. PRECIO CON EQUIPO OPERANDO"
    celda.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    celda.font = Font(color="FFFFFF", size=8, bold=True)
    celda.alignment = Alignment(horizontal="center", vertical="center")
    
    # Sección PRECIO CON EQUIPO STAND BY en X11:AD11
    hoja.merge_cells("X11:AD11")
    celda = hoja["X11"]
    celda.value = "II. PRECIO CON EQUIPO STAND BY"
    celda.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    celda.font = Font(color="FFFFFF", size=8, bold=True)
    celda.alignment = Alignment(horizontal="center", vertical="center")

    # Sección PRECIO EQUIPO DE APOYO en AF11:AK11
    hoja.merge_cells("AF11:AK11")
    celda = hoja["AF11"]
    celda.value = "III. PRECIO EQUIPO DE APOYO"
    celda.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    celda.font = Font(color="FFFFFF", size=8, bold=True)
    celda.alignment = Alignment(horizontal="center", vertical="center")

    # Sección OPERANDO en AW10:BA11
    hoja.merge_cells("AW10:BA11")
    celda = hoja["AW10"]
    celda.value = "OPERANDO"
    celda.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    celda.font = Font(color="FFFFFF", size=8, bold=True)
    celda.alignment = Alignment(horizontal="center", vertical="center")

    # Sección FALLAS/REPARACIONES en BM10:BV11
    hoja.merge_cells("BM10:BV11")
    celda = hoja["BM10"]
    celda.value = "FALLAS / REPARACIONES"
    celda.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    celda.font = Font(color="FFFFFF", size=8, bold=True)
    celda.alignment = Alignment(horizontal="center", vertical="center")

    # Ajustar ancho de columnas
    for col in range(13, 39):  # Columnas M (13) a AM (39)
        hoja.column_dimensions[get_column_letter(col)].width = 13.55

    for col in range(43, 47):  # Columnas AQ (43) a AU (47)
        hoja.column_dimensions[get_column_letter(col)].width = 15.45

    for col in range(51, 53):  # Columnas AY (51) a AZ (52)
        hoja.column_dimensions[get_column_letter(col)].width = 15.45

    for col in range(54, 78):  # Columnas BB (54) a BZ (78)
        hoja.column_dimensions[get_column_letter(col)].width = 15.45
        

    # Aplicar color de fuente negra a las celdas específicas (W, AE, AL)
    for col in columnas_negras:
        for fila in [10, 11, 12]:
            celda = hoja[f"{col}{fila}"]
            celda.font = Font(color="000000", size=8, bold=True)
            celda.alignment = wrap_text
        
        # Aplicar fondo amarillo y contenido a filas 10 y 11
        hoja[f"{col}10"].fill = fondo_amarillo
        hoja[f"{col}10"].value = "Costo"
        hoja[f"{col}11"].fill = fondo_amarillo
        hoja[f"{col}11"].value = 0  # Asignar como número
        hoja[f"{col}11"].number_format = estilo_moneda.number_format  # Aplicar formato de moneda

    return hoja

def filtrar_reportes_aprobados(reportes, sondajes):
    """Filtra reportes con progreso 'Aprobado'."""
    return [
        reporte for reporte in reportes
        if reporte.get("progreso") == "Aprobado" and
        f"{sondajes.get(reporte['sondajeCodigo'], 'Unknown')}-{reporte['sondajeSerie']}".strip()
    ]

def agrupar_reportes_por_sondaje(reportes_aprobados, sondajes, estados):
    """Agrupa reportes aprobados por código de sondaje."""
    reportes_agrupados = defaultdict(list)

    for reporte in reportes_aprobados:
        # Verificar si reporte es una lista y tiene elementos
        if isinstance(reporte, list) and reporte:
            reporte_data = reporte[0]  # Extrae el primer elemento
        elif isinstance(reporte, dict):  # Si es un diccionario, úsalo directamente
            reporte_data = reporte
        else:
            print(f"Advertencia: formato inesperado en reporte: {reporte}")
            continue  # Ignorar este reporte si no tiene formato válido
    
        codigo_sondaje = f"{sondajes.get(reporte_data.get('sondajeCodigo', ''), 'Unknown')}-{reporte_data.get('sondajeSerie', '')}{estados.get(reporte_data.get('sondajeEstado', ''), '')}".strip()
        reportes_agrupados[codigo_sondaje].append(reporte)

    return reportes_agrupados

def insertar_logo_empresa(hoja):
    """Inserta el logo de la empresa en la hoja de Excel."""

    # Ruta al logo desde el archivo actual (por ejemplo, desde documentation/)
    script_dir = Path(__file__).resolve().parent  # 'documentation/'
    logo_path = script_dir / "logo_informe" / "logo.jpg"

    if not logo_path.exists():
        print(f"No se encontró el logo en: {logo_path}")
        return hoja

    # Cargar imagen
    logo = Image(str(logo_path))

    # Ajustar tamaño si es necesario
    logo.width = 115  # Ancho en píxeles
    logo.height = 100  # Alto en píxeles

    # Posicionar en la celda B4 (se expandirá hasta C9 visualmente)
    hoja.add_image(logo, "B4")

    return hoja

def generar_celdas_blancas(hoja):
    """
    Aplica un fondo blanco a celdas de la hoja.
    """
    blanco = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Aplicar fondo blanco a Columnas A, B y C (Fila 3 - 11)
    for fila in range(3, 12):
        for columna in range(1, 4):  # A (1), B (2), C (3)
            hoja.cell(row=fila, column=columna).fill = blanco

    # Aplicar fondo blanco a Columnas D a K (Fila 3 - 11)
    for fila in range(3, 12):
        for columna in range(4, 12):  # D (4) hasta K (12)
            hoja.cell(row=fila, column=columna).fill = blanco

    # Aplicar fondo blanco a Columnas M a V (Fila 3 - 8)
    for fila in range(3, 9):
        for columna in range(13, 23):  # M (13) hasta V (22)
            hoja.cell(row=fila, column=columna).fill = blanco

    # Aplicar fondo blanco a Columnas L a V en la Fila 10 únicamente
    for columna in range(12, 23):  # L (12) hasta V (22)
        hoja.cell(row=10, column=columna).fill = blanco

    # Aplicar fondo blanco a Columnas AM, AN, AO en la Fila 10
    for columna in range(39, 42):  # AM (39) hasta AO (41)
        hoja.cell(row=10, column=columna).fill = blanco

    # Aplicar fondo blanco a Columnas AP a AV desde fila 10 hasta fila 11
    for fila in range(10, 12):
        for columna in range(42, 49):  # AP (42) hasta AV (48)
            hoja.cell(row=fila, column=columna).fill = blanco

    # Aplicar fondo blanco a Columnas BB a BL desde fila 10 hasta fila 11
    for fila in range(10, 12):
        for columna in range(54, 65):  # BB (54) hasta BL (65)
            hoja.cell(row=fila, column=columna).fill = blanco

    # Aplicar fondo blanco a Columnas BW a BZ desde fila 10 hasta fila 11
    for fila in range(10, 12):
        for columna in range(75, 86):  # BW (75) hasta BZ (79)
            hoja.cell(row=fila, column=columna).fill = blanco

    for columna in range(79, 82):  #CA (79) hasta CC (81) fila 9
        hoja.cell(row=9, column=columna).fill = blanco

    # Aplicar fondo blanco a Columnas AN a CC desde fila 2 hasta fila 8
    for fila in range(3, 9):
        for columna in range(40, 82):  # AN (40) hasta CC (81)
            hoja.cell(row=fila, column=columna).fill = blanco

    return hoja

def obtener_fecha_documento(hoja):
    """
    Combina las celdas A1:E1 de la hoja y agrega la fecha actual
    formateada en español con la hora chilena.
    """
    # Establecer el idioma a español
    locale.setlocale(locale.LC_TIME, "es_ES.UTF-8")

    # Obtener la hora actual en la zona horaria de Chile
    zona_chile = pytz.timezone("America/Santiago")
    fecha_actual = datetime.now(zona_chile)

    # Formatear la fecha actual
    fecha_formateada = fecha_actual.strftime("%A, %B %d, %Y").lower()

    # Combinar las celdas de A1 a E1
    hoja.merge_cells("A1:E1")
    
    # Establecer el valor en la celda combinada
    celda = hoja["A1"]
    celda.value = fecha_formateada

    # Aplicar formato: texto en negritas y tamaño 8
    celda.font = Font(bold=True, size=8)
    celda.alignment = Alignment(horizontal="center", vertical="center")

    return hoja

def inicializar_hoja(libro, codigo_sondaje, primera_hoja):
    """Crea una hoja de Excel y aplica formato inicial."""
    hoja = libro.active if primera_hoja else libro.create_sheet(title=codigo_sondaje)
    hoja.title = codigo_sondaje
    hoja = obtener_fecha_documento(hoja)
    hoja = obtener_detalle_general(hoja)
    return hoja

def aplicar_estilo_encabezados(hoja, titulos, colores_por_rango):
    """Aplica estilos a los encabezados en la fila 12."""
    for col_idx, titulo in enumerate(titulos, start=2):
        celda = hoja[f"{get_column_letter(col_idx)}12"]
        celda.value = titulo
        celda.font = Font(color="FFFFFF", size=8, bold=True)
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda.border = Border(top=Side(style="hair"))

        # Aplicar colores
        color_celda = next((color for rango, color in colores_por_rango.items() if col_idx in rango), "FFFFFF")
        celda.fill = PatternFill(start_color=color_celda, end_color=color_celda, fill_type="solid")

def configurar_columnas(hoja):
    """Agrupa y da formato a las columnas de la hoja de Excel."""
    # Agrupar columnas de "M" hasta "W" (Grupo 1)
    for col in range(13, 23):
        hoja.column_dimensions[get_column_letter(col)].outline_level = 1  

    # Agrupar columnas de "X" hasta "AD" (Grupo 2)
    for col in range(24, 31):
        hoja.column_dimensions[get_column_letter(col)].outline_level = 1  

    # Agrupar columnas de "AF" hasta "AK" (Grupo 3)
    for col in range(32, 38):
        hoja.column_dimensions[get_column_letter(col)].outline_level = 1 
        hoja.column_dimensions[get_column_letter(col)].hidden = True  # Mantener contraído por defecto 

    # Agrupar columnas de "AP" hasta "BX" (Grupo 4 - Gran Bloque)
    for col in range(42, 77):
        hoja.column_dimensions[get_column_letter(col)].outline_level = 1 

def aplicar_estilo_columnas_especificas(hoja, columnas_negras):
    """Aplica color de fuente negra a columnas específicas."""
    for col in columnas_negras:
        for fila in [10, 11, 12]:
            celda = hoja[f"{col}{fila}"]
            celda.font = Font(color="000000", size=8, bold=True)
            celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def aplicar_estilo_datos(hoja, cells_to_fill):
    """Aplica estilos a las celdas de datos."""
    for cell, value in cells_to_fill.items():
        hoja[cell] = value
        if isinstance(value, float):
            hoja[cell].number_format = '#,##0.00'
        hoja[cell].font = Font(size=8)
        hoja[cell].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def forzar_formato_numerico(hoja, fila_inicio_tabla, fila_fin_tabla, columnas_con_totales):
    """
    Aplica formato de número con dos decimales a todas las celdas dentro de la tabla,
    excepto en las columnas "W", "AE" y "AL".
    """
    columnas_excluir = {"W", "AE", "AL"}  # Conjunto de columnas a excluir

    for col_letter in columnas_con_totales:
        if col_letter not in columnas_excluir:  # Verificar si la columna no está en las excluidas
            for fila in range(fila_inicio_tabla, fila_fin_tabla + 1):
                celda = hoja[f"{col_letter}{fila}"]
                celda.number_format = '#,##0.00'  # Formato numérico con dos decimales

def aplicar_totales(hoja, columnas_con_totales, fila_inicio_tabla, fila_totales, titulos):
    """Aplica la función SUBTOTAL en las columnas numéricas y formatea los totales."""
    col_fin_tabla = 2 + len(titulos) - 1  # Ajusta el rango de columnas según los títulos
    columnas_monedas = ["W", "AE", "AL"]  # Columnas donde el total debe mostrarse como "$ 0"

    for col_idx in range(7, col_fin_tabla + 1):  # Ajusta dinámicamente el rango de columnas
        col_letter = get_column_letter(col_idx)
        celda_total = hoja[f"{col_letter}{fila_totales}"]
        
        if col_letter in columnas_con_totales or col_letter in columnas_monedas:
            celda_total.value = f"=SUBTOTAL(109, {col_letter}{fila_inicio_tabla}:{col_letter}{fila_totales - 1})"

            if col_letter in columnas_monedas:
                celda_total.number_format = '"$"#,##0'  # Formato de moneda "$ 0"
            else:
                celda_total.number_format = '#,##0.00'  # Formato decimal normal

        celda_total.font = Font(bold=True, color="000000", size=8)
        celda_total.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def aplicar_formato_celdas(hoja, fila_inicio_tabla, fila_fin_tabla, col_fin_tabla):
    """Aplica formato a todas las celdas desde la fila 13 en adelante."""
    for fila in hoja.iter_rows(min_row=fila_inicio_tabla, max_row=fila_fin_tabla, min_col=2, max_col=col_fin_tabla):
        for celda in fila:
            celda.font = Font(size=8)  # Asegurar tamaño de fuente 8
            celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def agregar_tabla(hoja, titulos, fila_actual):
    """Agrega una tabla a la hoja de Excel."""
    col_fin_tabla = 2 + len(titulos) - 1
    rango_tabla = f"B12:{get_column_letter(col_fin_tabla)}{fila_actual}"

    tabla = Table(displayName=f"Tabla_{hoja.title.replace('-', '_')}", ref=rango_tabla)
    tabla.tableColumns = [TableColumn(id=i+1, name=titulos[i]) for i in range(len(titulos))]

    estilo_tabla = TableStyleInfo(
        name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tabla.tableStyleInfo = estilo_tabla
    tabla.headerRowCount = 1

    hoja.add_table(tabla)


def calcular_totales_horas(hoja, fila_totales):
    """
    Calcula los totales de distintas categorías de horas y los coloca en sus respectivas columnas.
    También aplica estilos y el color de fondo "A9D08E".
    """
    categorias = [
        {"col_inicio": 13, "col_fin": 22, "col_resultado": "W", "texto": "Horas Operando"},      # M - W
        {"col_inicio": 24, "col_fin": 30, "col_resultado": "AE", "texto": "Horas Standby"},      # X - AE
        {"col_inicio": 32, "col_fin": 37, "col_resultado": "AL", "texto": "Horas Equipo de Apoyo"}  # AF - AL
    ]

    total_fill = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")  # Verde claro

    for categoria in categorias:
        col_suma_inicio = categoria["col_inicio"]
        col_suma_fin = categoria["col_fin"]
        col_resultado = categoria["col_resultado"]
        fila_total_resultado = fila_totales + 1  # Fila donde se mostrará la suma total
        fila_texto = fila_totales + 2  # Fila donde se colocará el título

        # Fórmula de suma para la categoría correspondiente
        formula_suma = f"=SUM({get_column_letter(col_suma_inicio)}{fila_totales}:{get_column_letter(col_suma_fin)}{fila_totales})"
        
        # Aplicar el total en la columna correspondiente
        celda_total = hoja[f"{col_resultado}{fila_total_resultado}"]
        celda_total.value = formula_suma
        celda_total.number_format = '#,##0.00'  # Aplicar formato decimal
        celda_total.font = Font(size=8, bold=True)  # Aplicar mismo tamaño de fuente
        celda_total.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda_total.fill = total_fill  # Aplicar color de fondo

        # Agregar el texto correspondiente debajo del total
        celda_texto = hoja[f"{col_resultado}{fila_texto}"]
        celda_texto.value = categoria["texto"]
        celda_texto.font = Font(size=8)  # Aplicar mismo tamaño de fuente
        celda_texto.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda_texto.fill = total_fill  # Aplicar color de fondo

def insertar_formula_totales(hoja, fila_inicio, fila_fin):
    """
    Inserta las siguientes fórmulas:
    - `=SI.ERROR(SUMA(M13:V13) * AE$11, 0)` en la columna W.
    - `=SI.ERROR(SUMA(X13:AD13) * AE$11, 0)` en la columna AE.
    - `=SI.ERROR(SUMA(AF13:AK13), 0)` en la columna AL.
    - `=SI.ERROR(SUMA(M13:V13;X13:AD13), 0)` en la columna AM.
    - `=SI.ERROR(SUMA(AP13:BX13), 0)` en la columna BY.
    - `=SI.ERROR(BY13+AM13, 0)` en la columna BZ.
    
    Además, en la columna BZ, cambia el color a "c6efce" solo si el valor es exactamente "12,00".
    """

    categorias = [
        {"col_resultado": "W", "col_inicio": "M", "col_fin": "V"},
        {"col_resultado": "AE", "col_inicio": "X", "col_fin": "AD"},
        {"col_resultado": "AL", "col_inicio": "AF", "col_fin": "AK"},
        {"col_resultado": "AM", "col_inicio": "M", "col_fin": "V"},
        {"col_resultado": "BY", "col_inicio": "AP", "col_fin": "BX"},
        {"col_resultado": "BZ", "formula": "BY{fila}+AM{fila}"}  # Nueva fórmula para BZ
    ]

    # Detectar si Excel usa "SI.ERROR" (español) o "IFERROR" (inglés)
    try:
        hoja["Z1"] = "=SI.ERROR(1/0, 0)"  # Prueba en español
        if hoja["Z1"].value is None:
            formula_error = "SI.ERROR"  # Excel en español
            formula_sum = "SUMA"  # Función SUMA en español
            separator = ";"  # Separador de argumentos en español
        else:
            formula_error = "IFERROR"  # Excel en inglés
            formula_sum = "SUM"  # Función SUM en inglés
            separator = ","  # Separador de argumentos en inglés
        hoja["Z1"] = None  # Limpiar la celda de prueba
    except:
        formula_error = "IFERROR"  # Por defecto, usar inglés si hay un error
        formula_sum = "SUM"
        separator = ","  # Separador de argumentos en inglés

    for categoria in categorias:
        col_resultado = categoria["col_resultado"]
        col_inicio = categoria.get("col_inicio")
        col_fin = categoria.get("col_fin")
        formula_template = categoria.get("formula")

        for fila in range(fila_inicio, fila_fin):
            celda = hoja[f"{col_resultado}{fila}"]

            # Seleccionar la fórmula adecuada
            if col_resultado in ["W", "AE"]:
                formula = f"={formula_error}({formula_sum}({col_inicio}{fila}:{col_fin}{fila})*AE$11, 0)"
            elif col_resultado == "AM":
                formula = f"={formula_error}({formula_sum}({col_inicio}{fila}:{col_fin}{fila}{separator}X{fila}:AD{fila}), 0)"
            elif col_resultado == "BZ":
                formula = f"={formula_error}({formula_template.format(fila=fila)}, 0)"  # BZ: BY + AM
            else:
                formula = f"={formula_error}({formula_sum}({col_inicio}{fila}:{col_fin}{fila}), 0)"  # AL, BY con "$ 0"

            celda.value = formula
            celda.number_format = '"$ "#,##0'  # Formato de moneda: "$ 0"
            celda.font = Font(size=8)
            celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Aplicar formato condicional para cambiar color en BZ solo si el valor es 12,00
    color_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    for fila in range(fila_inicio, fila_fin):
        regla = CellIsRule(operator="equal", formula=["12"], stopIfTrue=True, fill=color_verde)
        hoja.conditional_formatting.add(f"BZ{fila}", regla)

def insertar_formula_columna_L(hoja, fila_inicio, fila_fin):
    """
    Inserta la fórmula `=SI.ERROR(G13-J13, 0)` en la columna L,
    asegurando que el formato de los valores sea numérico con dos decimales "0,00".
    """

    # Detectar si Excel usa "SI.ERROR" (español) o "IFERROR" (inglés)
    try:
        hoja["Z1"] = "=SI.ERROR(1/0, 0)"  # Prueba en español
        if hoja["Z1"].value is None:
            formula_error = "SI.ERROR"  # Excel en español
        else:
            formula_error = "IFERROR"  # Excel en inglés
        hoja["Z1"] = None  # Limpiar la celda de prueba
    except:
        formula_error = "IFERROR"  # Por defecto, usar inglés si hay un error

    formula_template = "G{fila}-J{fila}"

    for fila in range(fila_inicio, fila_fin):
        celda = hoja[f"L{fila}"]

        # Aplicar la fórmula con control de errores
        formula = f"={formula_error}({formula_template.format(fila=fila)}, 0)"
        celda.value = formula

        # Formato numérico con dos decimales "0,00"
        celda.number_format = "0.00"

        # Aplicar estilos a la celda
        celda.font = Font(size=8)
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def asignar_valor_total_a_I8(hoja, fila_fin_tabla):
    """Obtiene el valor total calculado de la última fila en la columna 'K' y lo asigna a la celda 'I8' con formato adecuado, sin fórmula."""

    # Buscar la última fila con datos en la columna "K"
    ultima_fila = hoja.max_row  # Última fila usada en la hoja

    while ultima_fila > 1 and hoja[f"K{ultima_fila}"].value in [None, ""]:  # Evitar celdas vacías o sin datos
        ultima_fila -= 1

    # Obtener la celda en la última fila de "K"
    celda_k = hoja[f"K{ultima_fila}"]
    
    # Obtener solo el valor calculado, ignorando la fórmula
    if celda_k.data_type == "f":  # Si la celda tiene una fórmula
        hoja[f"I8"] = celda_k.value  # Asignar el resultado de la celda
    else:
        valor_total = celda_k.value  # Solo copia el valor, no la fórmula
        if isinstance(valor_total, (int, float)):  # Verificar si es un número
            hoja["I8"] = valor_total  # Asigna el valor numérico en I8

            # Aplicar formato numérico con 2 decimales en "I8"
            forzar_formato_numerico(hoja, 8, 8, ["I"])  # Aplica formato solo en "I8"

            # Aplicar estilos adicionales
            cell = hoja["I8"]
            cell.font = Font(name="Arial", size=8, bold=True)  # Fuente Arial, tamaño 8, negrita
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # Centrado y wrap text

def obtener_diametro_mas_reciente(reporte_id, reportes_detalles, diametros_data):
    """Obtiene el diámetro más reciente asociado a un reporte."""
    detalles_reporte = [
        detalle for detalle in reportes_detalles if detalle["reporte"] == reporte_id
    ]

    if not detalles_reporte:
        return "Desconocido"

    # Ordenar por fecha y seleccionar el más reciente
    detalle_mas_reciente = max(
        detalles_reporte, key=lambda x: datetime.fromisoformat(x["fechacreacion"])
    )

    # Obtener el nombre del diámetro
    diametro_id = detalle_mas_reciente["diametros"]
    return diametros_data.get(diametro_id, "Desconocido")

def obtener_recomendacion_por_sondaje(codigo_sondaje, recomendaciones):
    """Busca la recomendación asociada a un código de sondaje (pozo)."""
    for recomendacion in recomendaciones:
        if recomendacion["pozo"] == codigo_sondaje:
            return recomendacion
    return None

def asignar_recomendaciones_hoja(hoja, codigo_sondaje, recomendaciones, sondas):
    """Inserta los datos de la recomendación en las celdas indicadas, manteniendo su cantidad original de decimales."""
    recomendacion = obtener_recomendacion_por_sondaje(codigo_sondaje, recomendaciones)

    if not recomendacion:
        print(f"No se encontró recomendación para el sondaje: {codigo_sondaje}")
        return

    # Función para detectar la cantidad de decimales en un número
    def detectar_decimales(valor):
        try:
            valor_str = str(valor).rstrip("0")  # Elimina ceros innecesarios al final, pero no antes del punto
            if "." in valor_str:
                decimales = len(valor_str.split(".")[1])
                return max(decimales, 2)  # Mínimo 2 decimales, pero conserva más si los hay
            return 2  # Si no tiene punto decimal, se muestra con 2 decimales
        except (ValueError, TypeError):
            return 2

    # Función para convertir a número sin alterar su cantidad original de decimales
    def convertir_a_decimal(valor):
        try:
            return float(valor)  # Se guarda como número real
        except (ValueError, TypeError):
            return None  # Excel lo interpreta como celda vacía

    # Mapeo de celdas con los datos de la recomendación
    datos_celdas = {
        "E5": recomendacion["recomendacion"],
        "E6": recomendacion["pozo"],
        "E7": sondas.get(recomendacion["sonda"], "Desconocido"),
        "E8": recomendacion["fecha_inicio"].split("T")[0],  # Extrae solo la fecha
        "E9": recomendacion["sector"],  # NO tendrá wrap_text
        "I5": convertir_a_decimal(recomendacion["azimut"]),  
        "I6": convertir_a_decimal(recomendacion["inclinacion"]),  
        "I7": convertir_a_decimal(recomendacion["largo_programado"]),  
        "I8": convertir_a_decimal(recomendacion["largo_real"]),  
        "L5": convertir_a_decimal(recomendacion["este"]),  
        "L6": convertir_a_decimal(recomendacion["norte"]),  
        "L7": convertir_a_decimal(recomendacion["cota"]),  
    }

    # Detectar la cantidad de decimales para cada número y aplicar formato adecuado
    formatos_decimales = {
        celda: f"0.{''.join(['0' for _ in range(detectar_decimales(valor))])}" if isinstance(valor, float) else None
        for celda, valor in datos_celdas.items()
    }

    # Insertar los datos en la hoja
    for celda, valor in datos_celdas.items():
        hoja[celda] = valor

    # Aplicar estilos después de insertar los datos
    aplicar_estilos_recomendaciones(hoja, datos_celdas, formatos_decimales)

def aplicar_estilos_recomendaciones(hoja, datos_celdas, formatos_decimales):
    """Aplica los estilos adecuados a las celdas de recomendaciones, asegurando el formato correcto de números sin alterar la cantidad de decimales."""
    for celda, valor in datos_celdas.items():
        cell = hoja[celda]

        # Aplicar formato general
        cell.font = Font(name="Arial", size=8, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Aplicar wrap_text a TODA la columna "E" excepto "E9"
        if celda.startswith("E") and celda != "E9":
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Aplicar formato numérico solo si es número y tiene decimales detectados
        if isinstance(valor, float) and formatos_decimales.get(celda):
            cell.number_format = formatos_decimales[celda]  # Aplica el formato con la cantidad exacta de decimales


def encontrar_observacion(observaciones, id_reporte):

    for obs in observaciones:
        
        if obs['reporte'] == id_reporte:
            if obs['observaciones'] == "":
                return "Sin observación"
            return obs['observaciones']

    return "Sin observación"
def procesar_registro_reporte(reporte, sondas, reportes_controles_horarios, detalle_column_map, fila_actual, codigo_sondaje, reportes_detalles, diametros_data, recomendaciones,observaciones):
    """Procesa un registro del reporte y devuelve un diccionario con los valores a insertar en la hoja."""
    nombre_sonda = sondas.get(reporte.get("sonda"), "Unknown")
    turno = "Día" if reporte.get("turno") == "1" else "Noche"

    fecha_hora = reporte.get("fechacreacion", "N/A")
    fecha_ajustada = "N/A"
    if fecha_hora != "N/A":
        dt = datetime.fromisoformat(fecha_hora)
        if turno == "Noche" and 0 <= dt.hour < 8:
            dt -= timedelta(days=1)
        fecha_ajustada = dt.strftime("%Y-%m-%d")


    observacion = encontrar_observacion(observaciones, reporte['id'])

    # if "reportesMaterialesCaseta" in reporte and reporte["reportesMaterialesCaseta"]:
    #     observacion = reporte["reportesMaterialesCaseta"][0].get("observacion", "Sin observación")

    # Obtener el diámetro más reciente para el reporte
    diametro_nombre = obtener_diametro_mas_reciente(reporte["id"], reportes_detalles, diametros_data)

    # Obtener la recomendación correspondiente al `codigo_sondaje`
    recomendacion = obtener_recomendacion_por_sondaje(codigo_sondaje, recomendaciones)
    recomendacion_texto = recomendacion["recomendacion"] if recomendacion else "Sin recomendación"

    # Obtener el "largo_programado" de la recomendación y asegurarse de que es float
    largo_programado = float(recomendacion["largo_programado"]) if recomendacion and "largo_programado" in recomendacion else 0.00

    # Procesar los valores de `reportes_controles_horarios` y convertirlos a formato decimal
    control_horario_data = {}
    for control_horario in reportes_controles_horarios:
        if control_horario["reporte"] == reporte["id"]:
            col = detalle_column_map.get(control_horario["detalleControlHorario"])
            if col:
                tiempo = control_horario["total"]
                horas, minutos, _ = map(int, tiempo.split(":"))
                valor_decimal = horas + minutos / 60
                control_horario_data[col] = round(valor_decimal, 2)

    return {
        f"B{fila_actual}": fecha_ajustada,
        f"C{fila_actual}": turno,
        f"D{fila_actual}": nombre_sonda,
        f"E{fila_actual}": recomendacion_texto,
        f"F{fila_actual}": codigo_sondaje,
        f"G{fila_actual}": largo_programado,
        f"H{fila_actual}": diametro_nombre,
        f"I{fila_actual}": float(reporte.get("metroInicial", 0.00)),
        f"J{fila_actual}": float(reporte.get("metroFinal", 0.00)),
        f"K{fila_actual}": float(reporte.get("totalPerforado", 0.00)),
        f"AN{fila_actual}": reporte.get("creador", "Desconocido"),
        f"AO{fila_actual}": observacion,
        **{f"{col}{fila_actual}": value for col, value in control_horario_data.items()}
    }


def generar_excel(libro, primera_hoja,reportes_agrupados,sondas,reportes_controles_horarios,reportes_detalles,diametros_data,recomendaciones,observaciones):

    libro = Workbook()
    primera_hoja = True
    titulos = [
            "Fecha",
            "Turno",
            "Sonda",
            "Rec",
            "Sondaje",
            "Largo Programado",
            "Diametro",
            "Desde (m)",
            "Hasta (m)",
            "Total Día (m)",
            "Faltante",
            "1. Instalación Casing",
            "2. Retiro Casing",
            "3. Cementación y Perforación",
            "4. Acondicionamiento de pozo",
            "5. Operación requerida por el cliente",
            "5. Rescate barras/herramientas (sólo condiciones de riesgo)",
            "6. Ensanche de pozo (reaming)",
            "7. Movimiento herramienta (excepto operación normal)",
            "8. Pruebas geotécnicas con equipo operando",
            "9. Instalación de instrumentación en los pozos con equipo operando",
            "Precio con equipo operando ($)",
            "1. Esperando instrucciones del cliente",
            "2. Medición de trayectorias",
            "3. Traslado entre plataformas y/o pozos",
            "4. Demoras por caminos o plataformas no habilitadas",
            "5. Detenciones solicitadas por MLP",
            "6. Pruebas geotécnicas con equipo detenido",
            "7. Instalación de instrumentación en los pozos con equipo detenido",
            "Precio con equipo Stand By ($)",
            "1. Esperando instrucciones del cliente2",
            "2. Traslado de equipo entre sectores",
            "3. Limpieza de camino",
            "4. Construcción de camino",
            "5. Construcción de plataforma",
            "6. Construcción de pretil",
            "Precio equipo de apoyo ($)",
            "Cliente (Hrs.)",
            "Controlador (Geoatacama)",
            "Observación",
            "01 - Esperando Insumos / repuestos",
            "02 - Esperando Agua",
            "03 - Espera instrucción Geotec",
            "04 - Esperando personal de apoyo",
            "05 - Detención por clima",
            "06 - Esperando traslado personal",
            "07 - Inspección / Charla / Reunion HSE",
            "08 - Movimiento de Herramienta",
            "09 - Reperforación - reaming2",
            "10 - Perforando",
            "11 - Rescate herramienta atrapada - cortada",
            "12 - Perforando matriz-fierro en el pozo",
            "13 - Rescatando muestra (fuera del pozo)",
            "14 - Cambio de diámetro o guías (HQ-PQ-HWT)",
            "15 - Preparando hta / tubo interior",
            "16 - Desarme - traslado - instalación de equipo",
            "17 - Limpieza y orden de plataforma",
            "18 - Preparación de Lodo",
            "19 - Retiro de lodo / Lavado de tarros",
            "20 - Desaguando / Armando circuito de lodo",
            "21 - Descongelando",
            "22 - Colación",
            "23 - Tiempo de traslado colación",
            "24 -Problema en equipo de apoyo (bus, retroexcavadora)",
            "25 - Falla alternador, motor de arranque o batería",
            "26 - Falla eléctrica",
            "27 - Defecto en bombas",
            "28 - Falla en rod sloop / rod handler",
            "29 - Falla en winches o poleas",
            "30 - Defecto en mangueras (fuga/rotura/cambio)",
            "31 - Falla en la mesa o el chuck",
            "32 - Falla mecánica del motor",
            "33 - Otra falla o reparación",
            "34 - Mantención inicial (cheq de niv, engrase)",
            "35 - Mantención programada",
            "Geotec (Hrs.)",
            "12/8 Horas",
        ]
    colores_por_rango = {
            range(2, 11): "70AD47", # VERDE
            range(11, 12): "0D0D0D", # NEGRO
            range(12, 23): "70AD47", # VERDE
            range(23, 24): "FFFF00", # AMARILLO
            range(24, 31): "70AD47", # VERDE
            range(31, 32): "FFFF00", # AMARILLO
            range(32, 38): "70AD47", # VERDE
            range(38, 39): "FFFF00", # AMARILLO
            range(39, 40): "0D0D0D", # NEGRO
            range(40, 49): "70AD47", # VERDE
            range(49, 54): "92D050", # VERDE CLARO
            range(54, 65): "70AD47", # VERDE
            range(65, 75): "00B050", # VERDE OSCURO
            range(75, 77): "70AD47", # VERDE
            range(77, 78): "0D0D0D", # NEGRO
            range(78, 79): "70AD47", # VERDE
            
        }
    columnas_negras = ["W", "AE", "AL"]
    columnas_con_totales = ["K", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", 
                            "AA", "AB", "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AP", "AQ", "AR", "AS", "AT", "AU", "AV", "AW", "AX", "AY", "AZ", 
                            "BA", "BB", "BC", "BD", "BE", "BF", "BG", "BH", "BI", "BJ", "BK", "BL", "BM", "BN", "BO", "BP", "BQ", "BR", "BS", "BT", "BU", "BV", "BW", "BX", "BY", "BZ"]
    detalle_column_map = {
            1: "S", 2: "AC", 3: "BH", 4: "AD", 5: "BP", 6: "O", 7: "Q",
            8: "BX", 9: "BS", 10: "AV", 11: "AW", 12: "R", 13: "V", 14: "M",
            15: "BV", 16: "Z", 18: "T", 19: "AS", 20: "BT", 21: "X", 22: "BR",
            23: "BE", 24: "AX", 25: "N", 26: "Y", 27: "AP", 28: "AB", 29: "BU",
            30: "BN", 31: "AT", 32: "BC", 33: "BF", 34: "AZ", 35: "U", 36: "BW",
            37: "BA", 38: "BG", 39: "BM", 40: "BQ", 41: "BK", 42: "AY", 43: "BB",
            44: "BI", 45: "AU", 46: "P", 47: "BL", 48: "BO", 49: "BD", 50: "AQ",
            51: "AR", 52: "BJ", 53: "AA"
        }


    for codigo_sondaje, registros in reportes_agrupados.items():

        hoja = inicializar_hoja(libro, codigo_sondaje, primera_hoja)

        generar_celdas_blancas(hoja)
        
        configurar_columnas(hoja)

        aplicar_estilo_encabezados(hoja, titulos, colores_por_rango)
        
        aplicar_estilo_columnas_especificas(hoja, columnas_negras)

        # Insertar el logo de la empresa
        insertar_logo_empresa(hoja)

        fila_actual = 13

        asignar_recomendaciones_hoja(hoja, codigo_sondaje, recomendaciones, sondas)

        if not registros:
            hoja[f"B{fila_actual}"] = "Sin datos"
            fila_actual += 1
            continue

        for reporte in registros:
            cells_to_fill = procesar_registro_reporte(
                reporte, sondas, reportes_controles_horarios, detalle_column_map, fila_actual, 
                codigo_sondaje, reportes_detalles, diametros_data, recomendaciones,observaciones
            )

            aplicar_estilo_datos(hoja, cells_to_fill)

            fila_actual += 1

        # Última fila de datos
        fila_fin_tabla = fila_actual 

        # Aplicar formato a todas las celdas desde la fila 13 en adelante
        aplicar_formato_celdas(hoja, 13, fila_fin_tabla, 2 + len(titulos) - 1)

        # Aplicar totales
        aplicar_totales(hoja, columnas_con_totales, 13, fila_actual, titulos)

        # Asignar el valor total a la celda I8
        asignar_valor_total_a_I8(hoja, fila_fin_tabla)

        # Insertar fórmula específica en la columna L
        insertar_formula_columna_L(hoja, 13, fila_actual)

        # Calcular totales de Horas Operando, Standby y Equipo de Apoyo
        calcular_totales_horas(hoja, fila_actual)

        # Insertar fórmulas en las columnas W, AE y AL
        insertar_formula_totales(hoja, 13, fila_fin_tabla)

        # Forzar el formato numérico en las columnas de totales para evitar la conversión automática a enteros
        forzar_formato_numerico(hoja, 13, fila_fin_tabla, columnas_con_totales)

        # Agregar la tabla
        agregar_tabla(hoja, titulos, fila_actual)

        # Aplicar estilo a la tabla
        primera_hoja = False

        

    return libro


