from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from core.models import TipoDocumentoFaenaGeneral, Faena
from user.models import UsuarioProfile
from .forms import FormSelectFaena, FormSelectTipo
from vehicle.tasks import generate_vehicle_pdfs_and_send_email
from django.http import HttpResponse, JsonResponse
import os
from pathlib import Path
import requests
import openpyxl
from django.conf import settings
from datetime import datetime, timedelta
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A0
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import locale
from django.http import FileResponse
from core.decorators import admin_required
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image
import json
from openpyxl import Workbook
from core.decorators import sondaje_admin_or_base_datos_or_supervisor_required, admin_required
import subprocess
import sys
from .reportes import get_report_sondajes as reportes_sondajes
from .reportes import get_report_perf_vs_rec as reportes_perf_vs_rec
from .reportes import get_report_recomendacion as reportes_recomendacion
from .reportes import get_report_avance_programa as reportes_avance_programa
from .reportes import get_report_m_trayectoria as reportes_m_trayectoria
from .reportes import get_report_avance_muestrera as reportes_avance_muestrera
from .reportes import get_report_rendimiento_mensual as reportes_rendimiento_mensual
from .reportes import get_report_programas as reportes_programas
from .reportes import get_report_avance_diario as reportes_avance_diario
from .reportes import get_report_gerenciales as reportes_gerenciales
from .reportes import get_graficos_avance_diario as graficos_avance_diario
from .reportes import get_report_detalle_r_sonda_dia as reportes_detalle_r_sonda_dia
from .reportes import main as main_reportes
from drilling.models import DocumentosReportesPerforaciones
from django.core.files import File
import subprocess
import sys
from pathlib import Path

@login_required
def view_general_mining_documents(request): 
    storage = messages.get_messages(request)
    storage.used = True
    tipo_documentos = list(TipoDocumentoFaenaGeneral.objects.all().order_by('fechacreacion'))
    faenausuario = UsuarioProfile.objects.get(user=request.user)
    context = {
        'faenausuario': faenausuario.faena,
        'tipo_documentos': tipo_documentos,
        'sidebar': 'manage_viewdocuemntationmining',
        'sidebarmain': 'manage_documentation', 
    }
    return render(request,'pages/documentation/view_general_mining_documents.html', context)

@login_required
def select_massive_vehicles(request):

    context = {
        'formfaenas': FormSelectFaena,
        'formtipos': FormSelectTipo,
        'sidebar': 'manage_massive_vehicles',
        'sidebarmain': 'manage_documentation', 
    }
    return render(request,'pages/documentation/select_massive_vehicles.html', context)

def request_massive_vehicle_pdf(request):
    if request.method == 'POST':
        generate_vehicle_pdfs_and_send_email.delay()
        return HttpResponse("done")
    else:
        return redirect('select_massive_vehicles')
    
def consumir_api(url_path):
    # 1. Tomamos la base (http://127.0.0.1:8000 o https://sicgeoatacama.cl)
    base_url = settings.BASE_API_URL
    
    if "sicgeoatacama.cl" in url_path:
        url_final = url_path.replace("https://sicgeoatacama.cl", base_url)
    elif url_path.startswith("http"):
        url_final = url_path 
    else:
        url_final = f"{base_url.rstrip('/')}/{url_path.lstrip('/')}"

    try:
        if settings.DEBUG:
            print(f"🛡️ LLAMADA SEGURA A: {url_final}")
            
        response = requests.get(url_final, timeout=10)
        response.raise_for_status()
        return response.json()
    except Exception as e:
        raise ValueError(f"Error en comunicación segura: {e}")
    

def cargar_plantilla(path):
    """
    Carga una plantilla de Excel.
    """
    if not Path(path).exists():
        raise FileNotFoundError("La plantilla especificada no existe.")
    return openpyxl.load_workbook(path)

def obtener_datos_vehiculos():

    try:
        
        faenas = consumir_api("/api/vehiculos_faenas/")
        vehiculos = consumir_api("/api/vehiculos/")
        kilometrajes = consumir_api("/api/vehiculos_kilometrajes/")

        # Filtrar datos con status=True
        faenas_filtradas = [item for item in faenas if item.get("status") is True]
        vehiculos_filtrados = {item["placaPatente"]: item for item in vehiculos if item.get("status") is True}

        # Obtener el kilometraje más alto por vehículo
        kilometrajes_filtrados = {}
        for item in kilometrajes:
            vehiculo = item["vehiculo"]
            if vehiculo not in kilometrajes_filtrados or item["kilometraje"] > kilometrajes_filtrados[vehiculo]:
                kilometrajes_filtrados[vehiculo] = item["kilometraje"]

        # Combinar datos
        datos_combinados = []
        for faena in faenas_filtradas:
            vehiculo_id = faena["vehiculo"]
            vehiculo_info = vehiculos_filtrados.get(vehiculo_id)
            if not vehiculo_info:
                continue  # Saltar si no hay información del vehículo

            datos_combinados.append({
                "vehiculo": vehiculo_id,
                "faena": faena["faena"],
                "tipo": vehiculo_info["tipo"],
                "marca": vehiculo_info["marca"],
                "modelo": vehiculo_info["modelo"],
                "color": vehiculo_info.get("color", "N/A"),
                "ano": vehiculo_info["ano"],
                "situacion": "Activo" if vehiculo_info["status"] else "Inactivo",
                "tenencia": vehiculo_info["tenencia"],
                "kilometraje": kilometrajes_filtrados.get(vehiculo_id, 0),
                "tieneTag": vehiculo_info.get("tieneTag", False),
                "nombrePropietario": vehiculo_info.get("nombrePropietario", "N/A"),
                "rutPropietario": vehiculo_info.get("rutPropietario", "N/A"),
                "fechaArriendoInicial": vehiculo_info.get("fechaArriendoInicial", "N/A"),
                "fechaArriendoFinal": vehiculo_info.get("fechaArriendoFinal", "N/A"),
                "numeroVin": vehiculo_info.get("numeroVin", "N/A"),
                "numeroMotor": vehiculo_info.get("numeroMotor", "N/A"),
                "numeroChasis": vehiculo_info.get("numeroChasis", "N/A"),
            })

        return datos_combinados
    except ValueError as e:
        raise ValueError(f"Error al combinar datos: {e}")

def generar_excel_dinamico(datos, carpeta_salida, campos_seleccionados):
    """
    Genera un archivo Excel dinámico con bordes en la fecha y el título,
    mantiene la tabla correctamente formateada y reserva espacio para el logo en B3:B6.
    """
    try:
        if not datos:
            return "No hay datos para procesar."
            
        # **Ruta del logo**
        ruta_logo = Path(settings.MEDIA_ROOT) / "plantilla_base_informe_vehiculo" / "logobg.jpeg"

        # **Diccionario de mapeo de nombres de columnas**
        mapeo_titulos = {
            "faena": "FAENA",
            "tipo": "TIPO VEHICULO",
            "vehiculo": "PATENTE",
            "marca": "MARCA",
            "modelo": "MODELO",
            "color": "COLOR",
            "ano": "AÑO",
            "situacion": "SITUACION",
            "tenencia": "TENENCIA",
            "kilometraje": "KILOMETRAJE",
            "tieneTag": "TAG",
            "nombrePropietario": "NOMBRE PROPIETARIO",
            "rutPropietario": "RUT PROPIETARIO",
            "fechaArriendoInicial": "INICIO FECHA ARRIENDO",
            "fechaArriendoFinal": "TERMINO FECHA ARRIENDO",
            "numeroVin": "N° VIN",
            "numeroMotor": "N° MOTOR",
            "numeroChasis": "N° CHASIS",
        }

        # **Función para parsear fechas al formato "DD-MM-YYYY"**
        def parsear_fecha(fecha):
            if fecha and "T" in fecha:
                try:
                    return datetime.fromisoformat(fecha.split("T")[0]).strftime("%d-%m-%Y")
                except ValueError:
                    return "N/A"
            return "N/A"

        # **Configurar bordes en "todos los bordes"**
        thin_border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )

        # Crear un nuevo libro de Excel y seleccionar la hoja activa
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "INFORME VEHICULAR"

        # **Configurar la columna "A" con ancho fijo de 2.5**
        sheet.column_dimensions["A"].width = 2.5

        # **Configurar la fecha dinámica en A1 (combinada en A1:D1)**
        sheet.merge_cells("A1:D1")
        cell_fecha = sheet["A1"]
        cell_fecha.value = '=TEXT(NOW(),"dddd, d"" de ""mmmm"" de ""yyyy, HH:mm")'
        cell_fecha.font = Font(bold=True, size=8)
        cell_fecha.alignment = Alignment(horizontal="center", vertical="center")

        # **Aplicar bordes a A1:D1**
        for col in range(1, 5):  # Columnas A (1) a D (4)
            sheet.cell(row=1, column=col).border = thin_border

        # **Configurar el título en la fila 3 (combinando E3:K3)**
        sheet.merge_cells("E3:K3")
        cell_titulo = sheet["E3"]
        cell_titulo.value = "SICGEOATACAMA - INFORME VEHICULAR POR FAENA"
        cell_titulo.font = Font(bold=True, size=14)
        cell_titulo.alignment = Alignment(horizontal="center", vertical="center")

        # **Aplicar bordes a E3:K3**
        for col in range(5, 12):  # Columnas E (5) a K (11)
            sheet.cell(row=3, column=col).border = thin_border

        # **Insertar el logo si el archivo existe**
        if ruta_logo.exists():
            img = Image(str(ruta_logo))
            img.width = 100  # Ajustar el ancho del logo
            img.height = 75  # Ajustar la altura del logo
            sheet.add_image(img, "B3")  # Insertar la imagen en la celda B3
            

        # **Estilos**
        titulo_font = Font(bold=True, color="FFFFFF")
        titulo_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")  # Verde #70AD47
        fila_impar_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Verde Claro 80%
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # **Configurar el alto de la fila 12 (títulos) en 36px**
        sheet.row_dimensions[12].height = 36

        # **Insertar encabezados en la fila 12, comenzando desde la columna B**
        col_start = 2  # Comenzar en la columna B (índice 2)
        col_end = col_start + len(campos_seleccionados) - 1  # Última columna según los campos seleccionados
        for idx, campo in enumerate(campos_seleccionados, start=col_start):
            col_letter = get_column_letter(idx)
            cell = sheet[f"{col_letter}12"]
            cell.value = mapeo_titulos.get(campo, campo.upper())
            cell.font = titulo_font
            cell.fill = titulo_fill
            cell.alignment = center_align
            sheet.column_dimensions[col_letter].width = len(cell.value) + 4

        # **Llenar los datos dinámicamente desde la fila 13**
        row_start = 13
        row_end = row_start + len(datos) - 1  # Última fila según los datos
        for row_idx, dato in enumerate(datos, start=row_start):
            sheet.row_dimensions[row_idx].height = 15  # Ajustar altura de fila automáticamente
            for col_idx, campo in enumerate(campos_seleccionados, start=col_start):
                col_letter = get_column_letter(col_idx)
                valor = parsear_fecha(dato.get(campo, "")) if campo in ["fechaArriendoInicial", "fechaArriendoFinal"] else dato.get(campo, "N/A")
                cell = sheet.cell(row=row_idx, column=col_idx, value=valor)
                cell.alignment = center_align

                # Aplicar color de relleno en filas impares
                if (row_idx - 13) % 2 == 0:  # Solo en filas impares (1°, 3°, 5°...)
                    cell.fill = fila_impar_fill

                # Formato especial para kilometraje
                if campo == "kilometraje":
                    cell.number_format = '#,##0'

                # Ajustar ancho de columna
                max_length = max(len(str(cell.value)) + 2, sheet.column_dimensions[col_letter].width)
                sheet.column_dimensions[col_letter].width = max_length

        # **Mantener la tabla**
        rango_tabla = f"B12:{get_column_letter(col_end)}{row_end}"
        tabla = Table(displayName="InformeVehicular", ref=rango_tabla)
        estilo_tabla = TableStyleInfo(
            name="TableStyleMedium4",  # Verde Oliva, Estilo de Tabla Medio 4
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,  # Activar bandas alternas predeterminadas
            showColumnStripes=False
        )
        tabla.tableStyleInfo = estilo_tabla
        sheet.add_table(tabla)

        # **Guardar el archivo**
        timestamp = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
        carpeta_salida = Path(settings.MEDIA_ROOT) / "reportes_generados_informe_vehiculo" / timestamp
        carpeta_salida.mkdir(parents=True, exist_ok=True)
        archivo_salida = carpeta_salida / f"INFORME_VEHICULAR_{timestamp}.xlsx"
        wb.save(archivo_salida)

        return archivo_salida

    except Exception as e:
        return f"Error desconocido al generar el reporte: {e}"

@login_required
@admin_required
def generar_excel(request):
    """
    Genera el informe Excel con los datos combinados de las APIs, filtrando los campos seleccionados.
    """
    if request.method != "POST":
        return JsonResponse({"success": False, "message": "Método no permitido."}, status=405)

    try:
        # Obtener los campos seleccionados desde la solicitud POST
        data = json.loads(request.body)
        campos_seleccionados = data.get("campos", [])

        if not campos_seleccionados:
            return JsonResponse({"success": False, "message": "Debes seleccionar al menos una opción."}, status=400)

        # Obtener los datos de la API
        datos_completos = obtener_datos_vehiculos()

        if not datos_completos:
            return JsonResponse({"success": False, "message": "No hay datos disponibles para procesar."}, status=404)

        # Filtrar solo los campos seleccionados
        datos_filtrados = [{campo: dato.get(campo, "N/A") for campo in campos_seleccionados} for dato in datos_completos]

        # Crear carpeta de salida
        timestamp = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
        carpeta_salida = Path(settings.MEDIA_ROOT) / "reportes_generados_informe_vehiculo" / timestamp
        carpeta_salida.mkdir(parents=True, exist_ok=True)

        # Generar el archivo Excel dinámico
        archivo_generado = generar_excel_dinamico(datos_filtrados, carpeta_salida, campos_seleccionados)

        if isinstance(archivo_generado, str):  # Si hay un error en la generación
            return JsonResponse({"success": False, "message": archivo_generado}, status=500)

        # Devolver el archivo generado como respuesta para descarga
        response = FileResponse(open(archivo_generado, 'rb'), as_attachment=True)
        response['Content-Disposition'] = f'attachment; filename="{archivo_generado.name}"'
        return response

    except Exception as e:
        return JsonResponse({"success": False, "message": str(e)}, status=500)

@login_required
@admin_required
def generar_pdf(request):
    """
    Genera un PDF con todas las columnas seleccionadas en procesar_reporte,
    guardándolo en la carpeta de reportes generados y devolviendo el archivo al usuario.
    """
    try:
        # Configurar el idioma de la fecha y hora en español
        locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')

        # Obtener datos combinados
        datos = obtener_datos_vehiculos()

        if not datos:
            return JsonResponse({"error": "No hay datos disponibles para procesar."}, status=404)

        # Ruta de la carpeta para guardar el archivo
        carpeta_salida = Path(settings.MEDIA_ROOT) / "reportes_generados_informe_vehiculo"
        carpeta_salida.mkdir(parents=True, exist_ok=True)

        # Crear el nombre del archivo con timestamp
        timestamp = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
        filename = f"INFORME VEHICULAR POR FAENA_{timestamp}.pdf"
        archivo_salida = carpeta_salida / filename

        # Crear el archivo PDF
        pdf = SimpleDocTemplate(
            str(archivo_salida), pagesize=A0, leftMargin=20, rightMargin=20, topMargin=30, bottomMargin=30
        )

        # Estilos y generación de contenido
        styles = getSampleStyleSheet()
        title_style = styles["Title"]
        title_style.fontSize = 24
        title_style.alignment = 1

        normal_style = styles["Normal"]
        normal_style.fontSize = 10

        fecha_actual = datetime.now().strftime("%A, %d de %B de %Y, %H:%M").capitalize()
        titulo = Paragraph("SICGEOATACAMA - INFORME VEHICULAR POR FAENA", title_style)
        fecha_parrafo = Paragraph(f"{fecha_actual}", normal_style)

        encabezados = [
            "FAENA", "TIPO VEHICULO", "PATENTE", "MARCA", "MODELO", "COLOR", "AÑO",
            "SITUACION", "TENENCIA", "KILOMETRAJE", "TAG", "NOMBRE PROPIETARIO",
            "RUT PROPIETARIO", "INICIO FECHA ARRIENDO", "TERMINO FECHA ARRIENDO",
            "NUMERO VIN", "NUMERO MOTOR", "NUMERO CHASIS"
        ]

        def formatear_fecha(fecha):
            if not fecha or fecha == "N/A":
                return "N/A"
            try:
                return datetime.strptime(fecha.split("T")[0], "%Y-%m-%d").strftime("%d/%m/%Y")
            except ValueError:
                return "N/A"

        filas = [encabezados]
        for dato in datos:
            filas.append([
                dato["faena"], 
                dato["tipo"], 
                dato["vehiculo"], 
                dato["marca"],
                dato["modelo"], 
                dato.get("color", "N/A"), 
                dato["ano"],
                dato["situacion"], 
                dato["tenencia"], 
                f"{dato['kilometraje']:,}",
                "Sí" if dato.get("tieneTag", False) else "No",
                dato.get("nombrePropietario", "N/A"), 
                dato.get("rutPropietario", "N/A"),
                formatear_fecha(dato.get("fechaArriendoInicial", "N/A")),
                formatear_fecha(dato.get("fechaArriendoFinal", "N/A")),
                dato.get("numeroVin", "N/A"), 
                dato.get("numeroMotor", "N/A"), 
                dato.get("numeroChasis", "N/A")
            ])

        colWidths = [max(len(str(row[col_idx])) for row in filas) * 8 for col_idx in range(len(encabezados))]
        tabla = Table(filas, repeatRows=1, colWidths=colWidths)
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))

        elementos = [fecha_parrafo, Spacer(1, 20), titulo, Spacer(1, 40), tabla]
        pdf.build(elementos)

        response = FileResponse(open(archivo_salida, 'rb'), as_attachment=True)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    except Exception as e:
        return JsonResponse({"error": f"Error al generar el PDF: {e}"}, status=500)

@login_required
@sondaje_admin_or_base_datos_or_supervisor_required
def manage_report_drilling(request): 
    lista = list(DocumentosReportesPerforaciones.objects.all().order_by('-fechacreacion'))
    faenas = list(Faena.objects.exclude(faena="SIN ASIGNAR").order_by('faena'))
    context = {
        'documentos': lista,
        'faenas': faenas,
        'sidebar': 'manage_report_drilling',
        'sidebarmain': 'system_report_drilling', 
    }
    return render(request,'pages/documentation/manage_report_drilling.html', context)

def ejecutar_scripts(fecha_inicial,fecha_final,faena_id):

    libro,error = main_reportes.ObtenerReportes(fecha_inicial,fecha_final, faena_id).run()

    return libro, error

def report_drilling_general(request):
    if request.method == 'POST':
        print(request.POST)

        fecha_inicial = request.POST.get('fechaInicial')
        fecha_final = request.POST.get('fechaFinal')
        faena_id = request.POST.get('faena')

        
        fecha_inicial = datetime.strptime(fecha_inicial, "%Y-%m-%d")
        fecha_final = datetime.strptime(fecha_final, "%Y-%m-%d")

        if fecha_inicial > fecha_final:
            return JsonResponse({'titleText': "La fecha inicial no puede ser mayor que la fecha final",'text': "Intenta nuevamente"}, status=400)

        libro, error = ejecutar_scripts(fecha_inicial,fecha_final, faena_id)  # Ejecuta el script en el archivo de Excel

        if error:
            return JsonResponse({'titleText': "No se puede crear el documento", 'text': error}, status=400)
        # Crear el nombre del archivo con timestamp
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        nombre_archivo = f"DB_Avance_Campaña_Perforación_{timestamp}.xlsx"

        # Guardar el archivo en memoria antes de pasarlo a Django
        ruta_archivo = Path(f"media/reportes_avance/{nombre_archivo}")
        ruta_archivo.parent.mkdir(parents=True, exist_ok=True)
        libro.save(str(ruta_archivo))

        documento = DocumentosReportesPerforaciones.objects.create(
            reporte= "Avance Campaña Perforación",
            creador=f"{request.user.first_name} {request.user.last_name}",
            status=True,
            archivo=f"reportes_avance/{nombre_archivo}"
        )

        print(f"Archivo guardado en: '{documento.archivo.url}'")

        return JsonResponse({
            'success': True,
            'file_id': documento.id,
            'file_name': documento.reporte,
            'file_url': documento.archivo.url
        })
    else:
        return redirect('manage_report_drilling')